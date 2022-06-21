# +++++-------Code Beginning-------+++++#
# FROM kivy.config IMPORT Config TO CONTROL APP CONFIGURATION SETTINGS.
from kivy.config import Config

# MAKE THE APP HAVE FIXED CONFIGURATION(BY PUT False) THAT'S MAKE THE USER CAN'T CHANGE ANY THING AS MEXIMIZE
# THE SCREEN FOR FULL SCREEN OR CHANGE THE SIZE, TO KEEP THE APP ORGANIZED.
Config.set('graphics', 'resizable', False)
# IMPORT (MDApp) TO CREATE THE APP
from kivymd.app import MDApp
# FROM kivy.core IMPORT Window TO BE ABLE TO CONTROL THE WINDOW SIZE.
from kivy.core.window import Window
# FROM kivy.uix.image IMPORT (AsyncImage) IF NEED TO USE IMAGE FROM WEBSITE , USE (Image) IF PHOTO ON LOCAL COMPUTER
from kivy.uix.image import AsyncImage
# FROM kivy.lang IMPORT Builder THAT'S A METHOD TO CREATE THE TEXT INPUT
from kivy.lang.builder import Builder
# FROM kivy.uix.screenmanager IMPORT ScreenManager, Screen TO CREATE APP SCREEN AND MANEGE THEM
from kivy.uix.screenmanager import ScreenManager, Screen
# FROM kivymd.uix IMPORT ALL WIDGETS(LABELS, BUTTONS,BoxLayout,...) THAT USED IN THE APP.
from kivymd.uix.label import MDLabel
from kivymd.uix.button import MDRaisedButton, MDRaisedButton, MDFloatingActionButton, MDRoundFlatButton, \
    MDFillRoundFlatButton
from kivymd.uix.gridlayout import MDGridLayout
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.textfield import MDTextField
from kivy.uix.textinput import TextInput
# USE MDDialog TO SHOW DIALOG WINDOW OF MESSAGES
from kivymd.uix.dialog import MDDialog
from kivymd.uix.list import MDCheckbox, OneLineAvatarListItem
from kivy.uix.widget import Widget

from kivy.uix.boxlayout import BoxLayout
# from kivy._event import fbind
# from kivy.core.text import
from kivymd.uix.button import MDFlatButton

from kivy.properties import StringProperty

from kivy.properties import ObjectProperty

# IMPORT MATH LIBRARY TO USE SIN(ANGLE)
import math

# USE glob() (BUILT IN FUNCTION IN PYTHON) TO SEARCH FILES INSIDE FOLDER USING FORGING NUMBER .
import glob

# **************************************#
# USE subprocess (BUILD IN FUNCTION) TO OPEN THE PROGRAM IN CIMCO
import subprocess

# **************************************#
# USE PANDAS LIBRARY TO READ HORIZONTAL DATA SHEET
import pandas as pd
# IMPORT ExcelWriter TO BA ABLE TO UPDATE THE EXCEL SHEET
from pandas import ExcelWriter

# **************************************#
# IMPORT openpyxl TO LOAD THE WORK SHEET THAT HAS OUR FORGING DATA
import openpyxl
from openpyxl import load_workbook, worksheet, workbook, writer

# IMPORT EMAIL PACKAGE
import smtplib
# IMPORT PASSWORD PACKAGE
import keyring

# **************************************#
# Date Variable , TO SET DATE OF TODAY BY DEFAULT
from datetime import date

today = date.today()
todaydate = today.strftime("%m/%d/%Y")

Screens_Builder = """
ScreenManager:
    LoginScreen:
    HomeScreen:
    PinBoreScreen:
    OldHorizontalScreen:
    NewHorizontalScreen:
    SettingScreen:
    PinBoreSettingScreen:
    OldHorizontalSettingScreen:
    AppSettingScreen:
    AddNewUserScreen:


<LoginScreen>:
    name: 'LoginScreen'
    MDLabel:
        text: 'Enter Login Information if You are One of Programming Team.'
        pos_hint: {'center_x':0.69,'center_y':0.85}
        font_size: '20sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDLabel:
        text: 'If You Are New Programmer, Ask One of Programmers to Add You.'
        pos_hint: {'center_x':0.73,'center_y':0.75}
        font_size: '16sp'
        bold: True
        italic: True
        theme_text_color: "Secondary"    
    MDLabel:
        text: ''
        pos_hint: {'center_x':0.81,'center_y':0.65}
        font_size: '15sp'
        bold: True
        italic: True
        theme_text_color: "Error"       
    MDTextField:
        id: Email
        text: self.text.lower() if self.text is not None else ''
        hint_text: "Enter Email Address"
        helper_text: "Use Email With Extension @rwbteam or @wiseco."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.45}
        size_hint_x:None
        width:300
        height:10      
    MDTextField:
        id: Password
        hint_text: "Enter Password"
        helper_text: "Use The Shared Password You Got by Email."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        password: True
        pos_hint: {'center_x': 0.50, 'center_y': 0.33}
        size_hint_x:None
        width:300
        height:10    
    MDRaisedButton:                                                                         
        text: 'LOGIN'
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        pos_hint: {'center_x':0.5,'center_y':0.23}
        on_press : 
            root.Login_Check()       
    MDLabel:
        text: 'Created by: Moemen Alatweh'
        pos_hint: {'center_x':1.32,'center_y':0.04}
        font_style: 'Caption'
        theme_text_color: "Custom"
        text_color: 175/255.0, 0/255.0, 0/255.0, 1
    MDLabel:
        text: 'malatweh@rwbteam.com'
        pos_hint: {'center_x':1.33,'center_y':0.01}
        font_style: 'Caption'
        theme_text_color: "Custom"
        text_color: 175/255.0, 0/255.0, 0/255.0, 1           


<HomeScreen>:
    name: 'HomeScreen'
    on_pre_enter: root.User_Name()
    MDLabel:
        text: 'Welcome to Wiseco Programs Maker'
        pos_hint: {'center_x':0.70,'center_y':0.8}
        font_size: '32sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDLabel:
        id: UserName
        text: ''
        halign:'center'
        valign: 'middle'
        pos_hint: {'center_y':0.7}
        font_size: '24sp'
        bold: True
        italic: True
        theme_text_color: "Error"  
    MDLabel:
        text: 'Choose Type of Program'
        pos_hint: {'center_x':0.885,'center_y':0.55}
        font_size: '18sp'
        bold: True
        italic: True
        theme_text_color: "Secondary"    
    MDRaisedButton:
        text: 'Pin Bore'
        pos_hint: {'center_x':0.5,'center_y':0.45}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'PinBoreScreen'    
    MDRaisedButton:
        text: 'Setting'
        halign:'left'
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'SettingScreen'
    MDRaisedButton:
        text: 'LOGOUT'
        pos_hint: {'center_x':0.945,'center_y':0.035}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : 
            root.Logout()         




<PinBoreScreen>:
    name: 'PinBoreScreen'
    MDLabel:
        text: 'Pin Bore Programs Maker'
        pos_hint: {'center_x':0.80,'center_y':0.8}
        font_size: '30sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDLabel:
        text: 'Choose Machine That Need Program'
        pos_hint: {'center_x':0.83,'center_y':0.6}
        font_size: '18sp'
        bold: True
        italic: True
        theme_text_color: "Secondary"    
    MDRaisedButton:
        text: 'Old Horizontal Machines (28,29,32)'
        pos_hint: {'center_x':0.5,'center_y':0.5}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'OldHorizontalScreen'
    MDRaisedButton:
        text: 'New Horizontal Machine (127)'
        pos_hint: {'center_x':0.5,'center_y':0.4}                    
        md_bg_color: 120/255, 0/255, 0/255, 1   
        font_size: "15sp"
        on_press : root.still_Work_On_It(object) 
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.5,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'    



<OldHorizontalScreen>:
    name: 'OldHorizontalScreen'
    MDLabel:
        text: 'Old Horizontal Machines (28,29,32)'
        pos_hint: {'center_x':0.73,'center_y':0.8}
        font_size: '30sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDTextField:
        id: JobNumber
        text: self.text.upper() if self.text is not None else ''
        hint_text: "Enter Job Number"
        helper_text: "Job number MUST match with number in Spec."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.60}
        size_hint_x:None
        width:300
        height:10      
    MDRaisedButton:
        text: 'Submit'
        pos_hint: {'center_x':0.5,'center_y':0.45}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : root.CreateProgram(object)    
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: 
            root.manager.current = 'PinBoreScreen'   
            root.ResetFields()     
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: 
            root.manager.current = 'HomeScreen'        
            root.ResetFields() 


<NewHorizontalScreen>:
    name: 'NewHorizontalScreen'
    MDLabel:
        text: 'New Horizontal Machine (127)'
        pos_hint: {'center_x':0.84,'center_y':0.8}
        font_size: '20sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDTextField:
        id: JobNumber
        hint_text: "Enter Job Number"
        helper_text: "Job number MUST match with number in Spec."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.60}
        size_hint_x:None
        width:300
        height:10      
    MDRaisedButton:
        text: 'Submit'
        pos_hint: {'center_x':0.5,'center_y':0.45}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : root.CreateProgram()

    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'PinBoreScreen'  
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'     


<SettingScreen>:
    name: 'SettingScreen'
    MDLabel:
        text: 'Setting'
        pos_hint: {'center_x':0.93,'center_y':0.9}
        font_size: '36sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDRaisedButton:
        text: 'Pin Bore Setting'
        pos_hint: {'center_x':0.5,'center_y':0.7}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'PinBoreSettingScreen'
    MDRaisedButton:
        text: 'Application Setting'
        pos_hint: {'center_x':0.5,'center_y':0.6}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : 
            root.Admin_Check() 
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.5,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'   


<PinBoreSettingScreen>:
    name: 'PinBoreSettingScreen'
    MDLabel:
        text: 'Pin Bore Setting'
        pos_hint: {'center_x':0.86,'center_y':0.9}
        font_size: '36sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDRaisedButton:
        text: 'Old Horizontal Setting Screen'
        pos_hint: {'center_x':0.5,'center_y':0.7}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'OldHorizontalSettingScreen'
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'SettingScreen'



<OldHorizontalSettingScreen>:
    name: 'OldHorizontalSettingScreen'
    MDLabel:
        text: 'Old Horizontal Setting Screen'
        pos_hint: {'center_x':0.77,'center_y':0.9}
        font_size: '32sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDTextField:
        id: HorizontalTemplate
        hint_text: "Horizontal Template Path"
        text: "H:/CNC_Programming/Moe A/WisecoApplications/TestStuff/HorizontalTemplate 01-05-21.MIN"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.75}
        size_hint_x:None
        width:800
        height:50       
    MDTextField:
        id: HorizontalToolList
        hint_text: "Horizontal Tool List Excel File Path"
        text: "H:/CNC_Programming/Moe A/WisecoApplications/TestStuff/HORIZONTAL_SHEETS_FOR_AUTOMATION.xlsx"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.65}
        size_hint_x:None
        width:800
        height:50          
    MDTextField:
        id: ProbePrograms
        hint_text: "Probe Programs File Path"
        text: "H:/CNCProgs/HOREBORE/Probe Programs"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.55}
        size_hint_x:None
        width:800
        height:50              
    MDTextField:
        id: NewHorizontalProgramRunningFolderPath
        hint_text: "New Horizontal Program (Running Folder) Path"
        text: "H:/CNC_Programming/Moe A/WisecoApplications/TestStuff/Horizontal"
        helper_text: "Folder that Use on Machine to Load the Program."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.45}
        size_hint_x:None
        width:800
        height:50
    MDTextField:
        id: NewHorizontalProgramOriginalFolderPath
        hint_text: "New Horizontal Program (Original Folder) Path"
        text: "H:/CNC_Programming/Moe A/WisecoApplications/TestStuff/Horizontal Original"
        helper_text: "Folder that Use as Backup For Programs."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.35}
        size_hint_x:None
        width:800
        height:50                
    MDTextField:
        id: CimcoEditorPath
        hint_text: "CIMCO Editor Path"
        text: "C:/CIMCO/CIMCOEdit8/CIMCOEdit.EXE"
        helper_text: "Path Should be Where CIMCO App Installed in User Computer."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.25}
        size_hint_x:None
        width:800
        height:50                          
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'PinBoreSettingScreen'      


<AppSettingScreen>:
    name: 'AppSettingScreen'
    MDLabel:
        text: 'Application Setting'
        pos_hint: {'center_x':0.87,'center_y':0.9}
        font_size: '32sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDTextField:
        id: EmailAddressList
        hint_text: "Email Address List File Path"
        text: "H:/CNC_Programming/Moe A/WisecoApplications/TestStuff/EMAIL_ADDRESS_LIST.xlsx"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.75}
        size_hint_x:None
        width:800
        height:50   
    MDRaisedButton:
        text: 'Add New User'
        pos_hint: {'center_x':0.5,'center_y':0.25}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'AddNewUserScreen'    
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'SettingScreen' 



<AddNewUserScreen>:
    name: 'AddNewUserScreen'
    MDLabel:
        text: 'Add New User'
        pos_hint: {'center_x':0.90,'center_y':0.9}
        font_size: '32sp'
        bold: True
        italic: True
        theme_text_color: "Primary"   
    MDTextField:
        id: UserName
        hint_text: "Enter User Name"
        helper_text: "First Name, Last Name."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.65}
        size_hint_x:None
        width:400
        height:10          
    MDTextField:
        id: NewRWBEmail
        hint_text: "Enter RWB Email Address"
        helper_text: "Email With Extension @rwbteam."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.55}
        size_hint_x:None
        width:400
        height:10
    MDTextField:
        id: NewWisecoEmail
        hint_text: "Enter Wiseco Email Address"
        helper_text: "Email With Extension @wiseco, If Not Applicable. Leave It Blank."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.45}
        size_hint_x:None
        width:400
        height:10                  
    MDRaisedButton:
        text: 'SING UP'
        pos_hint: {'center_x':0.5,'center_y':0.30}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : 
            root.Add_New_User()        
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: 
            root.manager.current = 'HomeScreen'
            root.ResetFields() 
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: 
            root.manager.current = 'AppSettingScreen'  
            root.ResetFields() 

"""


# Dialog_Builder = """
# <ConfirmationMessageTextField>:
#     MDTextField:
#         id: ConfirmationMessageTextField
#         text: "Enter the Right Value"
#
# """

#  MDTextField:
#     hint_text: "TEST"
#     helper_text_mode: "on_focus"
#     line_color_focus: 1, 1, 1, 1
#     size_hint_x:None

#  halign:'center'
#     valign: 'middle'
#     bold: True
#     italic: True
#     theme_text_color: "Secondary"

#   MDTextField:
#         id: ConfirmationMessageTextField
#         text: "Enter the Right Value"


# CREATE FUNCTION (load_horizontal_sheets_for_automation) TO LOAD HORIZONTAL TOOL LIST SHEET TO USE IT LATER FOR ALL
# HORIZONTAL PROGRAMS WHETHER 4-CYCLE, 2-CYCLE, OLD, OR NEW MACHINE
def load_horizontal_sheets_for_automation(self):
    # DEFINE VARIABLE(horizontal_tool_list_file) TO STORE horizontal_tool_list_file PATH FROM SETTING SCREEN
    # TO GET VALUE FROM ANOTHER SCREEN(SettingScreen IN THIS EXAMPLE) TO USE IT IN ANOTHER SCREEN CLASS
    # (OldHorizontalScreen IN THIS EXAMPLE) FOR LOGIC PURPOSES>>
    # >> USE (self.manager.get_screen('SettingScreen').ids["HorizontalToolList"].text)
    # self.manager.get_screen('SettingScreen'): PUT SCREEN YOU WANT TO ACCESS
    # ids["HorizontalToolList"].text : PUT WIDGET OR VALUE YOU NEED TO GET(IN THIS EXAMPLE WE GET THE TEXT OF
    # MDTextField THAT NAME "HorizontalToolList")
    horizontal_tool_list_file_path = self.manager.get_screen('OldHorizontalSettingScreen').ids[
        "HorizontalToolList"].text

    print(horizontal_tool_list_file_path)
    # TO READ EXCEL FILE THAT CONTAIN HORIZONTAL TOOL LIST AND PROBE PROGRAMS
    # (WE USE -sheet_name=None- TO READ ALL THE SHEETS)
    global horizontal_tool_list_file
    horizontal_tool_list_file = pd.read_excel(horizontal_tool_list_file_path, sheet_name=None)
    # LEAVE IT FOR TEST
    # print(horizontal_tool_list_file)
    # TO PRINT SHEETS NAME
    # print(horizontal_tool_list_file.keys())
    # TO PRINT SHEET DATA BY USING SHEET NAME('FINISH_BORE_TOOL_LIST' AS EXAMPLE)
    # print(horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'])
    # DEFINE CIMCO PATH
    global cimco_editor_path
    cimco_editor_path = self.manager.get_screen('OldHorizontalSettingScreen').ids["CimcoEditorPath"].text
    print("CIMCO EDITOR PATH: ", cimco_editor_path)

    # DEFINE LIST TO ADD ALL FINISH_BORE_TOOL_LIST TO THE LIST
    global finish_bore_tool_list
    finish_bore_tool_list = []
    # MAKE for LOOP TO READ ALL FINISH_BORE_TOOL_LIST IN SHEET['FINISH_BORE_TOOL_LIST'] IN COLUMN OF
    # 'PIN_BORE_DIAMETER', AND ADDED TO THE LIST TO USE THEM
    # horizontal_tool_list_file: THE EXCEL FILE THAT CONTAIN HORIZONTAL TOOL LIST AND PROBE PROGRAMS
    # ['FINISH_BORE_TOOL_LIST']: THE SHEET THAT'S CONTAIN TOOL LIST OF FINISH BORE
    # ['PIN_BORE_DIAMETER']: THE COLUMN THAT'S CONTAIN PIN_BORE_DIAMETER SIZES WE HAVE
    for tool in horizontal_tool_list_file['FINISH_BORE_TOOL_LIST']['PIN_BORE_DIAMETER']:
        # print(tool)
        finish_bore_tool_list.append(tool)
    print("FINISH_BORE_TOOL_LIST:", finish_bore_tool_list)

    # DEFINE LIST TO ADD ALL ROUGH_BORE_TOOL_LIST TO THE LIST
    global rough_bore_tool_list
    rough_bore_tool_list = []
    # MAKE for LOOP TO READ ALL ROUGH_BORE_TOOL_LIST IN SHEET['ROUGH_BORE_TOOL_LIST'] IN COLUMN OF 'DRILL_DIAMETER',
    # AND ADDED TO THE LIST TO USE THEM
    for tool in horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST']['DRILL_DIAMETER']:
        # print(tool)
        rough_bore_tool_list.append(tool)
    print("ROUGH_BORE_TOOL_LIST:", rough_bore_tool_list)

    # DEFINE LIST TO ADD ALL LOCK_RING_AND_CFREN_TOOL TO THE LIST
    global lock_ring_and_cfren_tool_list
    lock_ring_and_cfren_tool_list = []
    # MAKE for LOOP TO READ ALL LOCK_RING_AND_CFREN_TOOL_LIST IN SHEET['LOCK_RING_AND_CFREN_TOOL_LIST'] IN COLUMN
    # OF 'TOOL_WIDTH', AND ADDED TO THE LIST TO USE THEM
    for tool in horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST']['TOOL_WIDTH']:
        # print(tool)
        lock_ring_and_cfren_tool_list.append(tool)
    print("LOCK_RING_AND_CFREN_TOOL_LIST:", lock_ring_and_cfren_tool_list)

    # DEFINE LIST TO ADD ALL MISCELLANEOUS_TOOL_LIST TO THE LIST
    global miscellaneous_tool_list
    miscellaneous_tool_list = []
    # MAKE for LOOP TO READ ALL MISCELLANEOUS_TOOL_LIST IN SHEET['MISCELLANEOUS_TOOL_LIST'] IN COLUMN OF 'TOOL_USAGE',
    # AND ADDED TO THE LIST TO USE THEM
    for tool in horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST']['TOOL_USAGE']:
        # print(tool)
        miscellaneous_tool_list.append(tool)
    print("MISCELLANEOUS_TOOL_LIST:", miscellaneous_tool_list)

    # DEFINE LIST TO ADD ALL HORIZONTAL_SLOT_NUMBERS TO THE LIST
    global horizontal_slot_numbers
    horizontal_slot_numbers = []
    # MAKE for LOOP TO READ ALL HORIZONTAL_SLOT_NUMBERS IN SHEET['HORIZONTAL_SLOT_NUMBERS'] IN COLUMN
    # OF 'PIN_BORE_DIAMETER', AND ADDED TO THE LIST TO USE THEM
    for tool in horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS']['PIN_BORE_DIAMETER']:
        # print(tool)
        horizontal_slot_numbers.append(tool)
    print("HORIZONTAL_SLOT_NUMBERS:", horizontal_slot_numbers)

    # DEFINE LIST TO ADD ALL PROBE_PROGRAMS TO THE LIST
    # global PROBE_PROGRAMS_LIST
    # PROBE_PROGRAMS_LIST = []
    # # MAKE for LOOP TO READ ALL PROBE_PROGRAMS IN SHEET['PROBE_PROGRAMS'] IN COLUMN OF 'FORGING_NUMBER',
    # AND ADDED TO THE LIST TO USE THEM
    # for program in horizontal_tool_list_file['PROBE_PROGRAMS']['FORGING_NUMBER']:
    #     # print(tool)
    #     PROBE_PROGRAMS_LIST.append(program)
    # print("PROBE_PROGRAMS:", PROBE_PROGRAMS_LIST)

    # # define list here to store confirnation message
    # global confirmation_horizontal_email_message_list
    # confirmation_horizontal_email_message_list = []


# *****************just for now********************

# NEED TO ADD ALL 4 CYCLE PINBORE VARIABLES TO THIS FUNCTION TO AVOID REPETITION
def four_cycle_pinbore_variables():
    # NEED TO ADD LOGIC FOR SOME VARIABLES TO NOT PROCEED IF DOES NOT FIND THE VALUE FROM QANTEL DATA BASE

    global pin_hole_diameter
    pin_hole_diameter = 0.927  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global pilot_bore_depth
    pilot_bore_depth = 01.8481  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global pilot_to_pin
    pilot_to_pin = -0.25  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global x_distance_from_origin_to_pin_center
    x_distance_from_origin_to_pin_center = 0  # JUST DEFINE THAT TO USE IT ON MATH LATER
    # WE NEED LOGIC HERE TO CHECK THE OFFSET DIRECTION

    # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER   , WE JUST USE EMPTY STRING ("") WHEN CAN'T DETECT
    # THE OFFSET FROM DATA BASE
    global offset_amount
    offset_amount = ""
    # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    # , WE JUST USE EMPTY STRING ("") WHEN CAN'T DETECT THE OFFSET FROM DATA BASE
    global offset_direction  # OFFSET To0      OFFSET To180        OFFSET EACH WAY
    offset_direction = ""
    global rough_bore_speed
    rough_bore_speed = 8000  # THIS NUMBER COMES FROM HORIZONTAL TEMPLATE
    global rough_bore_feed
    rough_bore_feed = 100  # THIS NUMBER COMES FROM HORIZONTAL TEMPLATE
    global value_used_in_z_value_finish_bore_bottom
    value_used_in_z_value_finish_bore_bottom = 1.156  # THIS NUMBER COMES FROM HORIZONTAL TEMPLATE
    global ledge_tool_diameter
    ledge_tool_diameter = 0.625  # THIS NUMBER COMES FROM HORIZONTAL TEMPLATE
    # WE DEFINE IT HERE NOT BELOW BECAUSE APP CAN'T DECIDE THE STATUS IT WILL ASK FOE USER INPUT
    global ledge_cut_availability_status
    ledge_cut_availability_status = ""
    global lock_ring_cutter_width
    lock_ring_cutter_width = 0  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global lock_ring_id_spacing
    lock_ring_id_spacing = 0  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global lock_ring_diameter
    lock_ring_diameter = 0  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global lock_ring_tool_diameter
    lock_ring_tool_diameter = 0  # JUST DEFINE THAT TO USE IT ON MATH LATER , THIS NUMBER COMES FROM TOOL LIST SHEET
    global cfren_cutter_width
    cfren_cutter_width = 0.077  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global cfren_id_spacing
    cfren_id_spacing = 1.625  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global cfren_diameter
    cfren_diameter = 0.969  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global cfren_tool_diameter
    cfren_tool_diameter = 0  # JUST DEFINE THAT TO USE IT ON MATH LATER , THIS NUMBER COMES FROM TOOL LIST SHEET

    # NEED LOGIC RELATE ON QANTEL DATABASE TO KNOW IF JOB USE SAME TOOL FOR LOCKRING AND CFREN OR
    # DEFFERNT OR JUST HAVE CFREN

    global notch_angle_first_location
    notch_angle_first_location = 135  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global notch_angle_second_location
    notch_angle_second_location = 225  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global x_distance_from_origin_to_circlip_notch
    x_distance_from_origin_to_circlip_notch = 0  # JUST DEFINE THAT TO USE IT ON MATH OF NOTCH LOGIC
    global y_distance_from_origin_to_circlip_notch
    y_distance_from_origin_to_circlip_notch = 0  # JUST DEFINE THAT TO USE IT ON MATH OF NOTCH LOGIC
    global double_oil_hole_slot_id_spacing
    double_oil_hole_slot_id_spacing = 01.7823  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    # MAYBE NEED TO ADD LOGIC HERE TO DECIDE WHAT THE ID SPACING SHOULD BE, TAKE IT FROM QANTEL, OR
    # MAKE CALCULATION BY ADD LR ID SPACING + LR WIDTH IF HORIZ SLOTS STOP AT LOCKRING
    global horizontal_slots_id_spacing
    horizontal_slots_id_spacing = 2.153  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global horizontal_slots_arc_diameter
    horizontal_slots_arc_diameter = 0.375  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER

    # MAYBE THE WAY WE PROGRAM HORIZONTAL SLOTS WILL CHANGE SOON , BUT FOR NOW WE WILL USE THE WAY WE HAVE
    global i_start_horizontal_slot
    i_start_horizontal_slot = 0  # JUST DEFINE THAT TO USE IT LATER , THIS NUMBER COMES FROM TOOL LIST SHEET
    global j_start_horizontal_slot
    j_start_horizontal_slot = 0  # JUST DEFINE THAT TO USE IT LATER , THIS NUMBER COMES FROM TOOL LIST SHEET
    global horizontal_slot_radius
    horizontal_slot_radius = 0  # JUST DEFINE THAT TO USE IT LATER , THIS NUMBER COMES FROM TOOL LIST SHEET
    global ledge_counterbore_diameter
    ledge_counterbore_diameter = 0  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    # MAYBE NEED VARIABLE OF DISTANCE OF LEDGE_COUNTERBORE TO LOCKRING HERE AND IN THE HORIZ TAMPLATE

    # JUST DEFINE THAT TO USE IT ON MATH LATER
    global x_distance_375_slots_from_center_of_bore_to_center_of_horizontal_slot
    x_distance_375_slots_from_center_of_bore_to_center_of_horizontal_slot = 0
    # JUST DEFINE THAT TO USE IT ON MATH LATER
    global Y_DISTANCE_375_SLOTS_FROM_CENTER_OF_BORE_TO_CENTER_OF_HORIZONTAL_SLOT
    Y_DISTANCE_375_SLOTS_FROM_CENTER_OF_BORE_TO_CENTER_OF_HORIZONTAL_SLOT = 0

    global FORGING_NUMBER
    FORGING_NUMBER = "F6444X"  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    # NEED LOGIC TO CHECK IF WE HAVE THE FORGING IN THE FORGING DATA BASE
    # NEED LOGIC TO CHECK IF WE HAVE THE FORGE_REF_LENGTH NUMBER IN THE FORGING DATA BASE
    # DEFINE IT IN THIS WAY TO AVOID ANY CONFUSING LATER
    global FORGE_REF_LENGTH__IS_F_NUMBER_IN_EMSS
    FORGE_REF_LENGTH__IS_F_NUMBER_IN_EMSS = 2.514  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER
    global FORGING_DIAMETER__IS_FIRST_B_NUMBER_IN_EMSS
    FORGING_DIAMETER__IS_FIRST_B_NUMBER_IN_EMSS = 04.2  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER
    global FORGING_DIAMETER_OD_AT_ROUGHER__IS_SECOND_B_NUMBER_IN_EMSS
    FORGING_DIAMETER_OD_AT_ROUGHER__IS_SECOND_B_NUMBER_IN_EMSS = 4.12  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER
    global FORGING_OUTSIDE_BOSS_SPACING__IS_U_NUMBER_IN_EMSS
    FORGING_OUTSIDE_BOSS_SPACING__IS_U_NUMBER_IN_EMSS = 2.38  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER
    global FORGING_INSIDE_BOSS_SPACING__IS_J_NUMBER_IN_EMSS
    FORGING_INSIDE_BOSS_SPACING__IS_J_NUMBER_IN_EMSS = 0.800  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER

    # define list here to store confirnation message
    global confirmation_horizontal_email_message_list
    confirmation_horizontal_email_message_list = []


# MAYBE NEED TO MOVE THESE FUNCTIONS ABOVE (FOUR_CYCLE_PINBORE_VARIABLES) FUNCTION
def Horizontal_Program_Has_been_Created_Successfully_Original_Folder(self):
    print("Horizontal_Program_Has_been_Created_Successfully_Original_Folder " + " is called")
    # JUST FOR TESTING
    # SUCCESSFUL_MESSAGES_LIST = [self.ids["JobNumber"].text + ": Horizontal Program Has been Created Successfully."]
    # print('\n'.join(SUCCESSFUL_MESSAGES_LIST))
    # print()
    # print('\n'.join(HorizontalProgramLines))
    # WE NEED TO BACK HERE TO ADD if STATEMENT TO CHECK IF JOB HAS (OFFSET EACH WAY) WE NEED TO CREATE 2 PROGRAMS , ONE (TO0), AND ONE (TO180)
    # NEW_HORIZONTAL_PROGRAM = ("P" + NEW_PROGRAM_NUMBER + ".MIN")
    # print(NEW_HORIZONTAL_PROGRAM)
    # global NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH
    # NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH = self.manager.get_screen('OldHorizontalSettingScreen').ids["NewHorizontalProgramRunningFolderPath"].text
    # print(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH)
    global NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH
    NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH = self.manager.get_screen('OldHorizontalSettingScreen').ids
    ["NewHorizontalProgramOriginalFolderPath"].text


    # print(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH)
    # **************
    # TO CREATE THE NEW FILE ON THE PATH YOU WANT
    # open() WITH "x" IT will create a file, returns an error if the file exist(THAT WHY WE USE try/except)
    # IF IT DOES NOT EXIST IT WILL CREATE THE NEW FILE , IF IT IS EXIST IT WILL GO TO except BLOCK AND CHECK IF NEED TO SAVE OVER THE EXISTING FILE
    # **********************NEED TO BACK TO USE try & exept *************************

    # try:
    # TRY TO CREATE THE HORIZONTAL PROGRAM ON ORIGINAL FOLDER
    # DEFINE VARIABLES ON THIS FUNCTION AS global TO BE ABLE USE THEM OUT SIDE THE FUNCTION
    global NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER
    global NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0
    global NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180

    global HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER
    HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER = HorizontalProgramLines.copy()

    global HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER
    HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER = HorizontalProgramLines.copy()

    # # USED global TO BE ABLE USE THE LIST OUTSIDE THE FUNCTION
    # global result_of_existing_programs
    # # LIST TO STORE THE FOUND JOBS WITH WHOLE PATH INCLUDING FILE NAME
    # result_of_existing_programs = []
    # for file in glob.glob(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + '\*\*' + NEW_PROGRAM_NUMBER + '*'):
    #     # TO APPEND(ADD) EACH PROGRAM THAT FOUND TO THE RESULT LIST
    #     result_of_existing_programs.append(file)
    #     print("result_of_existing_programs: ",result_of_existing_programs)


    try:
        print("try original is called")

        # global result_of_existing_programs
        # # LIST TO STORE THE FOUND JOBS WITH WHOLE PATH INCLUDING FILE NAME
        # result_of_existing_programs = []
        # for file in glob.glob(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + '*\*' + "" + '*'):
        #     # TO APPEND(ADD) EACH PROGRAM THAT FOUND TO THE RESULT LIST
        #     result_of_existing_programs.append(file)
        # print("result_of_existing_programs: ", result_of_existing_programs)

        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):

            # TO 0
            # global HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER
            # WE USE COPY METHOD TO COPY (HorizontalProgramLines) THAT'S CONTAIN ORIGINAL LINES BEFORE MAKE 2 SEPARATE PROGRAMS
            # WE MAKE NEW LIST TO BE ABLE TO MODIFY IT WITHOUT ADJUST THE ORIGINAL LIST(HorizontalProgramLines) BECAUSE WE WANT TO USE IT LATER
            # HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER = HorizontalProgramLines.copy()
            print("ORIGINAL_FOLDER_TO0")
            print(HorizontalProgramLines)

            NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0 = \
                (NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO0.MIN")
            HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER[0] = (
                    '(PART ' + NEW_PROGRAM_NUMBER + " TO 0" + ' -- ' + todaydate + ' <SYS>' + ')')

            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0 = open(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0, "x")
            # CREATE SEPARETE LIST TO AVOID MIXED UP BETWEEN RUNNING AND ORIGINAL FOLDER
            # global HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER
            # HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER = HorizontalProgramLines
            # print("ORIGINAL_FOLDER_TO0")
            # print(HorizontalProgramLines)

            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0.write(
                '\n'.join(HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER))
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0.close()

            print("ORIGINAL_FOLDER_TO0   AFTER EXUTAION")
            print(HorizontalProgramLines)

            # TO 180
            # global HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER
            # HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER = HorizontalProgramLines.copy()
            print("ORIGINAL_FOLDER_TO180")
            print(HorizontalProgramLines)

            NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180 = (
                    NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO180.MIN")
            HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER[0] = (
                    '(PART ' + NEW_PROGRAM_NUMBER + " TO 180" + ' -- ' + todaydate + ' <SYS>' + ')')
            HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER[VC155_VARIABLE_INDEX] = (
                    'VC155=' + format((-1) * offset_amount) + '  (Offset)')
            # print("OFFSET_AMOUNT for original folder", OFFSET_AMOUNT)
            HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER[VC173_VARIABLE_INDEX] = (
                    'VC173=' + format(float(y_distance_from_origin_to_circlip_notch) - (2 * offset_amount),
                                      '.4f') + '  (yCirclipNotch)')

            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180 = open(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180,
                                                                       "x")
            # CREATE SEPARETE LIST TO AVOID MIXED UP BETWEEN RUNNING AND ORIGINAL FOLDER
            # global HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER
            # HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER = HorizontalProgramLines
            # print("ORIGINAL_FOLDER_TO180")
            # print(HorizontalProgramLines)

            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180.write(
                '\n'.join(HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER))
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180.close()

        else:
            # if ((NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO0.MIN") or
            #         (NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO180.MIN") in result_of_existing_programs):
            #     print("job saved as To0 & To180")
            #     Close_Button = MDRaisedButton(text='Close',on_release=self.Close_Old_Horizontal_Dialog_Original_Folder)  # Close_Old_Horizontal_Dialog_      '[color=ffffff] Folder.[/color]'
            #     self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=cccc00]Warning Message[/color]', text=(
            #                 '[b][i][u][color=ffffff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] saved as the job has Offset Each Way in [/color]' + '[color=ffff00]ORIGINAL[/color]' + '[color=ffffff] Folder.[/color]')
            #                                                   , size_hint=(0.7, 1.0), buttons=[Close_Button],auto_dismiss=False)
            #     # TO OPEN THE DIALOG WINDOW
            #     self.Old_Horizontal_Message_Dialog.open()
            NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER = (
                    NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + ".MIN")
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER = open(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER, "x")
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER.write('\n'.join(HorizontalProgramLines))
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER.close()

        # USED global TO BE ABLE USE THE LIST OUTSIDE THE FUNCTION
        # global result_of_existing_programs
        # # LIST TO STORE THE FOUND JOBS WITH WHOLE PATH INCLUDING FILE NAME
        # result_of_existing_programs = []
        # for file in glob.glob(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + '*\*' + NEW_PROGRAM_NUMBER + '*'):
        #     # TO APPEND(ADD) EACH PROGRAM THAT FOUND TO THE RESULT LIST
        #     result_of_existing_programs.append(file)
        #     print("result_of_existing_programs: ", result_of_existing_programs)
        #
        # print("result_of_existing_programs: ", result_of_existing_programs)

        # if ((NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO0.MIN") or
        #         (NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO180.MIN") in result_of_existing_programs):
        #     print("job saved as To0 & To180")

        # just for now              WARNING_MESSAGES_LIST
        if (confirmation_horizontal_email_message_list != []):
            horizontal_email_message_list.append("\n".join(confirmation_horizontal_email_message_list) + "\n")
        SUCCESSFUL_MESSAGES_LIST = ["Program has been created successfully in ORIGINAL Folder." + "\n"]
        # maybe need logic to send this message to trello in case there is warnning make this program
        horizontal_email_message_list.append("\n".join(SUCCESSFUL_MESSAGES_LIST))

        Close_Button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_window_of_original_folder,
                                      font_size=16)  # Close_Old_Horizontal_Dialog_      '[color=ffffff] Folder.[/color]'
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
                '[b][i][u][color=0099ff]' + self.ids[
            "JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] Program Has been Created Successfully in [/color]' + '[color=ffff00]ORIGINAL[/color]' + '[color=ffffff] Folder.[/color]')
                                                      , size_hint=(0.7, 1.0), buttons=[Close_Button],
                                                      auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        print()
        # print('\n'.join(HorizontalProgramLines))
    except(FileExistsError):
        print("exept original is called")
        # print(" Program is Exist in Original Folder")         ('[b][i][u][color=ffffff]'+self.ids["JobNumber"].text+'[/color][/u][/i][/b]' + '[color=ffffff] Program already Exists in [/color]'+'[color=ffff00]ORIGINAL[/color]'+'[color=ffffff] Do you want to replace it ?[/color]')
        Yes_Button = MDRaisedButton(text='Yes', on_release=self.replace_existing_old_horizontal_machine_program_in_original_folder,
                                    font_size=16)  # , on_press=self.Close_Old_Horizontal_Dialog_Original_Folder
        No_Button = MDRaisedButton(text='No', on_release=self.close_old_horizontal_window_of_original_folder, font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=0066ff]Confirmation Message[/color]',
                                                      text=('[b][i][u][color=0099ff]' + self.ids[
                                                          "JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] Program already Exists in [/color]' + '[color=ffff00]ORIGINAL[/color]' + '[color=ffffff] Folder.[/color]' + '\n' + '[color=ffffff]Do you want to replace it ?[/color]')
                                                      , size_hint=(0.7, 1.0), buttons=[Yes_Button, No_Button],
                                                      auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()

        # MAYBE WE DO NOT NEED IT
        # TO CLEAR THE LIST TO BE EMPTY IF ALREADY THE PROGRAM EXIST
        SUCCESSFUL_MESSAGES_LIST = []
        # maybe need logic to send this message to trello in case there is warnning make this program
        horizontal_email_message_list.append("\n".join(SUCCESSFUL_MESSAGES_LIST))

        # def SAVE_OVER_EXISTING_PROGRAM_ORIGINAL_FOLDER(self):
        #     NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER = (NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + ".MIN")
        #     CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER = open(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER, "w")
        #     CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER.write('\n'.join(HorizontalProgramLines))
        #     CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER.close()
        #     Close_Button = MDRaisedButton(text='Close', on_release=self.Close_Old_Horizontal_Dialog)
        #     self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
        #             self.ids["JobNumber"].text + ' Program Has been saved over Successfully in ORIGINAL Folder.'),size_hint=(0.7, 1.0), buttons=[Close_Button])
        #     # TO OPEN THE DIALOG WINDOW
        #     self.Old_Horizontal_Message_Dialog.open()
        #     print()
        #     print('\n'.join(HorizontalProgramLines))


# TRY TO CREATE THE HORIZONTAL PROGRAM ON RUNNING FOLDER
def Horizontal_Program_Has_been_Created_Successfully_Running_Folder(self):
    print("Horizontal_Program_Has_been_Created_Successfully_Running_Folder" + " is called")
    global NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH
    NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH = self.manager.get_screen('OldHorizontalSettingScreen').ids[
        "NewHorizontalProgramRunningFolderPath"].text
    global NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER
    global NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0
    global NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180

    global HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER
    HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER = HorizontalProgramLines.copy()

    global HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER
    HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER = HorizontalProgramLines.copy()

    try:
        print("try running is called")
        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):
            # TO 0
            # global HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER
            # HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER = HorizontalProgramLines.copy()
            print("RUNNING_FOLDER_TO0")
            print(HorizontalProgramLines)
            # TO 0
            NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0 = (
                    NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO0.MIN")
            HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER[0] = (
                    '(PART ' + NEW_PROGRAM_NUMBER + " TO 0" + ' -- ' + todaydate + ' <SYS>' + ')')

            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0 = open(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0, "x")
            # global HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER
            # HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER = HorizontalProgramLines            # need to fix here
            # print("RUNNING_FOLDER_TO0")
            # print(HorizontalProgramLines)

            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0.write(
                '\n'.join(HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER))
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0.close()

            # TO 180
            # global HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER
            # HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER = HorizontalProgramLines.copy()
            print("RUNNING_FOLDER_TO180")
            print(HorizontalProgramLines)
            # TO 180
            NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180 = (
                    NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO180.MIN")
            HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER[0] = (
                    '(PART ' + NEW_PROGRAM_NUMBER + " TO 180" + ' -- ' + todaydate + ' <SYS>' + ')')
            HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER[VC155_VARIABLE_INDEX] = (
                    'VC155=' + format((-1) * offset_amount) + '  (Offset)')
            HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER[VC173_VARIABLE_INDEX] = (
                    'VC173=' + format(float(y_distance_from_origin_to_circlip_notch) - (2 * offset_amount),
                                      '.4f') + '  (yCirclipNotch)')

            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180 = open(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180, "x")
            # global HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER
            # HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER = HorizontalProgramLines
            # print("RUNNING_FOLDER_TO180")
            # print(HorizontalProgramLines)

            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180.write(
                '\n'.join(HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER))
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180.close()

        else:
            NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER = (
                    NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + ".MIN")
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER = open(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER, "x")
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER.write('\n'.join(HorizontalProgramLines))
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER.close()
        # '[b][i][u][color=ffffff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] Program Has been Created Successfully in [/color]'+'[color=ffff00]RUNNING[/color]'+'[color=ffffff]Please Double Check and Report any Problem on Trello.[/color]'

        # JUST FOR NOW
        SUCCESSFUL_MESSAGES_LIST = ["Program has been created successfully in RUNNING Folder."]
        horizontal_email_message_list.append("\n".join(SUCCESSFUL_MESSAGES_LIST))

        Close_Button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_window_of_running_folder,
                                      font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
                '[b][i][u][color=0099ff]' + self.ids[
            "JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] Program Has been Created Successfully in [/color]' + '[color=33cc33]RUNNING[/color]' + '[color=ffffff] Folder.[/color]' + '\n' + '[color=ffffff]After closing this window, the program will open on CIMCO Editor.[/color]' + '\n' + '[color=ffffff]Please Double Check and Report any Problem on Trello.[/color]')
                                                      , size_hint=(0.7, 1.0), buttons=[Close_Button],
                                                      auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        print()
        # print('\n'.join(HorizontalProgramLines))
    except(FileExistsError):
        print("exept running is called")
        # print(" Program is Exist in Running Folder")
        Yes_Button = MDRaisedButton(text='Yes', on_release=self.replace_existing_old_horizontal_machine_program_in_running_folder,
                                    font_size=16)  # , on_release=self.Close_Old_Horizontal_Dialog
        No_Button = MDRaisedButton(text='No', on_release=self.close_old_horizontal_screen_window, font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=0066ff]Confirmation Message[/color]',
                                                      text=('[b][i][u][color=0099ff]' + self.ids[
                                                          "JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] Program already Exists in [/color]' + '[color=33cc33]Running[/color]' + '[color=ffffff] Folder.[/color]' + '\n' + '[color=ffffff]Do you want to replace it ?[/color]')
                                                      , size_hint=(0.7, 1.0), buttons=[Yes_Button, No_Button],
                                                      auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        # MAYBE WE DO NOT NEED IT
        # TO CLEAR THE LIST TO BE EMPTY IF ALREADY THE PROGRAM IS EXIST
        SUCCESSFUL_MESSAGES_LIST = []
        horizontal_email_message_list.append("\n".join(SUCCESSFUL_MESSAGES_LIST))

        # def Save_Over_Existing_Program(self):
        #     NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER = (NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + ".MIN")
        #     CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER = open(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER, "w")
        #     CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER.write('\n'.join(HorizontalProgramLines))
        #     CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER.close()
        #     Close_Button = MDRaisedButton(text='Close', on_release=self.Close_Old_Horizontal_Dialog)
        #     self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
        #             self.ids["JobNumber"].text + ' Program Has been saved over Successfully in RUNNING Folder.'+ '\n' + 'Shortly, It will open on CIMCO ' + '\n' + 'Please Double Check and Report any Problem on Trello'),
        #                                                   size_hint=(0.7, 1.0), buttons=[Close_Button])
        #     # TO OPEN THE DIALOG WINDOW
        #     self.Old_Horizontal_Message_Dialog.open()
        #     print()
        #     print('\n'.join(HorizontalProgramLines))


# def Reset_Everything(self):
#     self.Old_Horizontal_Message_Dialog.dismiss()


#     SUCCESSFUL_MESSAGES_LIST.append(self.ids["JobNumber"].text)
#     SUCCESSFUL_MESSAGES_LIST.append("Horizontal Program Has been Created Successfully.")
#     # #MAYBE WE WILL USE MESSAGES BELOW
#     # SUCCESSFUL_MESSAGES_LIST.append("Running Program : " + NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER)
#     # SUCCESSFUL_MESSAGES_LIST.append("Original Program : " + NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER)
#     Close_Button = MDRaisedButton(text='Close', on_release= self.Close_Old_Horizontal_Dialog)
#     self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]',text=('\n'.join(SUCCESSFUL_MESSAGES_LIST)), size_hint=(0.7, 1.0) , buttons=[Close_Button])
#     # TO OPEN THE DIALOG WINDOW
#     self.Old_Horizontal_Message_Dialog.open()
#     print('\n'.join(SUCCESSFUL_MESSAGES_LIST))
#     print()
#     print('\n'.join(HorizontalProgramLines))
#
# except:
#     print(" Program is Exist ")
#     Close_Button = MDRaisedButton(text='Close', on_release=self.Close_Old_Horizontal_Dialog)
#     self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=ffffff]Message[/color]',text=(self.ids["JobNumber"].text  +' is Exist on Horizontal Folders'), size_hint=(0.7, 1.0),buttons=[Close_Button])
#     # TO OPEN THE DIALOG WINDOW
#     self.Old_Horizontal_Message_Dialog.open()

#     # JUST FOR NOW
#     pass
# Close_Button = MDRaisedButton(text='Close', on_release=self.CloseDialog)
# self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]',text=('\n'.join(SUCCESSFUL_MESSAGES_LIST)), buttons=[Close_Button])
# # TO OPEN THE DIALOG WINDOW
# self.Old_Horizontal_Message_Dialog.open()
#               ,size_hint=(0.7, 1.0)           ,size_hint=self.size    ,size_hint=(0.95, 1.0)          self.texture_size

# *****************STILL WORK HERE****************
#   MAYBE NEED TO ADD ARGUMENT FOR THE FUNCTION FOR THE TEXT VARIABLE THAT'S WILL SEND TO (text) on (Confirmation_MDLabel)
def Need_Confermation_to_Create_Horizontal_Program(self, title, Sub_Function, Dialog_Type, Content):
    print("Need_Confermation_to_Create_Horizontal_Program" + " is called")
    print('\n'.join(CONFIRMATION_MESSAGES_LIST))
    # still need to add action to do
    Enter_Button = MDRaisedButton(text='Enter', on_press=Sub_Function, on_release=self.create_program_for_old_horizontal_machine,
                                  font_size=16)  # on_release=self.CreateProgram
    Close_Button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_screen_window, font_size=16)
    # Dialog_BoxLayout = MDBoxLayout(orientation='vertical')
    #     # NEED TO BAKE HERE AND MAKE (text) ARGUMENT TAKE VARIABLE TEXT RELATED ON THE SITUATION
    # Confirmation_MDLabel = MDLabel(text=('\n'.join(CONFIRMATION_MESSAGES_LIST)) , theme_text_color='Custom', text_color=(1, 1, 1, 1))
    # Dialog_BoxLayout.add_widget(Confirmation_MDLabel)
    # # global Confirmation_MDTextField
    # self.Confirmation_MDTextField = TextInput(hint_text= "Enter Value" , multiline=False , background_color= [60/255.0, 60/255.0, 60/255.0, 1], foreground_color= [1,1,1,1])  #,hint_text= "color_mode = 'accent'"   ,mode="fill",fill_color= (1, 1, 1, 1)   line_color_focus
    # Dialog_BoxLayout.add_widget(self.Confirmation_MDTextField)
    # self.Builder_Dialog = Builder.load_string(Dialog_Builder)
    # ,items=Content
    if (Dialog_Type == "custom"):
        self.Old_Horizontal_Message_Dialog = MDDialog(title=title, type=Dialog_Type, content_cls=Content,
                                                      size_hint=(0.7, 1.0), buttons=[Enter_Button, Close_Button],
                                                      auto_dismiss=False)
    elif (Dialog_Type == "confirmation"):
        self.Old_Horizontal_Message_Dialog = MDDialog(title=title, type=Dialog_Type, items=Content,
                                                      size_hint=(0.7, 1.0), buttons=[Enter_Button, Close_Button],
                                                      auto_dismiss=False)
    # TO OPEN THE DIALOG WINDOW             '[color=248f24]Successful Message[/color]'           ('\n'.join(FAILED_MESSAGES_LIST))          size_hint=(0.7, 1.0),     ,content_cls=Confirmation_MDTextField
    self.Old_Horizontal_Message_Dialog.open()


# ============NEED TO CREATE WARNING FUNCTION=================
def horizontal_Program_needs_Attention(self):
    print("horizontal_Program_needs_Attention" + " is called")
    # print('\n'.join(FAILED_MESSAGES_LIST))
    Close_Button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_screen_window, font_size=16)
    self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=cccc00]Warning Message[/color]', text=(
            '[color=ffffff]' + '\n'.join(WARNING_MESSAGES_LIST) + '[/color]'), size_hint=(0.7, 1.0),
                                                  buttons=[Close_Button], auto_dismiss=False)
    # TO OPEN THE DIALOG WINDOW             '[color=248f24]Successful Message[/color]'           ('\n'.join(FAILED_MESSAGES_LIST))
    self.Old_Horizontal_Message_Dialog.open()


def Failed_to_Create_Horizontal_Program(self):  # '[color=ffffff]' + '\n'.join(FAILED_MESSAGES_LIST) + '[/color]'
    print("Failed_to_Create_Horizontal_Program" + " is called")
    # print('\n'.join(FAILED_MESSAGES_LIST))
    Close_Button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_screen_window, font_size=16)
    self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=990000]Warning Message[/color]', text=(
            '[color=ffffff]' + '\n'.join(FAILED_MESSAGES_LIST) + '[/color]'), size_hint=(0.7, 1.0),
                                                  buttons=[Close_Button], auto_dismiss=False)
    # TO OPEN THE DIALOG WINDOW             '[color=248f24]Successful Message[/color]'           ('\n'.join(FAILED_MESSAGES_LIST))
    self.Old_Horizontal_Message_Dialog.open()


# still need work
def send_email(self):
    try:
        service_app = "outlook.office365.com"
        # SERVER NAME
        smtp_server = "smtp.office365.com"
        port = 587
        sender_email = user_emai_address
        # USING KEYRING PACKAGE TO GET PASSWORD FROM Windows Credential Manager WHILE THEY ARE SAVING IN USER COMPUTER
        sender_password = keyring.get_password(service_app, sender_email)
        # just for now
        recipient_email = "moemenatweh@hotmail.com"
        # creates SMTP session
        email_Server = smtplib.SMTP(smtp_server, port)
        # TLS for security
        email_Server.starttls()

        # authentication
        # compiler gives an error for wrong credential.
        email_Server.login(sender_email, sender_password)

        # message to be sent
        email_message = f"""From: Alatweh Moemen <malatweh@rwbteam.com>
To: <moemenatweh@hotmail.com>
Subject: Testing Email by python    


{"".join(horizontal_email_message_list)}"""

        email_Server.sendmail(sender_email, recipient_email, email_message)

        # terminating the session
        email_Server.quit()

    except Exception as e:
        FAILED_MESSAGES_LIST.append(
            "Failed to send Email to Trello board." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]' + str(
                e) + '[/color]' + "\n" + "Double Check Network and Login Authentication.")
        # horizontal_email_message_list.append("\n".join(FAILED_MESSAGES_LIST))
        Failed_to_Create_Horizontal_Program(self)


# def CloseDialog(self, obj):
#     self.Message_Dialog.dismiss()
#     # TO RESET JOB NUMBER FIELD TO START OVER
#     self.ids["JobNumber"].text = ""

class LoginScreen(Screen):
    def Login_Check(self):
        global user_emai_address
        user_emai_address = self.ids["Email"].text
        print("USER: ", user_emai_address)
        global EMAIL_ADDRESS_LIST_FILE_PATH
        EMAIL_ADDRESS_LIST_FILE_PATH = self.manager.get_screen('AppSettingScreen').ids["EmailAddressList"].text
        ##print(EMAIL_ADDRESS_LIST_FILE_PATH)
        # TO READ EXCEL FILE THAT CONTAIN EMAIL ADDRESS OF USERS (WE USE -sheet_name=None- TO READ ALL THE SHEETS)
        global EMAIL_ADDRESS_LIST_FILE
        EMAIL_ADDRESS_LIST_FILE = pd.read_excel(EMAIL_ADDRESS_LIST_FILE_PATH, sheet_name=None)
        ##print(EMAIL_ADDRESS_LIST_FILE)

        # ADD USERS NAME TO THE list
        global USERS_NAME_LIST
        USERS_NAME_LIST = []
        for user in EMAIL_ADDRESS_LIST_FILE['Email']['Users']:
            USERS_NAME_LIST.append(user)
        # print("USERS_NAME_LIST:", USERS_NAME_LIST)

        # ADD WISECO_EMAIL_ADDRESS TO THE list
        global WISECO_EMAIL_ADDRESS_LIST
        WISECO_EMAIL_ADDRESS_LIST = []
        for user in EMAIL_ADDRESS_LIST_FILE['Email']['Wiseco Email Address']:
            WISECO_EMAIL_ADDRESS_LIST.append(user)
        # print("WISECO_USERS_LIST:", WISECO_EMAIL_ADDRESS_LIST)

        # ADD RWB_EMAIL_ADDRESS TO THE list
        global RWB_EMAIL_ADDRESS_LIST
        RWB_EMAIL_ADDRESS_LIST = []
        for user in EMAIL_ADDRESS_LIST_FILE['Email']['RWB Email Address']:
            RWB_EMAIL_ADDRESS_LIST.append(user)
        # print("RWB_USERS_LIST:", RWB_EMAIL_ADDRESS_LIST)

        # ADD PASSWORD TO THE list
        global PASSWORD_LIST
        PASSWORD_LIST = []
        for password in EMAIL_ADDRESS_LIST_FILE['Email']['Pass']:
            PASSWORD_LIST.append(password)

        # TO CHECK LOGIN INFORMATION
        if (((self.ids["Email"].text in RWB_EMAIL_ADDRESS_LIST and self.ids[
            "Email"].text != 'malatweh@rwbteam.com') or (self.ids["Email"].text in WISECO_EMAIL_ADDRESS_LIST))
                and ((self.ids["Password"].text) in PASSWORD_LIST)):
            print("LOGIN SUCCESS")
            # THAT'S HOW YOU CAN GO TO OTHER SCREENS INSIDE THE CLASS
            self.manager.current = 'HomeScreen'

        elif (self.ids["Email"].text == 'malatweh@rwbteam.com' and (
                self.ids["Password"].text == 'moe' + PASSWORD_LIST[6])):
            print("ADMIN LOGIN SUCCESS")
            # THAT'S HOW YOU CAN GO TO OTHER SCREENS INSIDE THE CLASS
            self.manager.current = 'HomeScreen'
        else:
            print(
                "Wrong Email or Password, Try Again")  # ,text_color = (220/255, 0/255, 0/255, 1)             MDRaisedButton    font_size = 16          md_bg_color: 120/255, 0/255, 0/255, 1
            Close_Button = MDRaisedButton(text='Close', on_release=self.CloseDialog, font_size=16)
            self.Warning_Dialog = MDDialog(title='[color=990000]Warning Message[/color]',
                                           text=('[color=ffffff]Wrong Email or Password, Try Again[/color]'),
                                           size_hint=(0.7, 1.0), buttons=[Close_Button], auto_dismiss=False)
            # TO OPEN THE DIALOG WINDOW
            self.Warning_Dialog.open()

    def CloseDialog(self, obj):
        self.Warning_Dialog.dismiss()


class HomeScreen(Screen):
    # THAT'S HOW TO CALL A FUNCTION JUST BY ENTERING THR SCREEN, NEED TO ADD IT ON SCREEN ABOVE AS WELL
    def on_pre_enter(self):
        # YOU PUT THE FUNCTION YOU WANT TO CALL TO DO SOME ACTION
        self.User_Name()

    def User_Name(self):
        # print(USERS_NAME_LIST)
        # global UserName
        global user_name
        # self.ids["UserName"].text = USERS_NAME_LIST[3]
        # CHECK IF EMAIL THAT ENTERED IS EXIST OF ONE OF LISTS (RWB_EMAIL or WISECO_EMAIL), THEN FIND index OF EMAIL TO USE IT IN (USERS_NAME_LIST) list TO PRINT IT ON SCREEN
        if (self.manager.get_screen('LoginScreen').ids["Email"].text in RWB_EMAIL_ADDRESS_LIST):
            # print(self.manager.get_screen('LoginScreen').ids["Email"].text)
            # TO FIND INDEX OF USER IN EXCEL FILE
            USER_INDEX = RWB_EMAIL_ADDRESS_LIST.index(self.manager.get_screen('LoginScreen').ids["Email"].text)
            # print(USER_INDEX)
            self.ids["UserName"].text = USERS_NAME_LIST[USER_INDEX]
            user_name = self.ids["UserName"].text
        elif (self.manager.get_screen('LoginScreen').ids["Email"].text in WISECO_EMAIL_ADDRESS_LIST):
            # print(self.manager.get_screen('LoginScreen').ids["Email"].text)
            # TO FIND INDEX OF USER IN EXCEL FILE
            USER_INDEX = WISECO_EMAIL_ADDRESS_LIST.index(self.manager.get_screen('LoginScreen').ids["Email"].text)
            # print(USER_INDEX)
            self.ids["UserName"].text = (USERS_NAME_LIST[USER_INDEX])
            user_name = self.ids["UserName"].text

    def Logout(self):
        # THAT'S HOW YOU CAN GO TO OTHER SCREENS INSIDE THE CLASS
        self.manager.current = 'LoginScreen'
        # TO RESET LOGIN FIELDS
        self.manager.get_screen('LoginScreen').ids["Email"].text = ""
        self.manager.get_screen('LoginScreen').ids["Password"].text = ""


class PinBoreScreen(Screen):
    def still_Work_On_It(self, obj):
        Close_Button = MDRaisedButton(text='Close', on_release=self.Close_PinBore_Dialog, font_size=16)
        self.PinBore_Message_Dialog = MDDialog(title='[color=cccc00]Warning Message[/color]', text=(
            '[color=ffffff]Still Work On It, Thanks for your Patience. [/color]'), size_hint=(0.7, 1.0),
                                               buttons=[Close_Button], auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW             '[color=248f24]Successful Message[/color]'           ('\n'.join(FAILED_MESSAGES_LIST))
        self.PinBore_Message_Dialog.open()

    def Close_PinBore_Dialog(self, obj):
        print("Close_PinBore_Dialog" + " is called")
        # TO CLOSE THE DIALOGE
        self.PinBore_Message_Dialog.dismiss()


# CREATE THE CLASS FOE ITEM THAS'T DISPLAY IN CONFIRMATION DIALOG
class Item(OneLineAvatarListItem):
    divider = None

    # def set_item(self):
    #     print("choose item work fine")

    # if (OneLineAvatarListItem.on_release):
    #     print()
    #     print("choose item work fine")
    #     print()

    # def set_icon(self, instance_check):
    #     instance_check.active = True
    #     check_list = instance_check.get_widgets(instance_check.items)
    #     for check in check_list:
    #         if check != instance_check:
    #             check.active = False
    # source = StringProperty(fbind)


class OldHorizontalScreen(Screen):
    # # DEFINE (SUCCESSFUL_MESSAGES_LIST) list TO ADD IT TO DIALOG TO SHOW THE MESSAGE
    # global SUCCESSFUL_MESSAGES_LIST
    # SUCCESSFUL_MESSAGES_LIST = []
    #
    # # DEFINE (CONFIRMATION_MESSAGES_LIST) list TO ADD IT TO DIALOG TO SHOW THE MESSAGE
    # global CONFIRMATION_MESSAGES_LIST
    # CONFIRMATION_MESSAGES_LIST = []
    #
    # # DEFINE (FAILED_MESSAGES_LIST) list TO ADD IT TO DIALOG TO SHOW THE MESSAGE
    # global FAILED_MESSAGES_LIST
    # FAILED_MESSAGES_LIST = []

    # still work here
    # def EnterOffsetValue(self,obj):
    #     # global OFFSET_AMOUNT
    #     OFFSET_AMOUNT = self.Confirmation_MDTextField.text
    #     CONFIRMATION_MESSAGES_LIST = []
    #     print("offset value after ", OFFSET_AMOUNT)
    #
    # OldHorizontalScreen().EnterOffsetValue()

    # CALL (FOUR_CYCLE_PINBORE_VARIABLES) FUNCTION TO SET THE VARIABLES
    four_cycle_pinbore_variables()

    def CreateProgram(self, obj):
        print("CreateProgram" + " is called")

        print("offset value before OR AFTER ", offset_amount)

        # DEFINE (SUCCESSFUL_MESSAGES_LIST) list TO ADD IT TO DIALOG TO SHOW THE MESSAGE
        global SUCCESSFUL_MESSAGES_LIST
        SUCCESSFUL_MESSAGES_LIST = []

        # DEFINE (CONFIRMATION_MESSAGES_LIST) list TO ADD IT TO DIALOG TO SHOW THE MESSAGE
        global CONFIRMATION_MESSAGES_LIST
        CONFIRMATION_MESSAGES_LIST = []

        # DEFINE (FAILED_MESSAGES_LIST) list TO ADD IT TO DIALOG TO SHOW THE MESSAGE
        global FAILED_MESSAGES_LIST
        FAILED_MESSAGES_LIST = []

        # DEFINE (WARNING_MESSAGES_LIST) list TO ADD IT TO DIALOG TO SHOW THE MESSAGE
        global WARNING_MESSAGES_LIST
        WARNING_MESSAGES_LIST = []

        global horizontal_email_message_list
        horizontal_email_message_list = []
        horizontal_email_message_list.append(self.ids[
                                                 "JobNumber"].text + " Program on " + "Old Horizontal Machine" + "\n" + "Created by : " + user_name + "\n")

        # **************************************#
        # CALL (load_horizontal_sheets_for_automation) FUNCTION TO LOAD HORIZONTAL TOOL LIST SHEETS
        load_horizontal_sheets_for_automation(self)

        # **************************************#
        # CALL (FOUR_CYCLE_PINBORE_VARIABLES) FUNCTION TO SET THE VARIABLES
        # if (self.ids["JobNumber"].text == ""):
        #      FOUR_CYCLE_PINBORE_VARIABLES()

        # NEED LOGIC TO CHECK IF NUMBER THAT ENTERED MATCH WITH SPEC DATABASE
        # self.ids["JobNumber"].text: TO ACCESS TEXT FIELD WE USE ids THAT'S DEFINED ABOVE
        global NEW_PROGRAM_NUMBER
        NEW_PROGRAM_NUMBER = self.ids["JobNumber"].text
        print(NEW_PROGRAM_NUMBER)
        if (self.ids["JobNumber"].text == ""):
            FAILED_MESSAGES_LIST.append("Please Enter Job Number.")
            # horizontal_email_message_list.append("\n".join(FAILED_MESSAGES_LIST))
            Failed_to_Create_Horizontal_Program(self)
            return

        # # DEFINE VARIABLE(horizontal_tool_list_file) TO STORE horizontal_tool_list_file PATH FROM SETTING SCREEN
        # # TO GET VALUE FROM ANOTHER SCREEN(SettingScreen IN THIS EXAMPLE) TO USE IT IN ANOTHER SCREEN CLASS (OldHorizontalScreen IN THIS EXAMPLE) FOR LOGIC PURPOSES>>
        # # >> USE (self.manager.get_screen('SettingScreen').ids["HorizontalToolList"].text)
        # # self.manager.get_screen('SettingScreen'): PUT SCREEN YOU WANT TO ACCESS
        # # ids["HorizontalToolList"].text : PUT WIDGET OR VALUE YOU NEED TO GET(IN THIS EXAMPLE WE GET THE TEXT OF MDTextField THAT NAME "HorizontalToolList")
        # horizontal_tool_list_file_path = self.manager.get_screen('SettingScreen').ids["HorizontalToolList"].text
        # print(horizontal_tool_list_file_path)
        # # TO READ EXCEL FILE THAT CONTAIN HORIZONTAL TOOL LIST AND PROBE PROGRAMS (WE USE -sheet_name=None- TO READ ALL THE SHEETS)
        # horizontal_tool_list_file = pd.read_excel(horizontal_tool_list_file_path,sheet_name=None)
        # ### LEAVE IT FOR TEST
        # ##print(horizontal_tool_list_file)
        # # TO PRINT SHEETS NAME
        # ##print(horizontal_tool_list_file.keys())
        # # TO PRINT SHEET DATA BY USING SHEET NAME('FINISH_BORE_TOOL_LIST' AS EXAMPLE)
        # ##print(horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'])
        #
        # # DEFINE LIST TO ADD ALL FINISH_BORE_TOOL_LIST TO THE LIST
        # FINISH_BORE_TOOL_LIST = []
        # # MAKE for LOOP TO READ ALL FINISH_BORE_TOOL_LIST IN SHEET['FINISH_BORE_TOOL_LIST'] IN COLUMN OF 'PIN_BORE_DIAMETER', AND ADDED TO THE LIST TO USE THEM
        # # horizontal_tool_list_file: THE EXCEL FILE THAT CONTAIN HORIZONTAL TOOL LIST AND PROBE PROGRAMS
        # # ['FINISH_BORE_TOOL_LIST']: THE SHEET THAT'S CONTAIN TOOL LIST OF FINISH BORE
        # # ['PIN_BORE_DIAMETER']: THE COLUMN THAT'S CONTAIN PIN_BORE_DIAMETER SIZES WE HAVE
        # for tool in horizontal_tool_list_file['FINISH_BORE_TOOL_LIST']['PIN_BORE_DIAMETER']:
        #     # print(tool)
        #     FINISH_BORE_TOOL_LIST.append(tool)
        # print("FINISH_BORE_TOOL_LIST:", FINISH_BORE_TOOL_LIST)
        #
        # # DEFINE LIST TO ADD ALL ROUGH_BORE_TOOL_LIST TO THE LIST
        # ROUGH_BORE_TOOL_LIST = []
        # # MAKE for LOOP TO READ ALL ROUGH_BORE_TOOL_LIST IN SHEET['ROUGH_BORE_TOOL_LIST'] IN COLUMN OF 'DRILL_DIAMETER', AND ADDED TO THE LIST TO USE THEM
        # for tool in horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST']['DRILL_DIAMETER']:
        #     # print(tool)
        #     ROUGH_BORE_TOOL_LIST.append(tool)
        # print("ROUGH_BORE_TOOL_LIST:", ROUGH_BORE_TOOL_LIST)
        #
        # # DEFINE LIST TO ADD ALL LOCK_RING_AND_CFREN_TOOL TO THE LIST
        # LOCK_RING_AND_CFREN_TOOL_LIST = []
        # # MAKE for LOOP TO READ ALL LOCK_RING_AND_CFREN_TOOL_LIST IN SHEET['LOCK_RING_AND_CFREN_TOOL_LIST'] IN COLUMN OF 'TOOL_WIDTH', AND ADDED TO THE LIST TO USE THEM
        # for tool in horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST']['TOOL_WIDTH']:
        #     # print(tool)
        #     LOCK_RING_AND_CFREN_TOOL_LIST.append(tool)
        # print("LOCK_RING_AND_CFREN_TOOL_LIST:", LOCK_RING_AND_CFREN_TOOL_LIST)
        #
        # # DEFINE LIST TO ADD ALL MISCELLANEOUS_TOOL_LIST TO THE LIST
        # MISCELLANEOUS_TOOL_LIST = []
        # # MAKE for LOOP TO READ ALL MISCELLANEOUS_TOOL_LIST IN SHEET['MISCELLANEOUS_TOOL_LIST'] IN COLUMN OF 'TOOL_USAGE', AND ADDED TO THE LIST TO USE THEM
        # for tool in horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST']['TOOL_USAGE']:
        #     # print(tool)
        #     MISCELLANEOUS_TOOL_LIST.append(tool)
        # print("MISCELLANEOUS_TOOL_LIST:", MISCELLANEOUS_TOOL_LIST)
        #
        # # DEFINE LIST TO ADD ALL HORIZONTAL_SLOT_NUMBERS TO THE LIST
        # HORIZONTAL_SLOT_NUMBERS = []
        # # MAKE for LOOP TO READ ALL HORIZONTAL_SLOT_NUMBERS IN SHEET['HORIZONTAL_SLOT_NUMBERS'] IN COLUMN OF 'PIN_BORE_DIAMETER', AND ADDED TO THE LIST TO USE THEM
        # for tool in horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS']['PIN_BORE_DIAMETER']:
        #     # print(tool)
        #     HORIZONTAL_SLOT_NUMBERS.append(tool)
        # print("HORIZONTAL_SLOT_NUMBERS:", HORIZONTAL_SLOT_NUMBERS)
        #
        # # DEFINE LIST TO ADD ALL PROBE_PROGRAMS TO THE LIST
        # PROBE_PROGRAMS = []
        # # MAKE for LOOP TO READ ALL PROBE_PROGRAMS IN SHEET['PROBE_PROGRAMS'] IN COLUMN OF 'FORGING_NUMBER', AND ADDED TO THE LIST TO USE THEM
        # for program in horizontal_tool_list_file['PROBE_PROGRAMS']['FORGING_NUMBER']:
        #     # print(tool)
        #     PROBE_PROGRAMS.append(program)
        # print("PROBE_PROGRAMS:", PROBE_PROGRAMS)
        # ===========NEED TO ADD CHECK ABOUT CAN'T FIND FORGING NUMBER==================

        # ===============================================================================================================
        # region <<<<========================================[PROBE PROGRAMS]========================================>>>>

        # self.manager.get_screen('SettingScreen'): SCREEN WE WANT TO ACCESS
        # ids["ProbePrograms"].text : TO GET THE TEXT OF ["ProbePrograms"] THAT'S CONTAIN PROBE PROGRAMS FILE PATH TO SEARCH INSIDE THE FILE
        PROBE_PROGRAMS_FILE_PATH = self.manager.get_screen('OldHorizontalSettingScreen').ids["ProbePrograms"].text
        # DEFINE AN EMPTY LIST TO STORE RESULT OF PROBE PROGRAM SEARCH.
        RESULT_OF_PROBE_PROGRAM_SEARCH = []
        for file in glob.glob(PROBE_PROGRAMS_FILE_PATH + '*\*' + FORGING_NUMBER + '*'):
            # print(file)
            RESULT_OF_PROBE_PROGRAM_SEARCH.append(file)
        print(len(RESULT_OF_PROBE_PROGRAM_SEARCH))
        print(RESULT_OF_PROBE_PROGRAM_SEARCH)
        # DEFINE THAT JUST TO AVOID ERROR IF RESULT OF SEARCH IS MORE THAN ONE
        ## PROBE_PROGRAM = 'OXXXX'
        # DEFINE AN EMPTY LIST TO ADD PROBE PROGRAM LINES ONE BY ONE.
        PROBE_PROGRAMS_LINES = []
        # IF ONE PROBE PROGRAM FOUND , GO A HEAD AND OPEN IT AND TAKE FIRST LINE OF PROBE PROGRAM TO USE IT IN HORIZ TEMPLATE
        if ((len(RESULT_OF_PROBE_PROGRAM_SEARCH) == 1) and (FORGING_NUMBER != "")):
            print("RESULT_OF_PROBE_PROGRAM_SEARCH ", RESULT_OF_PROBE_PROGRAM_SEARCH)
            # print(RESULT_OF_PROBE_PROGRAM_SEARCH[0])
            with open(RESULT_OF_PROBE_PROGRAM_SEARCH[0], 'rt') as CurrentProgram:
                for line in CurrentProgram:  # For each line in the file,
                    PROBE_PROGRAMS_LINES.append(line.rstrip('\n'))  # strip newline and add to list.
                # print(PROBE_PROGRAMS_LINES)
                ## print("LINE OF PROBE PROGRAM THAT NEED TO ADD TO HORIZONTAL PROGRAM: " + PROBE_PROGRAMS_LINES[0])
                PROBE_PROGRAM = PROBE_PROGRAMS_LINES[0]
        # maybe we don't need it
        elif (self.ids["JobNumber"].text == ""):
            FAILED_MESSAGES_LIST.append("Please Enter Job Number.")
            # horizontal_email_message_list.append("\n".join(FAILED_MESSAGES_LIST))
            Failed_to_Create_Horizontal_Program(self)
            return
        elif (FORGING_NUMBER == ""):
            FAILED_MESSAGES_LIST.append(
                "Forging Number does NOT found." + '\n' + "Double Check Job Spec with Engineering and Try Again.")
            horizontal_email_message_list.append("\n".join(FAILED_MESSAGES_LIST))
            Failed_to_Create_Horizontal_Program(self)
            return
        # IF NO PROGRAM FOUND, WE NEED TO CHECK MANUALLY IF IT IS THERE , OTHERWISE CREATE NEW PROBE PROGRAM AND TRY AGAIN
        elif ((len(RESULT_OF_PROBE_PROGRAM_SEARCH) == 0) and (FORGING_NUMBER != "")):
            # WE NEED TO CREATE DIALOG TO POP UP THE MESSAGE FOR USER
            FAILED_MESSAGES_LIST = [
                "Probe Program does NOT found, Double Check in case it is there," + '\n' + "OTHERWISE Create new Probe Program and Try Again."]
            print(FAILED_MESSAGES_LIST)
            horizontal_email_message_list.append("\n".join(FAILED_MESSAGES_LIST))
            Failed_to_Create_Horizontal_Program(self)
            return
        elif ((len(RESULT_OF_PROBE_PROGRAM_SEARCH) > 1) and (FORGING_NUMBER != "")):
            # WE NEED TO CREATE DIALOG TO POP UP THE MESSAGE FOR USER
            if (len(RESULT_OF_PROBE_PROGRAM_SEARCH) <= 20):
                FAILED_MESSAGES_LIST = ["Many Probe Programs found for this Forging :" + '\n' + '\n' + (
                    '\n'.join(RESULT_OF_PROBE_PROGRAM_SEARCH)) + '\n' + '\n' + "Fix the Confusion and Try Again."]
            else:
                FAILED_MESSAGES_LIST = [
                    "Many Probe Programs found for this Forging" + '\n' + "Fix the Confusion and Try Again."]
            print(FAILED_MESSAGES_LIST)
            horizontal_email_message_list.append("\n".join(FAILED_MESSAGES_LIST))
            print(len(RESULT_OF_PROBE_PROGRAM_SEARCH))
            print(RESULT_OF_PROBE_PROGRAM_SEARCH)
            Failed_to_Create_Horizontal_Program(self)
            return
        else:
            FAILED_MESSAGES_LIST = [
                "UNEXPECTED issue about Probe Programs." + '\n' + "Double Check Network and probe program folder and Try Again."]
            print(FAILED_MESSAGES_LIST)
            horizontal_email_message_list.append("\n".join(FAILED_MESSAGES_LIST))
            Failed_to_Create_Horizontal_Program(self)
            return
        # endregion
        # ===============================================================================================================

        # DEFINE VARIABLE(HORIZONTAL_TEMPLATE_PATH) TO STORE HorizontalTemplate PATH FROM SETTING SCREEN
        # TO GET VALUE FROM ANOTHER SCREEN(SettingScreen IN THIS EXAMPLE) TO USE IT IN ANOTHER SCREEN CLASS (OldHorizontalScreen IN THIS EXAMPLE) FOR LOGIC PURPOSES>>
        # >> USE (self.manager.get_screen('SettingScreen').ids["HorizontalTemplate"].text)
        # self.manager.get_screen('SettingScreen'): PUT SCREEN YOU WANT TO ACCESS
        # ids["HorizontalTemplate"].text : PUT WIDGET OR VALUE YOU NEED TO GET(IN THIS EXAMPLE WE GET THE TEXT OF MDTextField THAT NAME "HorizontalTemplate")
        HORIZONTAL_TEMPLATE_PATH = self.manager.get_screen('OldHorizontalSettingScreen').ids["HorizontalTemplate"].text
        ###LEAVE IT FOR TEST
        print(HORIZONTAL_TEMPLATE_PATH)

        # DEFINE list TO CONTAIN HORIZONTAL PROGRAM
        global HorizontalProgramLines
        HorizontalProgramLines = []
        # TO OPEN HORIZONTAL_TEMPLATE AND ADD EACH SINGLE LINE TO list WE CREATE ABOVE(HorizontalProgramLines)
        with open(HORIZONTAL_TEMPLATE_PATH, 'rt') as CurrentProgram:
            for line in CurrentProgram:  # For each line in the file,
                HorizontalProgramLines.append(line.rstrip('\n'))  # strip newline and add to list.
            ###LEAVE IT FOR TEST
            ## TO PRINT ORIGINAL HORIZONTAL TEMPLATE
            # print(HorizontalProgramLines)
            # print()
            # print('\n'.join(HorizontalProgramLines))
            # print()
            ## TO REPLACE THE FIRST ITEM OF LIST WITH JOB NUMBER AND TODAYDATE
            HorizontalProgramLines[0] = ('(PART ' + NEW_PROGRAM_NUMBER + ' -- ' + todaydate + ' <SYS>' + ')')
            ## WE CAN USE CODE BELOW TO ADD THE TOOL LIST AFTER CHECK EACH TOOL LOGIC
            # DEFINE VARIABLE TO FINISH THE while LOOP OF TOOL LIST
            HORIZONTAL_PROGRAM_TOOL_LIST_END = 0
            for line in HorizontalProgramLines:
                while HORIZONTAL_PROGRAM_TOOL_LIST_END == 0:
                    # WE CAN USE ((**********TOOL LIST**********)) LINE IN HORIZONTAL TEMPLATE TO LOCATE THE index OF list TO START ADD OTHER TOOLS(IF APPLICABLE)(BY ADD 1 TO INDEX WE FOUND)
                    if '(**********TOOL LIST**********)' in HorizontalProgramLines:
                        # TO FIND THE INDEX OF ITEM INSIDE LIST
                        index = HorizontalProgramLines.index('(**********TOOL LIST**********)')
                        ###LEAVE IT FOR TEST
                        # print(index)
                        index += 1
                    else:
                        print("CHECK HORIZONTAL TEMPLATE AND CORRECTED TO HAVE '(**********TOOL LIST**********)'")
                    # JUST FOR NOW, NEED TO ADD LOGIC TO CHECK IF THE FORGING HAS PILOT DIAMETER WITH 2.25 OR SMALLER TO 1.70
                    PILOT_DIAMETER = 2.25
                    PILOT_AVAILABILITY_STATUS = 0
                    if (PILOT_DIAMETER == 2.25):
                        PILOT_AVAILABILITY_STATUS = 1
                        # TO ADD TOOL DESCRIPTION FROM TOOL LIST FROM EXCEL FILE,
                        # NEED TO FIND THE INDEX OF THE TOOL IN THE TOOL LIST IN EXCEL FILE BY PUT THE TOOL USAGE FROM 1ST COLUMN IN MISCELLANEOUS SHEET(WHICH IS: '2.250 PILOT BORE FOR 2.25 DIA')
                        # STORE THE INDEX OF THE TOOL IN VARIABLE(MISCELLANEOUS_TOOL_LIST_INDEX)
                        # INSERT THE NEW ELEMENT TO HorizontalProgramLines LIST BY USE insert METHOD, AND index THAT FOUND FROM PREVIOSE if STATEMENT
                        # AND USE PANDAS METHOD TO ACCESS EXCEL FILE(horizontal_tool_list_file), AND SHEET('MISCELLANEOUS_TOOL_LIST')>>
                        # >> WITH INDEX OF THE TOOL(MISCELLANEOUS_TOOL_LIST_INDEX) WE FOUND, AND COLUMN NAME(DESCRIPTION(FOR_MACHINES_27/28/32))
                        MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index('2.250 PILOT BORE FOR 2.25 DIA')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                            MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # HorizontalProgramLines.insert(index, '(T01 IS A 2.250 PILOT BORE)')
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                        # NEED TO PUT THIS LINE IN THE LAST if STATEMENT IN THIS while LOOP TO AVOID THE for LOOP STATRT AGAIN
                        ## HorizontalProgramEnd = 1
                        # NEED TO MAKE SURE IF TEMPLATE OF 1.70 PILOT DIAMETER IS DEFFERENT FROM TEMPLATE OF 2.25 PILOT DIAMETER
                    elif (PILOT_DIAMETER == 1.70):
                        PILOT_AVAILABILITY_STATUS = 1
                        MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index('1.70 PILOT BORE FOR 1.70 DIA')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                            MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # HorizontalProgramLines.insert(index, '(T57 IS A 1.7 PILOT BORE)')
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                    # NEED TO ADD LOGIC OF LEDGE TOOL THAT WILL CHECK IF (DISTANCE_FROM_ORIGIN_TO_HIGHEST_POINT_OF_BORE) SMALLER THAN>>
                    # >>((X) Outsd Ring Belt Ht FROM EMSS: IT IS DISTANCE FROM ORIGIN TO END POINT BEFORE BOSS)
                    ##OUTSIDE_RING_BELT_HIEGHT= 1  # JUST FOR NOW NEED TO HAVE THIS VALUE FROM FORGING DATABASE
                    ##DISTANCE_FROM_ORIGIN_TO_HIGHEST_POINT_OF_BORE = PILOT_TO_DOME_DEPTH- PILOT_TO_PIN_DISTANCE - PIN_HOLE_DIAMETER
                    # just for now                                              # =============NEED TO FIX HERE================
                    # global LEDGE_CUT_AVAILABILITY_STATUS
                    if (ledge_cut_availability_status == ""):
                        CONFIRMATION_MESSAGES_LIST = [
                            "Ledge status can't detect," + '\n' + "Does this job need to use Ladge Tool ?"]
                        print(CONFIRMATION_MESSAGES_LIST)
                        confirmation_horizontal_email_message_list.append(
                            "# It was need confirmation of :" + "\n" + "\n".join(CONFIRMATION_MESSAGES_LIST))
                        self.title = '[color=0066ff]Confirmation Message[/color] ' + '\n' + '\n' + '[b][i][color=ffffff]' + \
                                     CONFIRMATION_MESSAGES_LIST[0] + '[/color][/i][/b]'
                        # _Option
                        self.Need_Ledge_Tool_Option = Item(text="Yes, it is Need", on_release=self.Need_Ledge_Tool)
                        self.Does_Not_Need_Ledge_Tool_Option = Item(text="No, it Does NOT Need",
                                                                    on_release=self.Does_Not_Need_Ledge_Tool)

                        self.items = [self.Need_Ledge_Tool_Option, self.Does_Not_Need_Ledge_Tool_Option]
                        Need_Confermation_to_Create_Horizontal_Program(self, self.title, self.Decide_Ledge_Tool_Status,
                                                                       "confirmation", self.items)

                        print("LEDGE_CUT_AVAILABILITY_STATUS in create function: ", ledge_cut_availability_status)
                        return

                    if (ledge_cut_availability_status == 1 or ledge_counterbore_diameter != 0):
                        if (pin_hole_diameter >= 0.629):
                            MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index('LEDGE TOOL 0.625 DIA')
                            HorizontalProgramLines.insert(index,
                                                          horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                                                              MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (pin_hole_diameter < 0.629):
                            MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index('LEDGE TOOL 0.375 DIA')
                            HorizontalProgramLines.insert(index,
                                                          horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                                                              MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])

                            # WE NEED TO REPLACE THE TOOL ON ((*******CONSTANTS********)) SECTION ON TEMPLATE
                            # FIND LOCATION OF NOTE OF LEDGE TOOL ON HorizontalProgramLines List
                            LEDGE_TOOL_NOTE_INDEX = HorizontalProgramLines.index(
                                '(T06 IS THE STD .625 CARBIDE END MILL LEDGE TOOL)')
                            # REPLACE THE NOTE TO USE TOOL #16
                            HorizontalProgramLines[LEDGE_TOOL_NOTE_INDEX] = ('(T16 IS A 3/8 END MILL - LEDGE TOOL)')

                            # FIND LOCATION OF VARIABLE ON HorizontalProgramLines List
                            VC103_VARIABLE_INDEX = HorizontalProgramLines.index('VC103=06')
                            # REPLACE THE TOOL TO USE TOOL #16
                            HorizontalProgramLines[VC103_VARIABLE_INDEX] = ('VC103=16')
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                    # LOGIC FOR ADD THE ROUGH BORE TOOL TO THE TOOL LIST IN THE PROGRAM AND TO THE TOOLS SECTION
                    if (0.4500 <= pin_hole_diameter < 0.4960):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('11MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.4960 <= pin_hole_diameter < 0.5355):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('12MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.5355 <= pin_hole_diameter < 0.5750):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('13MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.5750 <= pin_hole_diameter < 0.6142):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('14MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.6142 <= pin_hole_diameter < 0.6536):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('15MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.6536 <= pin_hole_diameter < 0.6930):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('16MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.6930 <= pin_hole_diameter < 0.7323):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('17MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.7323 <= pin_hole_diameter < 0.7717):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('18MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.7717 <= pin_hole_diameter < 0.8111):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('19MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.8111 <= pin_hole_diameter < 0.8504):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('20MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.8504 <= pin_hole_diameter < 0.8898):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('21MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.8898 <= pin_hole_diameter < 0.9292):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('22MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9292 <= pin_hole_diameter < 0.9686):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('23MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9686 <= pin_hole_diameter < 1.0079):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('24MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (1.0079 <= pin_hole_diameter < 1.0473):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('1_INCH')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (1.0473 <= pin_hole_diameter < 1.0942):
                        ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('26MM')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                            ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (pin_hole_diameter == 0):
                        # MAYBE WE CAN PUT MESSAGE ON TOP OF PROGRAM AND ON THE DIALOG OF APP IF ANYTHING UNEXPECTED HAPPEN
                        print("PIN HOLE DIAMETER IS MISSING, SEE PROGRAMMING")
                        # JUST TO NOT GIVE AN ERROR WHEN USE IT ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE, BECAUSE THE VERIABLE CAN NOT BE NOTHING, SHOULD HAVE A VALUE
                        ROUGH_BORE_TOOL_NUMBER = 0
                    else:
                        # NEED TO BACK AND WORK ON IT
                        print("MAYBE WE CAN USE SWAP TOOL T60 OR JUST PUT MESSAGE SEE PROGRAMMING")

                    # LOGIC FOR ADD THE FINISH BORE TOOL TO THE TOOL LIST IN THE PROGRAM AND TO THE TOOLS SECTION
                    # NEED TO CHECK IF WE GONNA USE MAPEL TOOL OR BORING BAR
                    if (0.4719 <= pin_hole_diameter <= 0.4729):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.4724)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    # NEED TO CHECK IF WE GONNA USE MAPEL TOOL OR BORING BAR
                    elif (0.4895 <= pin_hole_diameter <= 0.4905):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.49)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.5115 <= pin_hole_diameter <= 0.5125):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.512)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.5505 <= pin_hole_diameter <= 0.5515):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.551)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.5905 <= pin_hole_diameter <= 0.5915):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.591)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.6295 <= pin_hole_diameter <= 0.6305):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.63)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.6685 <= pin_hole_diameter <= 0.6695):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.669)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.6715 <= pin_hole_diameter <= 0.6725):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.672)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.7085 <= pin_hole_diameter <= 0.7095):  # IT SEEMS YOU CAN USE IF YPU HAVE 0.708 AS WELL
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.709)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.7475 <= pin_hole_diameter <= 0.7485):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.748)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.7865 <= pin_hole_diameter <= 0.7875):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.787)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.7905 <= pin_hole_diameter <= 0.7915):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.791)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.7915 <= pin_hole_diameter <= 0.7925):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.792)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                        # NEED TO CHECK IF WE CAN USE ROUGH TOOL OF 20MM INSTEAD 19MM AS OLD PROGRAMS DONE
                    elif (0.8119 <= pin_hole_diameter <= 0.8129):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.8124)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.8265 <= pin_hole_diameter <= 0.8275):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.827)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.8655 <= pin_hole_diameter <= 0.8665):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.866)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9045 <= pin_hole_diameter <= 0.9055):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.905)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9115 <= pin_hole_diameter <= 0.9125):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.912)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9265 <= pin_hole_diameter <= 0.9275):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.927)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                        # ******************MAYBE NEED TO ADD 0.943 AS MAPLE USE T45**********************

                    elif (0.9445 <= pin_hole_diameter <= 0.9455):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.945)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9745 <= pin_hole_diameter <= 0.9755):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.975)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9795 <= pin_hole_diameter <= 0.9805):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.98)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9835 <= pin_hole_diameter <= 0.9845):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.984)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9895 <= pin_hole_diameter <= 0.9905):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.99)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (0.9995 <= pin_hole_diameter <= 1.0005):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(1.00)
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                        # MAYBE NEED TO ADD ANOTHER LOGIC TO USE 1.094 MAPEL TOOL IF NEED IT

                    elif (1.0935 <= pin_hole_diameter <= 1.0945):
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index('1.094_BORING')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                    elif (pin_hole_diameter == 0):
                        # MAYBE WE CAN PUT MESSAGE ON TOP OF PROGRAM AND ON THE DIALOG OF APP IF ANYTHING UNEXPECTED HAPPEN
                        print("PIN HOLE DIAMETER IS MISSING, SEE PROGRAMMING")
                        # JUST TO NOT GIVE AN ERROR WHEN USE IT ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE, BECAUSE THE VERIABLE CAN NOT BE NOTHING, SHOULD HAVE A VALUE
                        FINISH_BORE_TOOL_NUMBER = 0
                    else:
                        FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index('BORING_BAR_TOOL')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                            pin_hole_diameter) + ' BORING BAR)')
                        # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                            FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                        # WE NEED TO BACK TO ADD IF WE NEED SWAP TOOL TO USE T45 MAPEL

                        # WE NEED TO BACK TO WORK IF NEED TO MAKE LOGIC OF (T03 IS A .927 MAPAL REAMER - WILL HONE TO .928)
                        #               ************(MAYBE WE CAN MAKE THE RULES FOR ALL MAPLE TOOLS BETWEEN -0.001 AND 0.001) EXAMPLE WD-13731**************

                    # NEED LOGIC TO ADD BOTH LOCKRING AND CFREN TOOL DESCRIPTION TO THE TOOL LIST IN THE PROGRAM IF THEY ARE NOT THE SAME TOOL EX WD-13100
                    # NEED TO ADJUST HORIZONTAL TEMPLATE ALSO

                    # LOGIC FOR ADD THE LOCKRING/C_FREN TOOL TO THE TOOL LIST IN THE PROGRAM
                    if (lock_ring_cutter_width == 0.0 and cfren_cutter_width == 0.0):
                        # IT IS MEAN JOB DOESN'T HAVE LOCKRING NOR CFREN
                        pass
                        # JUST TO NOT GIVE AN ERROR WHEN USE IT ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE, BECAUSE THE VERIABLE CAN NOT BE NOTHING, SHOULD HAVE A VALUE
                        LOCK_RING_TOOL_NUMBER = 0
                    # WE SEPARATE LOGIC OF LOCKRING AND CFREN FOR THE NEXT elif STATEMENTS BECAUSE WE NEED TO USE TOOL CUTTER WIDTH IN THE DESCRIPTION
                    elif (lock_ring_cutter_width == 0.039):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.039)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                                          lock_ring_cutter_width) + ' SQ X .465 DIA. LOCK RING)')
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (cfren_cutter_width == 0.039):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.039)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                                          cfren_cutter_width) + ' SQ X .465 DIA. LOCK RING)')
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                        # NEED TO CHECK WITCH TOOL WITH 0.042 WIDTH WILL USE WHILE WE HAVE THREE TOOLS WITH DIFFERENT DIAMETER (OR ADD LOGIC TO CHECK WICTH TOOL WILL USE),
                        # NEED TO ADD ANOTHER TOOL TO EXCEL SHEET AS WELL
                    elif (lock_ring_cutter_width == 0.042):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.042)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                                          lock_ring_cutter_width) + ' SQ X .465 PH HORN LOCK RING)')
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (cfren_cutter_width == 0.042):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.042)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                                          cfren_cutter_width) + ' SQ X .465 PH HORN LOCK RING)')
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))


                    # BACK TO USE LOGIC OF LOCKRING AND CFREN IN THE ONE if STATEMENT BECAUSE WH HAVE THE TOOLS FOR THE NEXT WIDTH

                    # NEED TO CHECK WITCH TOOL WITH 0.044 WIDTH WILL USE WHILE WE HAVE THREE TOOLS WITH DIFFERENT DIAMETER (OR ADD LOGIC TO CHECK WICTH TOOL WILL USE),
                    # FOR NOW WE WILL USE (T25 IS A .044 RAD X .465 PH HORN LOCK RING)
                    elif (lock_ring_cutter_width == 0.044 or cfren_cutter_width == 0.044):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index('0.044(DIA 0.465)')
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (lock_ring_cutter_width == 0.047 or cfren_cutter_width == 0.047):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.047)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                    # WE SEPARATE LOGIC OF LOCKRING AND CFREN FOR THE NEXT elif STATEMENTS BECAUSE WE NEED TO USE TOOL CUTTER WIDTH IN THE DESCRIPTION

                    elif (lock_ring_cutter_width == 0.048):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.048)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                                          lock_ring_cutter_width) + ' RAD X .575 DIA. LOCK RING)')
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (cfren_cutter_width == 0.048):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.048)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                                          cfren_cutter_width) + ' RAD X .575 DIA. LOCK RING)')
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                    # BACK TO USE LOGIC OF LOCKRING AND CFREN IN THE ONE if STATEMENT BECAUSE WH HAVE THE TOOLS FOR THE NEXT WIDTH

                    # NEED TO CHECK WITCH TOOL WITH 0.053 WIDTH WILL USE WHILE WE HAVE THREE TOOLS WITH DIFFERENT DIAMETER (OR ADD LOGIC TO CHECK WICTH TOOL WILL USE),
                    # FOR NOW WE WILL USE (T09 IS A .053 RAD X .618 PH HORN LOCK RING)
                    elif (lock_ring_cutter_width == 0.053 or cfren_cutter_width == 0.053):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index('0.053(DIA 0.618)')
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (lock_ring_cutter_width == 0.059 or cfren_cutter_width == 0.059):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.059)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (lock_ring_cutter_width == 0.063 or cfren_cutter_width == 0.063):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.063)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (lock_ring_cutter_width == 0.065 or cfren_cutter_width == 0.065):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.065)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (lock_ring_cutter_width == 0.067 or cfren_cutter_width == 0.067):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.067)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (lock_ring_cutter_width == 0.076 or cfren_cutter_width == 0.076):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.076)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (lock_ring_cutter_width == 0.077 or cfren_cutter_width == 0.077):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.077)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (lock_ring_cutter_width == 0.088 or cfren_cutter_width == 0.088):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.088)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))
                    elif (lock_ring_cutter_width == 0.11 or cfren_cutter_width == 0.11):
                        LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.11)
                        HorizontalProgramLines.insert(index,
                                                      horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                                          LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                        CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                    # NEED TO BACK AND WORK HERE
                    # NEED TO CHECK IF WE CAN USE T43 WITH ANY TOOL WIDTH COMES WITH JOB(LIKE WHAT WE DO FOR BORING BAR),
                    # OR NEED TO PUT MESSAGE ON TOP OF PROGRAM OR ON DIALOG OF APP TO QUESTION THE TOOL AVAILABILITY
                    else:
                        print("LOCK RING TOOL IS NOT ON THE TOOL LIST, SEE PROGRAMMING")
                        # JUST TO NOT GIVE AN ERROR WHEN USE IT ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE, BECAUSE THE VERIABLE CAN NOT BE NOTHING, SHOULD HAVE A VALUE
                        LOCK_RING_TOOL_NUMBER = 0

                    DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS = 0  # JUST FOR NOW
                    # NEED TO DOUBLE CHECK IF PIN_HOLE_DIAMETER SHOULD BE 0.900 OR SOMETHING ELSE FROM HORIZ TEMPLATE,  MAYBE WE CAN MAKE LOGIC IN HORIZ TEMPLATE AS RANGE INSTEAD IT IS EXACT VALUE
                    # NEED TO CHECK IF WE NEED TO ADJUST THE CONDITION OF (PIN_HOLE_DIAMETER >= 0.901) TO CHECK IF PIN HOLE DIAMETER IS COVERED ON HORIZ TEMPLATE LOGIC
                    # NEED TO DOUBLE CHECK IF LOGIC IN HORIZ TEMPLATE COVER ALL FINISH BORE SIZES (LIKE: 0.9839 NOT COVERED )
                    if (double_oil_hole_slot_id_spacing != 0 and pin_hole_diameter >= 0.901):
                        DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS = 1
                        MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index(
                            'DOUBLE OIL HOLES SLOTS(DOHS) 0.750 PH')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc
                        [MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        # DEFINE VARIABLE (DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                        DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER = horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                            MISCELLANEOUS_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                    NOTCH_AVAILABILITY_STATUS = 0
                    if (notch_angle_first_location != 0):
                        NOTCH_AVAILABILITY_STATUS = 1
                        MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index('NOTCH TOOL 5/32 DIA')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc
                        [MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        index += 1
                        ##print('\n'.join(HorizontalProgramLines))

                    HORIZONTAL_SLOTS_AVAILABILITY_STATUS = 0
                    HORIZONTAL_SLOTS_STRAIGHT_THROUGH_AVAILABILITY_STATUS = 0
                    if (horizontal_slots_arc_diameter != 0):
                        MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index('HORIZONTAL SLOTS TOOL 0.375 DIA')
                        HorizontalProgramLines.insert(index, horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc
                        [MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                        index += 1
                        # MAYBE WE NEED TO FIX THIS CONDITION LATER
                        ## WE STILL NEED TO BACK HERE AND WORK ON CONDITION, MAYBE NEED TO ADD SOMETHING FROM QANTEL DATABASE
                        if (horizontal_slots_arc_diameter != 0.375):
                            HORIZONTAL_SLOTS_AVAILABILITY_STATUS = 1
                        elif (horizontal_slots_arc_diameter == 0.375):
                            HORIZONTAL_SLOTS_STRAIGHT_THROUGH_AVAILABILITY_STATUS = 1
                        ##print('\n'.join(HorizontalProgramLines))

                        # MAYBE WE DON'T WANT THIS SECTION IF WE MAKE HORIZ SLOT STANDER WITH BORE SIZE
                    # +++++-----------------------------******************************************************************************--------------------------------------+++++#
                    # +++++-------LOGIC TO SET HORIZONTAL SLOT NUMBERS ACCORDING TO PIN_HOLE_DIAMETER TO USE THEM LATER IN VARIABLES(VC176,VC177, AND VC178) SECTIONS-------+++++#
                    # +++++-----------------------------******************************************************************************--------------------------------------+++++#
                    # MAYBE THE WAY WE PROGRAM HORIZONTAL SLOTS WILL CHANGE SOON , BUT FOR NOW WE WILL USE THE WAY WE HAVE
                    ### SOME OF THESE NUMBERS ARE MISSING, NEED TO BACK AND FIGURE OUT THESE NUMBERS OR MAKE MESSAGE LIKE: CAN'T FIND THEM OR SOMETHING SIMILER,
                    ### ALSO NEED A LOGIC TO CHECK WHEN WE HAVE TO MAKE MANUAL HORIZ SLOT PROGRAM BY MASTERCAM
                    if (0.4719 <= pin_hole_diameter <= 0.4729):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.4724)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.4895 <= pin_hole_diameter <= 0.4905):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.49)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.5115 <= pin_hole_diameter <= 0.5125):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.512)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.5505 <= pin_hole_diameter <= 0.5515):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.551)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.5905 <= pin_hole_diameter <= 0.5915):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.591)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.6295 <= pin_hole_diameter <= 0.6305):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.63)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.6685 <= pin_hole_diameter <= 0.6695):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.669)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.6715 <= pin_hole_diameter <= 0.6725):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.672)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.7085 <= pin_hole_diameter <= 0.7095):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.709)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.7278 <= pin_hole_diameter <= 0.7288):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.7283)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.7475 <= pin_hole_diameter <= 0.7485):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.748)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.7865 <= pin_hole_diameter <= 0.7875):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.787)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.7905 <= pin_hole_diameter <= 0.7915):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.791)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.7915 <= pin_hole_diameter <= 0.7925):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.792)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.7995 <= pin_hole_diameter <= 0.8005):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.800)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.8119 <= pin_hole_diameter <= 0.8129):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.8124)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.8265 <= pin_hole_diameter <= 0.8275):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.827)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.8265 <= pin_hole_diameter <= 0.8275):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.827)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.8655 <= pin_hole_diameter <= 0.8665):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.866)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.8745 <= pin_hole_diameter <= 0.8755):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.875)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9005 <= pin_hole_diameter <= 0.9015):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.901)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9045 <= pin_hole_diameter <= 0.9055):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.905)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9115 <= pin_hole_diameter <= 0.9125):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.912)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9265 <= pin_hole_diameter <= 0.9275):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.927)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9395 <= pin_hole_diameter <= 0.9405):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.94)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9445 <= pin_hole_diameter <= 0.9455):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.945)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9745 <= pin_hole_diameter <= 0.9755):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.975)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9795 <= pin_hole_diameter <= 0.9805):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.98)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9820 <= pin_hole_diameter <= 0.9830):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.9825)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9835 <= pin_hole_diameter <= 0.9845):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.984)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9895 <= pin_hole_diameter <= 0.9905):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.99)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (0.9995 <= pin_hole_diameter <= 1.0005):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(1.00)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (1.0305 <= pin_hole_diameter <= 1.0315):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(1.031)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    elif (1.0935 <= pin_hole_diameter <= 1.0945):
                        HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(1.094)
                        # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                        i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                        j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                        # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                        HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                            HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                    else:
                        ### STILL NEED TO BACK AND WORK HERE
                        # IN CASE THE NUMBERS NOT FOUND FOR PIN HOLE DIAMETER SIZE , MAKE A MESSAGE TO SAY THAT , MAYBE NEED TO CREATE DIALOG
                        print(
                            "LOOKS LIKE HORIZONTAL SLOT NUMBERS FOR THIS PIN HOLE DIAMETER NOT EXIST, MAYBE YOU CAN FIGURE OUT THEM OR NEED TO MAKE MANUAL HORIZ SLOTS ")

                    # END OF ADDING THE TOOL LIST FOR THE TEMPLATE
                    # NEED TO PUT THIS LINE IN END OF THE while LOOP TO AVOID THE for LOOP START AGAIN
                    HORIZONTAL_PROGRAM_TOOL_LIST_END = 1

            # WE CAN USE CODE BELOW TO ADJUST EACH VARIABLE DEPENDS ON JOB WE NEED TO MAKE PROGRAM FOR
            # MAKE for LOOP TO GO AND ITERATES THROUGH THE HORIZONTAL TEMPLATE (HorizontalProgramLines List)

            for line in HorizontalProgramLines:  # TO READ EACH LINE IN THE TEMPLATE

                # BEGINNING OF FEATURE LIST STATUS    (*******FEATURE LIST********)

                ### LEDGE CUT VARIABLE (VC118)
                substr = "VC118"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC118_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC118_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC118_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    # (LEDGE_CUT_AVAILABILITY_STATUS) COMES FROM LOGIC OF LEDGE CUT STATUS OF TOOL LIST ABOVE
                    HorizontalProgramLines[VC118_VARIABLE_INDEX] = (
                            'VC118=' + format(ledge_cut_availability_status) + '  (LedgeCut)')
                    print("LEDGE_CUT_AVAILABILITY_STATUS in program: ", ledge_cut_availability_status)

                ### PILOT VARIABLE (VC119)
                substr = "VC119"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC119_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC119_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC119_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    # (PILOT_AVAILABILITY_STATUS) COMES FROM LOGIC OF PILOT STATUS OF TOOL LIST ABOVE
                    HorizontalProgramLines[VC119_VARIABLE_INDEX] = (
                            'VC119=' + format(PILOT_AVAILABILITY_STATUS) + '  (Pilot)')

                ### LOCK RING VARIABLE (VC121)
                # WE DEFINE VARIABLE TO STORE LOCK RING AVAILABILITY STATUS
                # WE DEFINE IT HERE NOT ON TOOL LIST LOGIC ABOVE ,BECAUSE LOGIC ABOVE CHECK LOCK RING AND CFREN IN THE SAME STATMENT (MAYBE WE WILL CHANGE IT LATER, WILL SEE)
                LOCK_RING_AVAILABILITY_STATUS = 0
                substr = "VC121"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # MAYBE NEED TO ADJUST THE CONDITION BELOW TO NOT CHANGE THE STATUS IF WE DON'T HAVE THE TOOL WIDTH IN THE TOOL LIST
                if (index == 0 and lock_ring_cutter_width != 0):
                    LOCK_RING_AVAILABILITY_STATUS = 1
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC121_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC121_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC121_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC121_VARIABLE_INDEX] = (
                            'VC121=' + format(LOCK_RING_AVAILABILITY_STATUS) + '  (LockRing)')

                ### CFREN VARIABLE (VC122)
                # WE DEFINE VARIABLE TO STORE CFREN AVAILABILITY STATUS
                # WE DEFINE IT HERE NOT ON TOOL LIST LOGIC ABOVE ,BECAUSE LOGIC ABOVE CHECK LOCK RING AND CFREN IN THE SAME STATMENT (MAYBE WE WILL CHANGE IT LATER, WILL SEE)
                CFren_AVAILABILITY_STATUS = 0
                substr = "VC122"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0 and cfren_cutter_width != 0):
                    CFren_AVAILABILITY_STATUS = 1
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC122_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC122_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC122_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC122_VARIABLE_INDEX] = (
                            'VC122=' + format(CFren_AVAILABILITY_STATUS) + '  (C-Fren)')

                ### DOUBLE_OIL_HOLE_SLOT VARIABLE (VC123)
                substr = "VC123"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC123_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC123_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC123_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    # (DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS) COMES FROM LOGIC OF DOUBLE_OIL_HOLE_SLOT OF TOOL LIST ABOVE
                    HorizontalProgramLines[VC123_VARIABLE_INDEX] = (
                            'VC123=' + format(DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS) + '  (DOHS)')

                ### NOTCH VARIABLE (VC124)
                substr = "VC124"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC124_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC124_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC124_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    # (NOTCH_AVAILABILITY_STATUS) COMES FROM LOGIC OF NOTCH STATUS OF TOOL LIST ABOVE
                    HorizontalProgramLines[VC124_VARIABLE_INDEX] = (
                            'VC124=' + format(NOTCH_AVAILABILITY_STATUS) + '  (CirclipNotch)')

                ### HORIZONTAL SLOTS VARIABLE (VC125)
                substr = "VC125"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC125_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC125_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC125_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    # (HORIZONTAL_SLOTS_AVAILABILITY_STATUS) COMES FROM LOGIC OF HORIZONTAL_SLOTS STATUS OF TOOL LIST ABOVE
                    HorizontalProgramLines[VC125_VARIABLE_INDEX] = (
                            'VC125=' + format(HORIZONTAL_SLOTS_AVAILABILITY_STATUS) + '  (H-Slots)')

                ### LEDGE COUNTERBORE VARIABLE (VC126)
                # WE DEFINE VARIABLE TO STORE LEDGE COUNTERBORE AVAILABILITY STATUS
                LEDGE_COUNTERBORE_AVAILABILITY_STATUS = 0
                substr = "VC126"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0 and ledge_counterbore_diameter != 0):
                    LEDGE_COUNTERBORE_AVAILABILITY_STATUS = 1
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC126_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC126_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC126_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC126_VARIABLE_INDEX] = (
                            'VC126=' + format(LEDGE_COUNTERBORE_AVAILABILITY_STATUS) + '  (LedgeCounterbore)')

                ### DOUBLE NOTCH VARIABLE (VC127)
                # WE DEFINE VARIABLE TO STORE DOUBLE NOTCH AVAILABILITY STATUS
                DOUBLE_NOTCH_AVAILABILITY_STATUS = 0
                substr = "VC127"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0 and notch_angle_second_location != 0):
                    DOUBLE_NOTCH_AVAILABILITY_STATUS = 1
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC127_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC127_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC127_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC127_VARIABLE_INDEX] = (
                            'VC127=' + format(DOUBLE_NOTCH_AVAILABILITY_STATUS) + '   (Double Notch)')

                ### HORIZONTAL_SLOTS_STRAIGHT_THROUGH(0.375 DIA) VARIABLE (VC129)
                substr = "VC129"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR

                ## WE STILL NEED TO BACK HERE AND WORK ON CONDITION, MAYBE NEED TO ADD SOMETHING FROM QANTEL DATABASE
                if (index == 0 and horizontal_slots_arc_diameter == 0.375):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC129_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC129_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC129_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC129_VARIABLE_INDEX] = ('VC129=' + format(
                        HORIZONTAL_SLOTS_STRAIGHT_THROUGH_AVAILABILITY_STATUS) + '   (.375SlotsStraightThough)')

                # END OF FEATURE LIST STATUS (*******FEATURE LIST********)

                # BIGINNIG OF TOOLS NUMBER (*****TOOLS*****)

                ### ROUGH BORE TOOL NUUMBER VARIABLE (VC130)
                substr = "VC130"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC130_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC130_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC130_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    # WE JUST USE if STATEMENT BELOW TO MAKE THE TOOL NUMBER ALWAYS PRINTED WITH TWO DIGITS EVEN IT WAS ONE DIGIT(LIKE: 1,3,5,7)
                    # **IT IS NOT NECESSARY JUST TO KEEP TOOLS ORGANIZED AND HAVE 2 DIGITS**
                    if (ROUGH_BORE_TOOL_NUMBER < 10):
                        HorizontalProgramLines[VC130_VARIABLE_INDEX] = (
                                'VC130=0' + format(ROUGH_BORE_TOOL_NUMBER) + '  (RoughBoreToolNo)')
                    elif (ROUGH_BORE_TOOL_NUMBER >= 10):
                        HorizontalProgramLines[VC130_VARIABLE_INDEX] = (
                                'VC130=' + format(ROUGH_BORE_TOOL_NUMBER) + '  (RoughBoreToolNo)')

                ### FINISH BORE TOOL NUUMBER VARIABLE (VC132)
                substr = "VC132"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC132_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC132_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC132_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    # WE JUST USE if STATEMENT BELOW TO MAKE THE TOOL NUMBER ALWAYS PRINTED WITH TWO DIGITS EVEN IT WAS ONE DIGIT(LIKE: 1,3,5,7)
                    # **IT IS NOT NECESSARY JUST TO KEEP TOOLS ORGANIZED AND HAVE 2 DIGITS**
                    if (FINISH_BORE_TOOL_NUMBER < 10):
                        HorizontalProgramLines[VC132_VARIABLE_INDEX] = (
                                'VC132=0' + format(FINISH_BORE_TOOL_NUMBER) + '  (FinishBoreToolNo)')
                    elif (FINISH_BORE_TOOL_NUMBER >= 10):
                        HorizontalProgramLines[VC132_VARIABLE_INDEX] = (
                                'VC132=' + format(FINISH_BORE_TOOL_NUMBER) + '  (FinishBoreToolNo)')

                ### LOCK RING TOOL NUUMBER VARIABLE (VC134)
                substr = "VC134"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC134_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC134_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC134_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    # WE JUST USE if STATEMENT BELOW TO MAKE THE TOOL NUMBER ALWAYS PRINTED WITH TWO DIGITS EVEN IT WAS ONE DIGIT(LIKE: 1,3,5,7)
                    # **IT IS NOT NECESSARY JUST TO KEEP TOOLS ORGANIZED AND HAVE 2 DIGITS**
                    if (LOCK_RING_TOOL_NUMBER < 10):
                        HorizontalProgramLines[VC134_VARIABLE_INDEX] = (
                                'VC134=0' + format(LOCK_RING_TOOL_NUMBER) + '  (LockRingToolNo)')
                    elif (LOCK_RING_TOOL_NUMBER >= 10):
                        HorizontalProgramLines[VC134_VARIABLE_INDEX] = (
                                'VC134=' + format(LOCK_RING_TOOL_NUMBER) + '  (LockRingToolNo)')

                ### DOUBLE_OIL_HOLE_SLOT TOOL NUUMBER VARIABLE (VC136)
                substr = "VC136"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0 and DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS == 1):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC136_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC136_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC136_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    # WE JUST USE if STATEMENT BELOW TO MAKE THE TOOL NUMBER ALWAYS PRINTED WITH TWO DIGITS EVEN IT WAS ONE DIGIT(LIKE: 1,3,5,7)
                    # **IT IS NOT NECESSARY JUST TO KEEP TOOLS ORGANIZED AND HAVE 2 DIGITS**
                    # ** WE KNOW THE TOOL NUMBER USED FOR DOUBLE_OIL_HOLE_SLOT IS (5) WHICH MEAN WE DON'T NEED THE elif STATEMENT, BUT PUT HERE IN CASE THE TOOL NUMBER CHANGE ON FUTURE
                    if (DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER < 10):
                        HorizontalProgramLines[VC136_VARIABLE_INDEX] = (
                                'VC136=0' + format(DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER) + '  (DOHSToolNo)')
                    elif (DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER >= 10):
                        HorizontalProgramLines[VC136_VARIABLE_INDEX] = (
                                'VC136=' + format(DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER) + '  (DOHSToolNo)')

                # END OF TOOLS NUMBER (*****TOOLS*****)

                # BEGINNING OF DIMENSIONAL VARIABLES (****DIMENSIONAL VARIABLES****)

                # +++++-------PIN_HOLE_DIAMETER VARIABLE (VC149)-------+++++#
                substr = "VC149"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and PIN_HOLE_DIAMETER != 0) : JUST IN CASE (PIN_HOLE_DIAMETER) IS MISSING FOR SOME REASON
                if (index == 0 and pin_hole_diameter != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC149_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC149_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC149_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC149_VARIABLE_INDEX] = (
                            'VC149=' + format(pin_hole_diameter) + '  (PinHoleDiameter)')

                # +++++-------FORGE_REF_LENGTH VARIABLE (VC150)-------+++++#
                substr = "VC150"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and FORGE_REF_LENGTH__IS_F_NUMBER_IN_EMSS != 0) : JUST IN CASE (FORGE_REF_LENGTH__IS_F_NUMBER_IN_EMSS) IS MISSING FOR SOME REASON
                if (index == 0 and FORGE_REF_LENGTH__IS_F_NUMBER_IN_EMSS != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC150_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC150_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC150_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC150_VARIABLE_INDEX] = (
                            'VC150=' + format(FORGE_REF_LENGTH__IS_F_NUMBER_IN_EMSS) + '  (ForgeRefLength)')

                # +++++-------PILOT_BORE_DEPTH VARIABLE (VC151)-------+++++#
                substr = "VC151"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and PILOT_BORE_DEPTH != 0) : JUST IN CASE (PILOT_BORE_DEPTH) IS MISSING FOR SOME REASON
                if (index == 0 and pilot_bore_depth != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC151_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC151_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC151_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC151_VARIABLE_INDEX] = (
                            'VC151=' + format(pilot_bore_depth) + '  (PilotBoreDepth)')

                # +++++-------(ROUGH_BORE_SPEED) VARIABLE (VC152)-------+++++#
                substr = "VC152"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and FINISH_BORE_TOOL_NUMBER == 28) : CHECK IF WE USE BORING BAR TOOL (TOOL NUMBER IS 28), WE NEED TO SLOW THE SPEED TO 6000, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (8000)
                if (index == 0 and FINISH_BORE_TOOL_NUMBER == 28):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC152_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC152_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC152_VARIABLE_INDEX])
                    # WE NEED TO SLOW THE SPEED TO 6000 WHEN USING THE BORING BAR TOOL
                    ROUGH_BORE_SPEED = 6000
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC152_VARIABLE_INDEX] = (
                            'VC152=' + format(ROUGH_BORE_SPEED) + '  (RoughBoreSpeed)')

                # +++++-------(ROUGH_BORE_FEED) VARIABLE (VC153)-------+++++#
                substr = "VC153"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and FINISH_BORE_TOOL_NUMBER == 28) : CHECK IF WE USE BORING BAR TOOL (TOOL NUMBER IS 28), WE NEED TO SLOW THE FEED TO 60, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (100)
                if (index == 0 and FINISH_BORE_TOOL_NUMBER == 28):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC153_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC153_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC153_VARIABLE_INDEX])
                    # WE NEED TO SLOW THE FEED TO 60 WHEN USING THE BORING BAR TOOL
                    ROUGH_BORE_FEED = 60
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC153_VARIABLE_INDEX] = (
                            'VC153=' + format(ROUGH_BORE_FEED) + '  (RoughBoreFeed)')

                # +++++-------(X_DISTANCE_FROM_ORIGIN_TO_PIN_CENTER) VARIABLE (VC154)-------+++++#
                substr = "VC154"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and FINISH_BORE_TOOL_NUMBER == 28) : CHECK IF WE USE BORING BAR TOOL (TOOL NUMBER IS 28), WE NEED TO SLOW THE FEED TO 60, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (100)
                if (index == 0 and pilot_bore_depth != 0 and pilot_to_pin != 0 and pin_hole_diameter != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC154_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC154_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC154_VARIABLE_INDEX])
                    # MATH CALCULATION OF X_DISTANCE_FROM_ORIGIN_TO_PIN_CENTER
                    # (abs(PILOT_TO_PIN)): WE USE abs HERE TO MAKE THE PILOT_TO_PIN NUMBER POSITIVE BECAUSE MOST LIKELY IT IS NEGETIVE IN QANTEL DATA BASE
                    # (, '.4f'): TO MAKE THE NUMBER HAVE JUST 4 DIGITS (0.0000)
                    # WD DEFINE (X_DISTANCE_FROM_ORIGIN_TO_PIN_CENTER) AS global VARIABLE AGAIN TO CAN USE LATER ON NOTCH LOGIC
                    global x_distance_from_origin_to_pin_center
                    x_distance_from_origin_to_pin_center = format(
                        pilot_bore_depth - abs(pilot_to_pin) - (pin_hole_diameter / 2), '.4f')
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC154_VARIABLE_INDEX] = (
                            'VC154=-' + format(x_distance_from_origin_to_pin_center) + '  (xPinCenter)')

                # +++++-------(OFFSET) VARIABLE (VC155)-------+++++#
                substr = "VC155"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):

                    # **********STILL WORK HER************
                    # just for now
                    # something should happen here
                    # CONFIRMATION_MESSAGES_LIST = ["Offset can't find, Please Enter the value"]
                    # OFFSET_AMOUNT= self.Confirmation_MDTextField.text
                    if (offset_amount == ""):
                        CONFIRMATION_MESSAGES_LIST = ["Offset Amount does NOT found, Please Enter the value :"]
                        print(CONFIRMATION_MESSAGES_LIST)
                        confirmation_horizontal_email_message_list.append(
                            "# It was need confirmation of :" + "\n" + "\n".join(CONFIRMATION_MESSAGES_LIST))
                        self.title = '[color=0066ff]Confirmation Message[/color]' + '\n' + '\n' + '[b][i][color=ffffff]' + \
                                     CONFIRMATION_MESSAGES_LIST[0] + '[/color][/i][/b]'
                        # ******************STILL WORK HERE================
                        # MAKE IT MATCH WITH HEIGHT OF (Confirmation_MDTextField) TO NOT HAVE ANY EMPTY SPACE
                        self.Dialog_BoxLayout = BoxLayout(height=30)
                        # # NEED TO BAKE HERE AND MAKE (text) ARGUMENT TAKE VARIABLE TEXT RELATED ON THE SITUATION
                        # Confirmation_MDLabel = MDLabel(text=('\n'.join(CONFIRMATION_MESSAGES_LIST)),
                        #                                theme_text_color='Custom', text_color=(1, 1, 1, 1))
                        # self.Dialog_BoxLayout.add_widget(Confirmation_MDLabel)
                        # global Confirmation_MDTextField
                        self.Confirmation_MDTextField = TextInput(hint_text="Enter Value", multiline=False,
                                                                  input_filter="float",
                                                                  background_color=[60 / 255.0, 60 / 255.0, 60 / 255.0,
                                                                                    1], foreground_color=[1, 1, 1, 1],
                                                                  size_hint_y=None,
                                                                  height=30)  # ,hint_text= "color_mode = 'accent'"   ,mode="fill",fill_color= (1, 1, 1, 1)   line_color_focus
                        self.Dialog_BoxLayout.add_widget(self.Confirmation_MDTextField)

                        Need_Confermation_to_Create_Horizontal_Program(self, self.title, self.EnterOffsetValue,
                                                                       "custom", self.Dialog_BoxLayout)
                        # OFFSET_AMOUNT = self.Confirmation_MDTextField.text
                        print("offset value before ", offset_amount)
                        return

                    # **************************************************************************************#
                    if (offset_direction == "" and offset_amount != 0):
                        CONFIRMATION_MESSAGES_LIST = ["Offset Direction does NOT found, Please Enter the value :"]
                        print(CONFIRMATION_MESSAGES_LIST)
                        confirmation_horizontal_email_message_list.append(
                            "# It was need confirmation of :" + "\n" + "\n".join(CONFIRMATION_MESSAGES_LIST))

                        # (OFFSET_AMOUNT != 0)
                        self.title = '[color=0066ff]Confirmation Message[/color] ' + '\n' + '\n' + '[b][i][color=ffffff]' + \
                                     CONFIRMATION_MESSAGES_LIST[0] + '[/color][/i][/b]'

                        self.OFFSET_To0_Option = Item(text="OFFSET To0", on_release=self.set_Direction_as_OFFSET_To0)
                        self.OFFSET_To180_Option = Item(text="OFFSET To180",
                                                        on_release=self.set_Direction_as_OFFSET_To180)
                        self.OFFSET_EACH_WAY_Option = Item(text="OFFSET EACH WAY",
                                                           on_release=self.set_Direction_as_OFFSET_EACH_WAY)
                        # self.items = [self.OFFSET_To0_Option, self.OFFSET_To180_Option,self.OFFSET_EACH_WAY_Option]
                        self.items = [self.OFFSET_To0_Option, self.OFFSET_To180_Option, self.OFFSET_EACH_WAY_Option]
                        # '[b][i][u][color=ffffff]' + CONFIRMATION_MESSAGES_LIST[0] + '[/color][/u][/i][/b]'

                        # self.Dialog_BoxLayout.add_widget(items)

                        Need_Confermation_to_Create_Horizontal_Program(self, self.title, self.ChooseOffsetDirection,
                                                                       "confirmation", self.items)
                        # OFFSET_AMOUNT = self.Confirmation_MDTextField.text
                        print("offset direction before ", offset_direction)
                        return
                    # # just for now
                    # CONFIRMATION_MESSAGES_LIST = ["Offset can't find, Please Enter the value"]
                    # # OFFSET_AMOUNT= self.Confirmation_MDTextField.text
                    # if (CONFIRMATION_MESSAGES_LIST != 0):
                    #     Need_Confermation_to_Create_Horizontal_Program(self)
                    #     # OFFSET_AMOUNT = self.Confirmation_MDTextField.text
                    #     print("offset section ", OFFSET_AMOUNT)
                    #     return

                    # ADJUST OFFSET ACCORDING THE DIRECTION
                    # global OFFSET_AMOUNT
                    if (offset_amount == 0):
                        DIRECTION = 1
                        # just to make thing work
                    elif (offset_direction == "OFFSET EACH WAY" or offset_direction == "OFFSET To0"):
                        DIRECTION = 1
                    elif (offset_direction == "OFFSET To180"):
                        DIRECTION = -1
                    else:
                        Need_Confermation_to_Create_Horizontal_Program(self, self.title, self.EnterOffsetValue,
                                                                       "custom", self.Dialog_BoxLayout)
                        return

                    # print("offset section ", OFFSET_AMOUNT)

                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    global VC155_VARIABLE_INDEX
                    VC155_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC155_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC155_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC155_VARIABLE_INDEX] = (
                            'VC155=' + format(offset_amount * DIRECTION) + '  (Offset)')

                # +++++-------(Z_VALUE_PIN_BORE_TOP) VARIABLE (VC156)-------+++++#
                substr = "VC156"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC156_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC156_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC156_VARIABLE_INDEX])
                    if (FORGING_DIAMETER_OD_AT_ROUGHER__IS_SECOND_B_NUMBER_IN_EMSS != 0):
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC156_VARIABLE_INDEX] = ('VC156=[' + format(
                            FORGING_DIAMETER_OD_AT_ROUGHER__IS_SECOND_B_NUMBER_IN_EMSS) + '/2]' + '  (zPinBoreTop - ? IS Forging Diameter)')
                    elif (FORGING_DIAMETER__IS_FIRST_B_NUMBER_IN_EMSS != 0):
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC156_VARIABLE_INDEX] = ('VC156=[' + format(
                            FORGING_DIAMETER__IS_FIRST_B_NUMBER_IN_EMSS) + '/2]' + '  (zPinBoreTop - ? IS Forging Diameter)')

                # +++++-------(Z_VALUE_ROUGH_BORE_BOTTOM) VARIABLE (VC157)-------+++++#
                # NO NEED TO DO ANYTHING HERE, IT WILL JUST USE HOW IT IS IN HORIZONTAL TEMPLATE

                # +++++-------(Z_VALUE_FINISH_BORE_BOTTOM) VARIABLE (VC158)-------+++++#
                substr = "VC158"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and FINISH_BORE_TOOL_NUMBER == 28) : CHECK IF WE USE BORING BAR TOOL (TOOL NUMBER IS 28), WE NEED TO CHANGE THE NUMBER TO 0.1, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (1.156)
                if (index == 0 and FINISH_BORE_TOOL_NUMBER == 28):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC158_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC158_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC158_VARIABLE_INDEX])
                    # WE NEED TO TO CHANGE THE NUMBER TO 0.1 WHEN USING THE BORING BAR TOOL
                    VALUE_USED_IN_Z_VALUE_FINISH_BORE_BOTTOM = 0.1
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC158_VARIABLE_INDEX] = ('VC158=-[VC156+' + format(
                        VALUE_USED_IN_Z_VALUE_FINISH_BORE_BOTTOM) + ']' + '  (zFinishBoreBottom)')

                # +++++-------(FORGING_OUTSIDE_BOSS_SPACING__IS_U_NUMBER_IN_EMSS ) VARIABLE (VC159)-------+++++#
                substr = "VC159"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                if (index == 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC159_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC159_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC159_VARIABLE_INDEX])
                    if (FORGING_OUTSIDE_BOSS_SPACING__IS_U_NUMBER_IN_EMSS != 0):
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC159_VARIABLE_INDEX] = ('VC159=' + format(
                            FORGING_OUTSIDE_BOSS_SPACING__IS_U_NUMBER_IN_EMSS) + '  (OutsideBossSpacing)')
                    elif (FORGING_OUTSIDE_BOSS_SPACING__IS_U_NUMBER_IN_EMSS == 0):
                        if (FORGING_DIAMETER_OD_AT_ROUGHER__IS_SECOND_B_NUMBER_IN_EMSS != 0):
                            # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                            HorizontalProgramLines[VC159_VARIABLE_INDEX] = ('VC159=' + format(
                                FORGING_DIAMETER_OD_AT_ROUGHER__IS_SECOND_B_NUMBER_IN_EMSS) + '  (OutsideBossSpacing)')
                        elif (FORGING_DIAMETER__IS_FIRST_B_NUMBER_IN_EMSS != 0):
                            # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                            HorizontalProgramLines[VC159_VARIABLE_INDEX] = ('VC159=' + format(
                                FORGING_DIAMETER__IS_FIRST_B_NUMBER_IN_EMSS) + '  (OutsideBossSpacing)')

                # +++++-------(LEDGE_TOOL_DIAMETER) VARIABLE (VC160)-------+++++#
                substr = "VC160"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and (LEDGE_CUT_AVAILABILITY_STATUS == 1 or LEDGE_COUNTERBORE_DIAMETER != 0)) : TO CHECK IF WE USE THE LEDGE TOOL OR NOT,
                if (index == 0 and (ledge_cut_availability_status == 1 or ledge_counterbore_diameter != 0)):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC160_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC160_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC160_VARIABLE_INDEX])
                    # WE NEED TO CHANGE DIAMETER OF LEDGE TOOL TO 0.375 (TOOL DIAMETER OF TOOL WE USE WHEN HAVE SMALL BORE) IF PIN HOLE DIAMETER SMALLER THAN THE STD LEDGE TOOL DIAMETER(0.625),
                    # OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (0.625)
                    if (pin_hole_diameter < 0.629):
                        LEDGE_TOOL_DIAMETER = 0.375
                        HorizontalProgramLines[VC160_VARIABLE_INDEX] = (
                                'VC160=' + format(LEDGE_TOOL_DIAMETER) + '  (LedgeToolDiameter)')

                # +++++-------(Z_VALUE_OF_TOP_OF_LOCKRING) VARIABLE (VC161)-------+++++#
                substr = "VC161"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and LOCK_RING_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS LOCK RING, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                if (index == 0 and lock_ring_cutter_width != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC161_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC161_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC161_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC161_VARIABLE_INDEX] = ('VC161=[' + format(
                        lock_ring_id_spacing) + '/2]' + '  (zTopLockRing = Replace 0 with Lockring ID Spacing)')

                # +++++-------(Z_VALUE_OF_BOTTOM_OF_LOCKRING) VARIABLE (VC162)-------+++++#
                substr = "VC162"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and LOCK_RING_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS LOCK RING, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                if (index == 0 and lock_ring_cutter_width != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC162_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC162_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC162_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC162_VARIABLE_INDEX] = ('VC162=-[VC161+' + format(
                        lock_ring_cutter_width) + ']' + '  (zBottomLockRing IS zTopLockRing + LR CUTTER WIDTH)')

                # +++++-------(LOCKRING_CUT_RADIUS) VARIABLE (VC163)-------+++++#
                substr = "VC163"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and LOCK_RING_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS LOCK RING, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                if (index == 0 and lock_ring_cutter_width != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC163_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC163_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC163_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC163_VARIABLE_INDEX] = (
                            'VC163=[[' + format(lock_ring_diameter) + '-' + format(
                        LOCK_RING_TOOL_DIAMETER) + ']/2]' + '  (LRCutRadius)')

                # +++++-------(Z_VALUE_OF_TOP_OF_CFREN) VARIABLE (VC164)-------+++++#
                substr = "VC164"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and CFren_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS CFREN, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                if (index == 0 and cfren_cutter_width != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC164_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC164_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC164_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC164_VARIABLE_INDEX] = ('VC164=[' + format(
                        cfren_id_spacing) + '/2]' + '  (zTopCFREN = Replace 0 with CFren ID Spacing)')

                # +++++-------(Z_VALUE_OF_BOTTOM_OF_CFREN) VARIABLE (VC165)-------+++++#
                substr = "VC165"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and CFren_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS CFren, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                if (index == 0 and cfren_cutter_width != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC165_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC165_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC165_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC165_VARIABLE_INDEX] = ('VC165=-[VC164+' + format(
                        cfren_cutter_width) + ']' + '  (zBottomCFREN IS zTopCFREN + LR CUTTER WIDTH)')

                # +++++-------(CFREN_CUT_RADIUS) VARIABLE (VC166)-------+++++#
                substr = "VC166"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and CFren_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS LOCK RING, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                if (index == 0 and cfren_cutter_width != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC166_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC166_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC166_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC166_VARIABLE_INDEX] = ('VC166=[[' + format(cfren_diameter) + '-' + format(
                        CFren_TOOL_DIAMETER) + ']/2]' + '  (CFCutRadius)')

                # +++++-------(CFREN_CUT_OFFSET) VARIABLE (VC167)-------+++++#
                # NO NEED TO DO ANYTHING HERE FOR NOW, IT WILL JUST USE HOW IT IS IN HORIZONTAL TEMPLATE

                # +++++-------(DOUBLE_OIL_HOLE_SLOT_ID_SPACING) VARIABLE (VC168)-------+++++#
                substr = "VC168"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and DOUBLE_OIL_HOLE_SLOT_ID_SPACING != 0 and PIN_HOLE_DIAMETER >= 0.901) : CHECK IF JOB HAS DOUBLE_OIL_HOLE_SLOT OR NOT AND PIN HOLE SIZE IS BIGGER THAN 0.901,
                # OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                if (index == 0 and double_oil_hole_slot_id_spacing != 0 and pin_hole_diameter >= 0.901):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC168_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC168_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC168_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC168_VARIABLE_INDEX] = (
                            'VC168=' + format(double_oil_hole_slot_id_spacing) + '  (DOHS_ID_Spacing)')

                # +++++-------NOTCH CALCULATION TO USE ON VARIABLES (VC172 AND VC173) SECTIONS LATER -------+++++#

                # NEED TO BACK LATER AND EXPLAIN HOW YOU FIGURED OUT THING BELOW
                # (and OFFSET_AMOUNT != ""): JUST TO MAKE THE CODE WORK IF CAN'T DETECT THE OFFSET FROM DATA BASE
                if (notch_angle_first_location != 0 and offset_amount != ""):
                    # ADJUST OFFSET ACCORDING THE DIRECTION
                    if (offset_amount == 0):
                        DIRECTION = 1
                    elif (offset_direction == "OFFSET EACH WAY" or offset_direction == "OFFSET To0"):
                        DIRECTION = 1
                    elif (offset_direction == "OFFSET To180"):
                        DIRECTION = -1
                    else:
                        # just to make the code work
                        DIRECTION = 1
                    #     Need_Confermation_to_Create_Horizontal_Program(self , self.title ,self.EnterOffsetValue , "custom", self.Dialog_BoxLayout)
                    #     return
                    # Need_Confermation_to_Create_Horizontal_Program(self)

                    # print("in notch section ", OFFSET_AMOUNT)

                    # IT FIGURES OUT BY DRAWING TRIANGLE IN SOLID WORK AND USE TRIGONOMETRIC MATH
                    # WE USE (math.radians(NOTCH_ANGLE)) BECAUSE WE NEED PUT ANGLE IN DEGREES
                    ## print('NOTCH_ANGLE =', NOTCH_ANGLE_FIRST_LOCATION)
                    ## print("OFFSET =", OFFSET)
                    Y_VALUE_FOR_NOTCH_MATH = format(
                        (pin_hole_diameter / 2) * (math.sin(math.radians(notch_angle_first_location))), '.4f')

                    global y_distance_from_origin_to_circlip_notch
                    y_distance_from_origin_to_circlip_notch = format(
                        float(Y_VALUE_FOR_NOTCH_MATH) + (float(offset_amount) * DIRECTION), '.4f')
                    ## print("Y_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH =", Y_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH)

                    # IT IS USE PYTHAGORAS THEOREM FORMULA (HYPOTENUSE^2=(PERPENDICULAR^2)+(BASE^2))>>>(PERPENDICULAR=sqrt((HYPOTENUSE^2)-(BASE^2))
                    X_VALUE_FOR_NOTCH_MATH = math.sqrt(
                        (math.pow((pin_hole_diameter / 2), 2)) - (math.pow(float(Y_VALUE_FOR_NOTCH_MATH), 2)))

                    if (notch_angle_first_location <= 90):
                        X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH = format(
                            float(x_distance_from_origin_to_pin_center) - float(X_VALUE_FOR_NOTCH_MATH), '.4f')
                    elif (notch_angle_first_location > 90):
                        X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH = format(
                            float(x_distance_from_origin_to_pin_center) + float(X_VALUE_FOR_NOTCH_MATH), '.4f')
                    ## print("X VALUE IS: " + X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH)

                # +++++-------(X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH) VARIABLE (VC172)-------+++++#
                substr = "VC172"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and NOTCH_ANGLE_FIRST_LOCATION != 0) : CHECK IF JOB HAS NOTCH, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                if (index == 0 and notch_angle_first_location != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC172_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC172_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC172_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC172_VARIABLE_INDEX] = (
                            'VC172=-' + format(X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH) + '  (xCirclipNotch)')

                # +++++-------(Y_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH) VARIABLE (VC173)-------+++++#
                substr = "VC173"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and NOTCH_ANGLE_FIRST_LOCATION != 0) : CHECK IF JOB HAS NOTCH, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                if (index == 0 and notch_angle_first_location != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    global VC173_VARIABLE_INDEX
                    VC173_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC173_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC173_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC173_VARIABLE_INDEX] = (
                            'VC173=' + format(y_distance_from_origin_to_circlip_notch) + '  (yCirclipNotch)')

                # +++++-------(HORIZONTAL_SLOTS_ID_SPACING) VARIABLE (VC175)-------+++++#
                substr = "VC175"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and HORIZONTAL_SLOTS_ID_SPACING != 0) : CHECK IF JOB HAS HORIZ SLOTS, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                # (and HORIZONTAL_SLOTS_ARC_DIAMETER != 0.375) : ***JUST FOR NOW*** CHECK IF JOB HAS HORIZ SLOTS WITH RADIUS NOT EQUAL 0.375,>>>
                # >>>BECAUSE HORIZ SLOTS WITH 0.375 RADIUS CONSIDER (HORIZONTAL SLOTS STRAIGHT_THROUGH), WHICH NEED TO ADJUST OTHER VARIABLES
                if (index == 0 and horizontal_slots_id_spacing != 0 and horizontal_slots_arc_diameter != 0.375):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC175_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC175_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC175_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC175_VARIABLE_INDEX] = (
                            'VC175=' + format(horizontal_slots_id_spacing) + '  (HSlot_ID_Spacing)')

                # +++++-------(i_START_HORIZONTAL_SLOT) VARIABLE (VC176)-------+++++#
                # IF WE USE STANDERD HORIZ SLOT MAYBE WE WILL MAKE IT EQUAL 1
                substr = "VC176"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and HORIZONTAL_SLOTS_ID_SPACING != 0) : CHECK IF JOB HAS HORIZ SLOTS, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                # (and HORIZONTAL_SLOTS_ARC_DIAMETER != 0.375) : ***JUST FOR NOW*** CHECK IF JOB HAS HORIZ SLOTS WITH RADIUS NOT EQUAL 0.375,>>>
                # >>>BECAUSE HORIZ SLOTS WITH 0.375 RADIUS CONSIDER (HORIZONTAL SLOTS STRAIGHT_THROUGH), WHICH NEED TO ADJUST OTHER VARIABLES
                if (index == 0 and horizontal_slots_id_spacing != 0 and horizontal_slots_arc_diameter != 0.375):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC176_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC176_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC176_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC176_VARIABLE_INDEX] = (
                            'VC176=' + format(i_START_HORIZONTAL_SLOT) + '  (iStartHSlot)')

                # +++++-------(j_START_HORIZONTAL_SLOT) VARIABLE (VC177)-------+++++#
                # IF WE USE STANDERD HORIZ SLOT MAYBE WE WILL MAKE IT EQUAL 1
                substr = "VC177"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and HORIZONTAL_SLOTS_ID_SPACING != 0) : CHECK IF JOB HAS HORIZ SLOTS, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                # (and HORIZONTAL_SLOTS_ARC_DIAMETER != 0.375) : ***JUST FOR NOW*** CHECK IF JOB HAS HORIZ SLOTS WITH RADIUS NOT EQUAL 0.375,>>>
                # >>>BECAUSE HORIZ SLOTS WITH 0.375 RADIUS CONSIDER (HORIZONTAL SLOTS STRAIGHT_THROUGH), WHICH NEED TO ADJUST OTHER VARIABLES
                if (index == 0 and horizontal_slots_id_spacing != 0 and horizontal_slots_arc_diameter != 0.375):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC177_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC177_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC177_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC177_VARIABLE_INDEX] = (
                            'VC177=' + format(j_START_HORIZONTAL_SLOT) + '  (jStartHSlot)')

                # +++++-------(HORIZONTAL_SLOT_RADIUS) VARIABLE (VC178)-------+++++#
                # IF WE USE STANDERD HORIZ SLOT MAYBE WE WILL MAKE IT EQUAL 1
                substr = "VC178"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and HORIZONTAL_SLOTS_ID_SPACING != 0) : CHECK IF JOB HAS HORIZ SLOTS, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                # (and HORIZONTAL_SLOTS_ARC_DIAMETER != 0.375) : ***JUST FOR NOW*** CHECK IF JOB HAS HORIZ SLOTS WITH RADIUS NOT EQUAL 0.375,>>>
                # >>>BECAUSE HORIZ SLOTS WITH 0.375 RADIUS CONSIDER (HORIZONTAL SLOTS STRAIGHT_THROUGH), WHICH NEED TO ADJUST OTHER VARIABLES
                if (index == 0 and horizontal_slots_id_spacing != 0 and horizontal_slots_arc_diameter != 0.375):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC178_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC178_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC178_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC178_VARIABLE_INDEX] = (
                            'VC178=' + format(HORIZONTAL_SLOT_RADIUS) + '  (HSlotRadius)')

                # +++++-------(LEDGE_COUNTERBORE_DIAMETER) VARIABLE (VC179)-------+++++#
                substr = "VC179"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and LEDGE_COUNTERBORE_DIAMETER != 0) : CHECK IF JOB HAS LEDGE_COUNTERBORE, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                if (index == 0 and ledge_counterbore_diameter != 0):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC179_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC179_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC179_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC179_VARIABLE_INDEX] = (
                            'VC179=' + format(ledge_counterbore_diameter) + '  (CounterboreLedgeDiameter)')

                # +++++-------MAYBE NEED TO ADD VARIABLE OF DISTANCE OF LEDGE_COUNTERBORE TO LOCKRING TO HORIZONTAL TEMPLATE-------+++++#

                # +++++-------(X_DISTANCE_375_SLOTS_FROM_CENTER_OF_BORE_TO_CENTER_OF_HORIZONTAL_SLOT) VARIABLE (VC183)-------+++++#

                # NEED TO FIGURE OUT THE MATH WE WANT TO USE HERE

                # +++++-------(Y_DISTANCE_375_SLOTS_FROM_CENTER_OF_BORE_TO_CENTER_OF_HORIZONTAL_SLOT) VARIABLE (VC184)-------+++++#

                # NEED TO FIGURE OUT THE MATH WE WANT TO USE HERE

                # +++++-------(FORGING_OR_FINISHED_BOSS_WIDTH_FOR_375_SLOTS) VARIABLE (VC185)-------+++++#
                substr = "VC185"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and HORIZONTAL_SLOTS_ARC_DIAMETER == 0.375) : CHECK IF JOB HAS (HORIZONTAL SLOTS STRAIGHT_THROUGH), OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                if (index == 0 and horizontal_slots_arc_diameter == 0.375):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    VC185_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                    print(VC185_VARIABLE_INDEX)
                    print(HorizontalProgramLines[VC185_VARIABLE_INDEX])
                    # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                    HorizontalProgramLines[VC185_VARIABLE_INDEX] = ('VC185=' + format(
                        FORGING_INSIDE_BOSS_SPACING__IS_J_NUMBER_IN_EMSS) + '  (ForgingOrFinishedBossWidthFor375Slots)')

                # +++++-------(CALL PROBE PROGRAMS)-------+++++#
                # # self.manager.get_screen('SettingScreen'): SCREEN WE WANT TO ACCESS
                # # ids["ProbePrograms"].text : TO GET THE TEXT OF ["ProbePrograms"] THAT'S CONTAIN PROBE PROGRAMS FILE PATH TO SEARCH INSIDE THE FILE
                # PROBE_PROGRAMS_FILE_PATH = self.manager.get_screen('SettingScreen').ids["ProbePrograms"].text
                # # DEFINE AN EMPTY LIST TO STORE RESULT OF PROBE PROGRAM SEARCH.
                # RESULT_OF_PROBE_PROGRAM_SEARCH = []
                # for file in glob.glob(PROBE_PROGRAMS_FILE_PATH + '*\*' + FORGING_NUMBER + '*'):
                #     # print(file)
                #     RESULT_OF_PROBE_PROGRAM_SEARCH.append(file)
                #     # print(RESULT_OF_PROBE_PROGRAM_SEARCH)
                #     # print(len(RESULT_OF_PROBE_PROGRAM_SEARCH))
                # # DECLARE AN EMPTY LIST TO ADD PROBE PROGRAM LINES ONE BY ONE.
                # PROBE_PROGRAMS_LINES = []
                # if ((len(RESULT_OF_PROBE_PROGRAM_SEARCH) == 1) and (FORGING_NUMBER is not None)):
                #     # print(RESULT_OF_PROBE_PROGRAM_SEARCH[0])
                #     with open(RESULT_OF_PROBE_PROGRAM_SEARCH[0], 'rt') as CurrentProgram:
                #         for line in CurrentProgram:  # For each line in the file,
                #             PROBE_PROGRAMS_LINES.append(line.rstrip('\n'))  # strip newline and add to list.
                #         # print(PROBE_PROGRAMS_LINES)
                #         # print("LINE OF PROBE PROGRAM THAT NEED TO ADD TO HORIZONTAL PROGRAM: " + PROBE_PROGRAMS_LINES[0])
                #         PROBE_PROGRAM = PROBE_PROGRAMS_LINES[0]
                # elif ((len(RESULT_OF_PROBE_PROGRAM_SEARCH) == 0) and (FORGING_NUMBER is not None)):
                #     print("NO PROBE PROGRAM FOUND, MAKE ONE AND CLICK SUBMIT BUTTON AGAIN ")
                # elif ((len(RESULT_OF_PROBE_PROGRAM_SEARCH) > 1) and (FORGING_NUMBER is not None)):
                #     # NEED TO WORK ON THAT
                #     print("MORE RESULTS FOUND , CHOOSE THE RIGHT PROBE PROGRAM")
                #     print(RESULT_OF_PROBE_PROGRAM_SEARCH)
                #     print(len(RESULT_OF_PROBE_PROGRAM_SEARCH))
                # else:
                #     print("SOMETHING UNEXPECTED HAPPEN, SEE PROGRAMMING ")

                substr = "CALL OXXXX"  # VARIABLE WE LOOKING FOR
                # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                index = line.find(substr)
                # print(index)
                # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                # (and HORIZONTAL_SLOTS_ARC_DIAMETER == 0.375) : CHECK IF JOB HAS (HORIZONTAL SLOTS STRAIGHT_THROUGH), OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                if (index == 0 and FORGING_NUMBER is not None):
                    # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                    # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                    PROBE_PROGRAM_INDEX = HorizontalProgramLines.index(line)
                    print(PROBE_PROGRAM_INDEX)
                    print(HorizontalProgramLines[PROBE_PROGRAM_INDEX])
                    HorizontalProgramLines[PROBE_PROGRAM_INDEX] = ('CALL ' + format(PROBE_PROGRAM))

            # *****************STILL WORK HERE****************
            # JUST FOR TESTING
        ##FAILED_MESSAGES_LIST.append(self.ids["JobNumber"].text +": Failed to Create Horizontal Program.")

        # CONFIRMATION MESSAGE NEED TO BE JUST CALL WHEN NEED IT(ie: NEED TO GO THROUGH ALL VARIABLE AND CALL THE FUNCTION TO FIX IT RIGHT WAY) ,
        # PROBABLY WE DON'T NEED IT AS LIST
        # CONFIRMATION_MESSAGES_LIST.append(self.ids["JobNumber"].text + ": Confirmation Test.")
        # CONFIRMATION_MESSAGES_LIST.append("Can't find Offset Value, Please Enter")
        # SUCCESSFUL_MESSAGES_LIST.append(self.ids["JobNumber"].text +": Horizontal Program Has been Created Successfully.")

        # Need_Confermation_to_Create_Horizontal_Program(self)                                *****  START TOMOWROW FROM HERE.....and self.ids["JobNumber"].text != ""    *****
        # NEED TO BACK AND FIX THIS CONDISION
        if (FAILED_MESSAGES_LIST == [] and CONFIRMATION_MESSAGES_LIST == [] and self.ids["JobNumber"].text != ""):
            # just for now
            print("actual")
            print(HorizontalProgramLines)

            Horizontal_Program_Has_been_Created_Successfully_Original_Folder(self)
            # Horizontal_Program_Has_been_Created_Successfully_Running_Folder(self)
            # JUST FOR TESTING
            # print()
            # print('\n'.join(HorizontalProgramLines))
        elif (CONFIRMATION_MESSAGES_LIST != []):
            print("CONFIRMATION_MESSAGES_LIST STILL HAVE SOMTHING: ", CONFIRMATION_MESSAGES_LIST)


        elif (self.ids["JobNumber"].text == ""):
            FAILED_MESSAGES_LIST.append("Please Enter Job Number.")
            # horizontal_email_message_list.append("\n".join(FAILED_MESSAGES_LIST))
            Failed_to_Create_Horizontal_Program(self)

        #     Need_Confermation_to_Create_Horizontal_Program(self)                  "\n".join(FAILED_MESSAGES_LIST)
        else:
            Failed_to_Create_Horizontal_Program(self)

            # STILL NEED TO ADD DIALOG TO SHOW MESSAGE OF CREATE PROGRAM

    # def ResetFields(self):
    #     # TO RESET JOB NUMBER FIELD TO START OVER
    #     self.ids["JobNumber"].text = ""

    # def Horizontal_Program_Has_been_Created_Successfully(self):
    #     # JUST FOR TESTING
    #     print()
    #     print('\n'.join(HorizontalProgramLines))
    #     Close_Button = MDRaisedButton(text='Close', on_release=self.CloseDialog)
    #     self.Warning_Dialog = MDDialog(title='Message:',text=(self.ids["JobNumber"].text +": Horizontal Program Has been Created Successfully."),size_hint=(0.7, 1.0), buttons=[Close_Button])
    #     # TO OPEN THE DIALOG WINDOW
    #     self.Warning_Dialog.open()

    def Need_Ledge_Tool(self, obj):
        print("Need_Ledge_Tool function")
        self.Need_Ledge_Tool_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
        self.Does_Not_Need_Ledge_Tool_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.Need_Ledge_Tool_Status = True
        self.Does_Not_Need_Ledge_Tool_Status = False

    def Does_Not_Need_Ledge_Tool(self, obj):
        print("Does_Not_Need_Ledge_Tool function")
        self.Does_Not_Need_Ledge_Tool_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
        self.Need_Ledge_Tool_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.Does_Not_Need_Ledge_Tool_Status = True
        self.Need_Ledge_Tool_Status = False

    def Decide_Ledge_Tool_Status(self, obj):
        print("Decide_Ledge_Tool_Status function")
        global ledge_cut_availability_status
        if (self.Need_Ledge_Tool_Status == True):
            ledge_cut_availability_status = 1
        elif (self.Does_Not_Need_Ledge_Tool_Status == True):
            ledge_cut_availability_status = 0
        else:
            print("Ladge Tool Status get wrong")
        print("LEDGE_CUT_AVAILABILITY_STATUS in decide function ", ledge_cut_availability_status)
        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()

    def EnterOffsetValue(self, obj):
        global offset_amount
        offset_amount = float(self.Confirmation_MDTextField.text)
        CONFIRMATION_MESSAGES_LIST = []
        print("offset value after ", offset_amount)
        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()

        # ************** NEED TO FIX COLORS ISSUE

    def set_Direction_as_OFFSET_To0(self, obj):
        print("set_Direction_as_OFFSET_To0")
        self.OFFSET_To0_Option.bg_color = (
            20 / 255, 82 / 255, 20 / 255,
            1)  # '[b][i][color=33cc33]' + self.OFFSET_To0_Option.text + '[/color][/i][/b]'
        self.OFFSET_To180_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.OFFSET_EACH_WAY_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.OFFSET_To0_Option_Status = True
        print("To0 ", self.OFFSET_To0_Option_Status)
        # self.OFFSET_To180_Option.text = '[color=ffffff]' + self.OFFSET_To180_Option.text + '[/color]'
        self.OFFSET_To180_Option_Status = False
        print("To180 ", self.OFFSET_To180_Option_Status)
        # self.OFFSET_EACH_WAY_Option.text = '[color=ffffff]' + self.OFFSET_EACH_WAY_Option.text + '[/color]'
        self.OFFSET_EACH_WAY_Option_Status = False
        print("EACH_WAY ", self.OFFSET_EACH_WAY_Option_Status)
        print(self.OFFSET_To0_Option.text)
        print()

    def set_Direction_as_OFFSET_To180(self, obj):
        print("set_Direction_as_OFFSET_To180")
        self.OFFSET_To180_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
        self.OFFSET_To0_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.OFFSET_EACH_WAY_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)

        self.OFFSET_To180_Option_Status = True
        print("To180 ", self.OFFSET_To180_Option_Status)
        # self.OFFSET_To0_Option.text = '[color=ffffff]' + self.OFFSET_To0_Option.text + '[/color]'
        self.OFFSET_To0_Option_Status = False
        print("To0 ", self.OFFSET_To0_Option_Status)
        # self.OFFSET_EACH_WAY_Option.text = '[color=ffffff]' + self.OFFSET_EACH_WAY_Option.text + '[/color]'
        self.OFFSET_EACH_WAY_Option_Status = False
        print("EACH_WAY ", self.OFFSET_EACH_WAY_Option_Status)
        print(self.OFFSET_To180_Option.text)
        print()

    def set_Direction_as_OFFSET_EACH_WAY(self, obj):
        print("set_Direction_as_OFFSET_EACH_WAY")

        self.OFFSET_EACH_WAY_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
        self.OFFSET_To0_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.OFFSET_To180_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)

        self.OFFSET_EACH_WAY_Option_Status = True
        print("EACH_WAY ", self.OFFSET_EACH_WAY_Option_Status)
        # self.OFFSET_To180_Option.text = '[color=ffffff]' + self.OFFSET_To180_Option.text + '[/color]'
        self.OFFSET_To180_Option_Status = False
        print("To180 ", self.OFFSET_To180_Option_Status)
        # self.OFFSET_To0_Option.text = '[color=ffffff]' + self.OFFSET_To0_Option.text + '[/color]'
        self.OFFSET_To0_Option_Status = False
        print("To0 ", self.OFFSET_To0_Option_Status)
        print(self.OFFSET_EACH_WAY_Option.text)

        # **********still work here************

    def ChooseOffsetDirection(self, obj):
        global offset_direction
        if (self.OFFSET_To0_Option_Status == True):
            offset_direction = self.OFFSET_To0_Option.text
        elif (self.OFFSET_To180_Option_Status == True):
            offset_direction = self.OFFSET_To180_Option.text
        elif (self.OFFSET_EACH_WAY_Option_Status == True):
            offset_direction = self.OFFSET_EACH_WAY_Option.text
        else:
            print("OFFSET_DIRECTION get wrong")
        print("offset Direction after ", offset_direction)
        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()

    # def Horizontal_Program_Has_been_Created_Successfully_Original_Folder(self):
    #     pass

    # NEED TO PUT (obj) AS PARAMETER TO MAKE FUNCTION WORK
    def SAVE_OVER_EXISTING_PROGRAM_ORIGINAL_FOLDER(self, obj):
        # NEED TO PUT THIS ONE HERE TO MAKE SURE YOU CLOSE ALL THE DIALOG WINDOW WERE OPEN
        self.Old_Horizontal_Message_Dialog.dismiss()
        print("SAVE_OVER_EXISTING_PROGRAM_ORIGINAL_FOLDER" + " is called")
        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):

            # TO 0
            # JUST MAKE IT global HERE AGAIN TO CAN USE IT IN BELOW FUNCTIONS
            global NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0
            NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0 = (
                    NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO0.MIN")
            HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER[0] = (
                    '(PART ' + NEW_PROGRAM_NUMBER + " TO 0" + ' -- ' + todaydate + ' <SYS>' + ')')
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0 = open(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0, "w")
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0.write(
                '\n'.join(HORIZONTAL_PROGRAM_LINES_To0_ORIGINAL_FOLDER))
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO0.close()

            # TO 180
            # JUST MAKE IT global HERE AGAIN TO CAN USE IT IN BELOW FUNCTIONS
            global NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180
            NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180 = (
                    NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO180.MIN")
            HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER[0] = (
                    '(PART ' + NEW_PROGRAM_NUMBER + " TO 180" + ' -- ' + todaydate + ' <SYS>' + ')')
            HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER[VC155_VARIABLE_INDEX] = (
                    'VC155=' + format((-1) * offset_amount) + '  (Offset)')
            HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER[VC173_VARIABLE_INDEX] = (
                    'VC173=' + format(float(y_distance_from_origin_to_circlip_notch) - (2 * offset_amount),
                                      '.4f') + '  (yCirclipNotch)')
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180 = open(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180,
                                                                       "w")
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180.write(
                '\n'.join(HORIZONTAL_PROGRAM_LINES_To180_ORIGINAL_FOLDER))
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_TO180.close()

        else:
            NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER = (
                    NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + ".MIN")
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER = open(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER, "w")
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER.write('\n'.join(HorizontalProgramLines))
            CREATE_NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER.close()  # '[color=ffffff] Successfully in [/color]'
        # '[b][i][u][color=ffffff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] Program Has been '[/color] + '[b][color=ffffff]REPLACED[/color][/b]' + '[color=ffffff] Successfully in [/color]' + '[color=ffff00]ORIGINAL[/color]' + '[color=ffffff] Folder.[/color]'

        Close_Button = MDRaisedButton(text='Close', on_release=self.Close_Old_Horizontal_Dialog_Original_Folder,
                                      font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
                '[b][i][u][color=0099ff]' + self.ids[
            "JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] Program Has been [/color]' + '[b][color=ffffff]REPLACED[/color][/b]' + '[color=ffffff] Successfully in [/color]' + '[color=ffff00]ORIGINAL[/color]' + '[color=ffffff] Folder.[/color]')
                                                      , size_hint=(0.7, 1.0), buttons=[Close_Button],
                                                      auto_dismiss=False)

        if (confirmation_horizontal_email_message_list != []):
            horizontal_email_message_list.append("\n".join(confirmation_horizontal_email_message_list) + "\n")
        SUCCESSFUL_MESSAGES_LIST = ["Program has been REPLACED successfully in ORIGINAL Folder." + "\n"]
        # maybe need logic to send this message to trello in case there is warnning make this program
        horizontal_email_message_list.append("\n".join(SUCCESSFUL_MESSAGES_LIST))

        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        print()
        # print('\n'.join(HorizontalProgramLines))

    def SAVE_OVER_EXISTING_PROGRAM_RUNNING_FOLDER(self, obj):
        # NEED TO PUT THIS ONE HERE TO MAKE SURE YOU CLOSE ALL THE DIALOG WINDOW WERE OPEN
        self.Old_Horizontal_Message_Dialog.dismiss()
        print("SAVE_OVER_EXISTING_PROGRAM_RUNNING_FOLDER" + " is called")
        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):
            # TO 0
            # JUST MAKE IT global HERE AGAIN TO CAN USE IT IN BELOW FUNCTIONS
            global NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0
            NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0 = (
                    NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO0.MIN")
            HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER[0] = (
                    '(PART ' + NEW_PROGRAM_NUMBER + " TO 0" + ' -- ' + todaydate + ' <SYS>' + ')')
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0 = open(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0, "w")
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0.write(
                '\n'.join(HORIZONTAL_PROGRAM_LINES_To0_RUNNING_FOLDER))
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0.close()

            # TO 180
            # JUST MAKE IT global HERE AGAIN TO CAN USE IT IN BELOW FUNCTIONS
            global NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180
            NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180 = (
                    NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO180.MIN")
            HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER[0] = (
                    '(PART ' + NEW_PROGRAM_NUMBER + " TO 180" + ' -- ' + todaydate + ' <SYS>' + ')')
            HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER[VC155_VARIABLE_INDEX] = (
                    'VC155=' + format((-1) * offset_amount) + '  (Offset)')
            HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER[VC173_VARIABLE_INDEX] = (
                    'VC173=' + format(float(y_distance_from_origin_to_circlip_notch) - (2 * offset_amount),
                                      '.4f') + '  (yCirclipNotch)')
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180 = open(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180, "w")
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180.write(
                '\n'.join(HORIZONTAL_PROGRAM_LINES_To180_RUNNING_FOLDER))
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180.close()


        else:
            NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER = (
                    NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + ".MIN")
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER = open(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER, "w")
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER.write('\n'.join(HorizontalProgramLines))
            CREATE_NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER.close()
        # '[b][i][u][color=ffffff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] Program Has been [/color]' + '[b][color=ffffff]REPLACED[/color][/b]' + '[color=ffffff] Successfully in [/color]' + '[color=33cc33]RUNNING[/color]' + '[color=ffffff] Folder.[/color]'
        Close_Button = MDRaisedButton(text='Close', on_release=self.Close_Old_Horizontal_Dialog_Running_Folder,
                                      font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
                '[b][i][u][color=0099ff]' + self.ids[
            "JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] Program Has been [/color]' + '[b][color=ffffff]REPLACED[/color][/b]' + '[color=ffffff] Successfully in [/color]' + '[color=33cc33]RUNNING[/color]' + '[color=ffffff] Folder.[/color]'
                + '\n' + '[color=ffffff]After closing this window, the program will open on CIMCO Editor.[/color]' + '\n' + '[color=ffffff]Please Double Check and Report any Problem on Trello.[/color]')
                                                      , size_hint=(0.7, 1.0), buttons=[Close_Button],
                                                      auto_dismiss=False)

        SUCCESSFUL_MESSAGES_LIST = ["Program has been REPLACED successfully in RUNNING Folder." + "\n"]
        # maybe need logic to send this message to trello in case there is warnning make this program
        horizontal_email_message_list.append("\n".join(SUCCESSFUL_MESSAGES_LIST))
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        print()
        # print('\n'.join(HorizontalProgramLines))

    def Close_Old_Horizontal_Dialog_Original_Folder(self, obj):
        print("Close_Old_Horizontal_Dialog_Original_Folder" + " is called")

        global result_of_existing_programs
        # LIST TO STORE THE FOUND JOBS WITH WHOLE PATH INCLUDING FILE NAME
        result_of_existing_programs = []
        for file in glob.glob(NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + '\*' + NEW_PROGRAM_NUMBER + '*'):
            # TO APPEND(ADD) EACH PROGRAM THAT FOUND TO THE RESULT LIST
            result_of_existing_programs.append(file)
        print("result_of_existing_programs: ", result_of_existing_programs)

        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()

        if ((((
                      NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO0.MIN") in result_of_existing_programs) or
             ((
                      NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO180.MIN") in result_of_existing_programs)) and
                ((
                         NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + ".MIN") in result_of_existing_programs)):
            print("job saved as To0 & To180")
            WARNING_MESSAGES_LIST.append('[b][i][u][color=0099ff]' + self.ids[
                "JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] job saved in several ways :[/color]' + '\n' + '\n' +
                                         '[color=ffffff]' + ('\n'.join(
                result_of_existing_programs)) + '[/color]' + '\n' + '\n' + '[color=ffffff]Double check in [/color]' + '[color=ffff00]ORIGINAL[/color]' +
                                         '[color=ffffff] and [/color]' + '[color=33cc33]RUNNING[/color]' + '[color=ffffff] Folders , and[/color]' + '\n' + '[color=ffffff]DELETE the Wrong Program.[/color]')

            horizontal_email_message_list.append("Job saved in several ways" + '\n' + (
                '\n'.join(result_of_existing_programs)) + '\n' + "Double check in ORIGINAL and RUNNING Folders" +
                                                 '\n' + "and DELETE the Wrong Program.")
            horizontal_Program_needs_Attention(self)
            return

        # # TO CLOSE THE DIALOGE
        # self.Old_Horizontal_Message_Dialog.dismiss()
        # just for now
        # if (confirmation_horizontal_email_message_list != []):
        #     horizontal_email_message_list.append("\n".join(confirmation_horizontal_email_message_list)+"\n")
        # SUCCESSFUL_MESSAGES_LIST = ["Program has been created successfully in ORIGINAL Folder." + "\n"]
        # # maybe need logic to send this message to trello in case there is warnning make this program
        # horizontal_email_message_list.append("\n".join(SUCCESSFUL_MESSAGES_LIST))
        # TO CALL FUNCTION OF HORIZONTAL PROGRAM IN RUNNING FOLDER
        # if self.ids["JobNumber"].text != "":
        Horizontal_Program_Has_been_Created_Successfully_Running_Folder(self)
        # TO RESET JOB NUMBER FIELD TO START OVER
        # self.ids["JobNumber"].text = ""

        # ---------------------still work here--------------------

    def Close_Old_Horizontal_Dialog_Running_Folder(self, obj):
        print("Close_Old_Horizontal_Dialog_Running_Folder" + " is called")

        # global result_of_existing_programs
        # # LIST TO STORE THE FOUND JOBS WITH WHOLE PATH INCLUDING FILE NAME
        # result_of_existing_programs = []
        # for file in glob.glob(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + '\*' + NEW_PROGRAM_NUMBER + '*'):
        #     # TO APPEND(ADD) EACH PROGRAM THAT FOUND TO THE RESULT LIST
        #     result_of_existing_programs.append(file)
        # print("result_of_existing_programs: ", result_of_existing_programs)
        # print(horizontal_email_message_list)

        # print(NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180)
        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()
        # TO OPEN FILE ON WINDOWS USE CODE BELOW
        # subprocess.Popen([ApplicationName, FileName])
        # (subprocess): BUILT IN FUNCTION USED TO OPEN FILES
        # (Popen): USED TO OPEN THE FILE
        # (ApplicationName): KIND OF APP THAT USED TO OPEN THE FILE (LIKE NOTEPAD,EXCEL,WORD...ETC), CIMCO Editor IN OUR CASE
        # (FileName): PATH OF FILE YOU WANT TO OPEN
        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):
            subprocess.Popen([cimco_editor_path, NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO0])
            subprocess.Popen([cimco_editor_path, (NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_TO180)])
        else:
            subprocess.Popen([cimco_editor_path, NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER])

        # if ((((NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO0.MIN") in result_of_existing_programs) or
        #      ((NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO180.MIN") in result_of_existing_programs)) and
        #         ((NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + ".MIN")  in result_of_existing_programs)):
        #     print("job saved as To0 & To180")
        #     WARNING_MESSAGES_LIST.append('[b][i][u][color=0099ff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] job saved in several ways :[/color]'  + '\n' + '\n' +
        #                             '[color=ffffff]' + ('\n'.join(result_of_existing_programs)) + '[/color]' + '\n' + '\n' +  '[color=ffffff]Double check in [/color]' + '[color=ffff00]ORIGINAL[/color]' +
        #                                  '[color=ffffff] and [/color]' +  '[color=33cc33]RUNNING[/color]' + '[color=ffffff] Folders , and[/color]' + '\n' + '[color=ffffff]DELETE the Wrong Program.[/color]')
        #
        #     horizontal_Program_needs_Attention(self)

        # Close_Button = MDRaisedButton(text='Close',on_release=self.Close_Old_Horizontal_Dialog,font_size = 16)  # Close_Old_Horizontal_Dialog_      '[color=ffffff] Folder.[/color]'
        # self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=cccc00]Warning Message[/color]', text=(
        #                         '[b][i][u][color=0099ff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] job saved in several ways :[/color]'  + '\n' + '\n' +
        #                         '[color=ffffff]' + ('\n'.join(result_of_existing_programs)) + '[/color]' + '\n' + '\n' +  '[color=ffffff]Double check in [/color]' + '[color=ffff00]ORIGINAL[/color]' + '[color=ffffff] and [/color]' +  '[color=33cc33]RUNNING[/color]' + '[color=ffffff] Folders , and[/color]' + '\n' + '[color=ffffff]DELETE the Wrong Program.[/color]')
        #                                                           , size_hint=(0.7, 1.0), buttons=[Close_Button],auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        # self.Old_Horizontal_Message_Dialog.open()

        # maybe need to add logic of warnning and need manual stuff
        # SUCCESSFUL_MESSAGES_LIST = ["Program has been created successfully in RUNNING Folder."]
        # horizontal_email_message_list.append("\n".join(SUCCESSFUL_MESSAGES_LIST))

        send_email(self)

        # TO ADD MESSAGE ON TRELLO
        print("\n".join(horizontal_email_message_list))

        # TO RESET JOB NUMBER FIELD TO START OVER
        self.ids["JobNumber"].text = ""
        # Reset_Variables
        four_cycle_pinbore_variables()

    # def EnterOffsetValue(self,obj):
    #     global OFFSET_AMOUNT
    #     OFFSET_AMOUNT = self.Confirmation_MDTextField.text
    #     global CONFIRMATION_MESSAGES_LIST
    #     CONFIRMATION_MESSAGES_LIST = []
    #     print("offset value after ", OFFSET_AMOUNT)
    #     return OFFSET_AMOUNT

    def Close_Old_Horizontal_Dialog(self, obj):
        print("Close_Old_Horizontal_Dialog" + " is called")
        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()
        # TO ADD MESSAGE ON TRELLO
        if (self.ids["JobNumber"].text == ""):
            pass
        elif (FAILED_MESSAGES_LIST != []):
            print("\n".join(horizontal_email_message_list))
            send_email(self)
        elif (WARNING_MESSAGES_LIST != []):
            print("\n".join(horizontal_email_message_list))
            subprocess.Popen([cimco_editor_path, NEW_HORIZONTAL_PROGRAM_ORIGINAL_FOLDER])
            send_email(self)

        # TO RESET JOB NUMBER FIELD TO START OVER
        self.ids["JobNumber"].text = ""
        # Reset_Variables
        four_cycle_pinbore_variables()

    def ResetFields(self):
        # TO RESET JOB NUMBER FIELD TO START OVER
        self.ids["JobNumber"].text = ""

        # TO CALL FUNCTION TO START OVER
        # load_horizontal_sheets_for_automation(self)
        # self.manager.current = 'OldHorizontalScreen'
        # self.Old_Horizontal_Message_Dialog.dismiss()

    # def Need_Ledge_Tool(self, obj):
    #     print("Need_Ledge_Tool function")
    #     self.Need_Ledge_Tool_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
    #     self.Does_Not_Need_Ledge_Tool_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
    #     self.Need_Ledge_Tool_Status = True
    #     self.Does_Not_Need_Ledge_Tool_Status = False
    #
    # def Does_Not_Need_Ledge_Tool(self, obj):
    #     print("Does_Not_Need_Ledge_Tool function")
    #     self.Does_Not_Need_Ledge_Tool_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
    #     self.Need_Ledge_Tool_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
    #     self.Does_Not_Need_Ledge_Tool_Status = True
    #     self.Need_Ledge_Tool_Status = False
    #
    #
    # def Decide_Ledge_Tool_Status(self, obj):
    #     print("Decide_Ledge_Tool_Status function")
    #     global LEDGE_CUT_AVAILABILITY_STATUS
    #     if (self.Need_Ledge_Tool_Status == True):
    #         LEDGE_CUT_AVAILABILITY_STATUS = 1
    #     elif (self.Does_Not_Need_Ledge_Tool_Status == True):
    #         LEDGE_CUT_AVAILABILITY_STATUS = 0
    #     else:
    #         print("Ladge Tool Status get wrong")
    #     print("LEDGE_CUT_AVAILABILITY_STATUS in decide function ", LEDGE_CUT_AVAILABILITY_STATUS)
    #     # TO CLOSE THE DIALOGE
    #     self.Old_Horizontal_Message_Dialog.dismiss()
    #
    # def EnterOffsetValue(self,obj):
    #     global OFFSET_AMOUNT
    #     OFFSET_AMOUNT = float(self.Confirmation_MDTextField.text)
    #     CONFIRMATION_MESSAGES_LIST = []
    #     print("offset value after ", OFFSET_AMOUNT)
    #     # TO CLOSE THE DIALOGE
    #     self.Old_Horizontal_Message_Dialog.dismiss()
    #
    #
    #     #************** NEED TO FIX COLORS ISSUE
    # def set_Direction_as_OFFSET_To0(self, obj):
    #     print("set_Direction_as_OFFSET_To0")
    #     self.OFFSET_To0_Option.bg_color = (20/255,82/255,20/255,1)                             #'[b][i][color=33cc33]' + self.OFFSET_To0_Option.text + '[/color][/i][/b]'
    #     self.OFFSET_To180_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
    #     self.OFFSET_EACH_WAY_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
    #     self.OFFSET_To0_Option_Status = True
    #     print("To0 ",self.OFFSET_To0_Option_Status)
    #     # self.OFFSET_To180_Option.text = '[color=ffffff]' + self.OFFSET_To180_Option.text + '[/color]'
    #     self.OFFSET_To180_Option_Status = False
    #     print("To180 ",self.OFFSET_To180_Option_Status)
    #     # self.OFFSET_EACH_WAY_Option.text = '[color=ffffff]' + self.OFFSET_EACH_WAY_Option.text + '[/color]'
    #     self.OFFSET_EACH_WAY_Option_Status = False
    #     print("EACH_WAY ",self.OFFSET_EACH_WAY_Option_Status)
    #     print(self.OFFSET_To0_Option.text)
    #     print()
    #
    # def set_Direction_as_OFFSET_To180(self, obj):
    #     print("set_Direction_as_OFFSET_To180")
    #     self.OFFSET_To180_Option.bg_color = (20/255,82/255,20/255,1)
    #     self.OFFSET_To0_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
    #     self.OFFSET_EACH_WAY_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
    #
    #     self.OFFSET_To180_Option_Status = True
    #     print("To180 ",self.OFFSET_To180_Option_Status)
    #     # self.OFFSET_To0_Option.text = '[color=ffffff]' + self.OFFSET_To0_Option.text + '[/color]'
    #     self.OFFSET_To0_Option_Status = False
    #     print("To0 ",self.OFFSET_To0_Option_Status)
    #     # self.OFFSET_EACH_WAY_Option.text = '[color=ffffff]' + self.OFFSET_EACH_WAY_Option.text + '[/color]'
    #     self.OFFSET_EACH_WAY_Option_Status = False
    #     print("EACH_WAY ",self.OFFSET_EACH_WAY_Option_Status)
    #     print(self.OFFSET_To180_Option.text)
    #     print()
    #
    # def set_Direction_as_OFFSET_EACH_WAY(self, obj):
    #     print("set_Direction_as_OFFSET_EACH_WAY")
    #
    #     self.OFFSET_EACH_WAY_Option.bg_color = (20/255,82/255,20/255,1)
    #     self.OFFSET_To0_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
    #     self.OFFSET_To180_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
    #
    #     self.OFFSET_EACH_WAY_Option_Status = True
    #     print("EACH_WAY ",self.OFFSET_EACH_WAY_Option_Status)
    #     # self.OFFSET_To180_Option.text = '[color=ffffff]' + self.OFFSET_To180_Option.text + '[/color]'
    #     self.OFFSET_To180_Option_Status = False
    #     print("To180 ",self.OFFSET_To180_Option_Status)
    #     # self.OFFSET_To0_Option.text = '[color=ffffff]' + self.OFFSET_To0_Option.text + '[/color]'
    #     self.OFFSET_To0_Option_Status = False
    #     print("To0 ",self.OFFSET_To0_Option_Status)
    #     print(self.OFFSET_EACH_WAY_Option.text)
    #
    #
    #     # **********still work here************
    # def ChooseOffsetDirection(self, obj):
    #     global OFFSET_DIRECTION
    #     if (self.OFFSET_To0_Option_Status == True):
    #         OFFSET_DIRECTION = self.OFFSET_To0_Option.text
    #     elif (self.OFFSET_To180_Option_Status == True):
    #         OFFSET_DIRECTION = self.OFFSET_To180_Option.text
    #     elif (self.OFFSET_EACH_WAY_Option_Status == True):
    #         OFFSET_DIRECTION = self.OFFSET_EACH_WAY_Option.text
    #     else:
    #         print("OFFSET_DIRECTION get wrong")
    #     print("offset Direction after ", OFFSET_DIRECTION)
    #     # TO CLOSE THE DIALOGE
    #     self.Old_Horizontal_Message_Dialog.dismiss()


# ****************Work here******************
# OldHorizontalScreen().EnterOffsetValue(object)


class NewHorizontalScreen(Screen):
    # self.ids["JobNumber"].text: TO ACCESS TEXT FIELD WE USE ids THAT'S DEFINE ABOVE
    def CreateProgram(self):
        NEW_PROGRAM_NUMBER = self.ids["JobNumber"].text
        print(NEW_PROGRAM_NUMBER)


class SettingScreen(Screen):
    # CHECK FOR ADMIN TO NOT ALLOWED ANY ONE ELSE CHANE APP SETTING
    def Admin_Check(self):
        ADMIN = "malatweh@rwbteam.com"
        if (self.manager.get_screen('LoginScreen').ids["Email"].text == ADMIN):
            self.manager.current = 'AppSettingScreen'
        else:
            # print("SORRY, YOU ARE NOT AUTHORIZED TO ACCESS THIS SCREEN")
            Close_Button = MDRaisedButton(text='Close', on_release=self.CloseDialog, font_size=16)
            self.Warning_Dialog = MDDialog(title='[color=990000]Warning Message[/color]',
                                           text=(
                                               '[color=ffffff]Sorry, You are NOT Authorized to access this screen.[/color]'),
                                           size_hint=(0.7, 1.0), buttons=[Close_Button], auto_dismiss=False)
            # TO OPEN THE DIALOG WINDOW
            self.Warning_Dialog.open()

    def CloseDialog(self, obj):
        self.Warning_Dialog.dismiss()


class PinBoreSettingScreen(Screen):
    pass


class OldHorizontalSettingScreen(Screen):
    pass


class AppSettingScreen(Screen):
    pass


# ****************STILL NEED WORK HERE*******************    Add New User     UserName       NewRWBEmail         NewWisecoEmail
class AddNewUserScreen(Screen):
    def Add_New_User(self):
        try:
            print("New User Added Successfully, Make Sure You Send Him The Shared Password By Email")
            print("USER NAME:", self.ids["UserName"].text)
            print("RWB EMAIL:", self.ids["NewRWBEmail"].text)
            print("WISECO EMAIL:", self.ids["NewWisecoEmail"].text)
            # ADD NEW USER INFORMATION FOR EXCEL FILE(USER_NAME,WISECO_EMAIL, AND RWB_EMAIL)
            USER_NAME = EMAIL_ADDRESS_LIST_FILE['Email'].loc[self.ids["UserName"].text, 'Users'] = self.ids[
                "UserName"].text
            WISECO_EMAIL = EMAIL_ADDRESS_LIST_FILE['Email'].loc[
                self.ids["NewWisecoEmail"].text, 'Wiseco Email Address'] = self.ids["NewWisecoEmail"].text
            RWB_EMAIL = EMAIL_ADDRESS_LIST_FILE['Email'].loc[self.ids["NewRWBEmail"].text, 'RWB Email Address'] = \
                self.ids["NewRWBEmail"].text
            # CREATE DATA FRAME WITH THE NEW USER DATA TO ADD THEM TO THE EXCEL SHEET
            NEW_USER_DATA = pd.DataFrame(
                data={'Users': [USER_NAME], 'Wiseco Email Address': [WISECO_EMAIL], 'RWB Email Address': [RWB_EMAIL]})
            # LOAD THE EXCEL SHEET TO BE ABLE TO ADD NEW DATA
            Email_Workbook = openpyxl.load_workbook(EMAIL_ADDRESS_LIST_FILE_PATH)
            # ACCESS THE EXCEL SHEET AND USE (mode= 'a') TO ADD THE NEW DATA
            EMAIL_SHEET_UPDATE = pd.ExcelWriter(EMAIL_ADDRESS_LIST_FILE_PATH, engine='openpyxl', mode='a')
            # SET (EMAIL_SHEET_UPDATE) AS CURRENT EXCEL BOOK (EXCEL FILE IN ANOTHER WORD)
            EMAIL_SHEET_UPDATE.book = Email_Workbook
            # LOOP THROUGH THE EXCEL FILE TO SCAN ALL THE SHEETS (MUST PUT THIS LINE OF CODE WHEN WE USE THIS MODE (mode= 'a') )
            EMAIL_SHEET_UPDATE.sheets = dict((ws.title, ws) for ws in Email_Workbook.worksheets)
            ##print("EMAIL_SHEET_UPDATE.sheets:", EMAIL_SHEET_UPDATE.sheets)
            # UPDATE THE EMAIL SHEET WITH ADDING THE NEW USER DATA
            NEW_USER_DATA.to_excel(EMAIL_SHEET_UPDATE, sheet_name='Email', startrow=Email_Workbook['Email'].max_row,
                                   startcol=0, header=False, index=False)
            # SAVE CHANGES
            EMAIL_SHEET_UPDATE.save()
            # CLOSE SHEET
            EMAIL_SHEET_UPDATE.close()

            # SHOW MESSAGE OF New User Added Successfully
            Close_Button = MDRaisedButton(text='Close', on_release=self.CloseDialog, font_size=16)
            self.Warning_Dialog = MDDialog(title='', text=(
                "New User Added Successfully, Make Sure You Send to Him The Shared Password By Email"),
                                           size_hint=(0.7, 1.0), buttons=[Close_Button], auto_dismiss=False)
            # TO OPEN THE DIALOG WINDOW
            self.Warning_Dialog.open()


        # ***ADD USER WORK FINE ONLY FOR ADDING ONE USER , THE PROBLEM IT WILL CORRUPT THE EXCEL WHICH MAKE IT INACCESSIBLE FILE, WE NEED TO FIGURE OUT THAT LATER***
        except (PermissionError):
            # SHOW MESSAGE OF New User Added Successfully
            Close_Button = MDRaisedButton(text='Close', on_release=self.CloseDialog, font_size=16)
            self.Warning_Dialog = MDDialog(title='', text=(
                "PermissionError: Something Get Wrong to Access the File, Email malatweh@rwbteam.com "),
                                           size_hint=(0.7, 1.0), buttons=[Close_Button], auto_dismiss=False)
            # TO OPEN THE DIALOG WINDOW
            self.Warning_Dialog.open()

    def CloseDialog(self, obj):
        self.Warning_Dialog.dismiss()
        # TO RESET ADD USER FIELDS TO START OVER
        self.ids["UserName"].text = ""
        self.ids["NewRWBEmail"].text = ""
        self.ids["NewWisecoEmail"].text = ""

    def ResetFields(self):
        # TO RESET ADD USER FIELDS TO START OVER
        self.ids["UserName"].text = ""
        self.ids["NewRWBEmail"].text = ""
        self.ids["NewWisecoEmail"].text = ""


# Create the screen manager
sm = ScreenManager()
sm.add_widget(LoginScreen(name='LoginScreen'))
sm.add_widget(HomeScreen(name='HomeScreen'))
sm.add_widget(PinBoreScreen(name='PinBoreScreen'))
sm.add_widget(OldHorizontalScreen(name='OldHorizontalScreen'))
sm.add_widget(NewHorizontalScreen(name='NewHorizontalScreen'))
sm.add_widget(SettingScreen(name='SettingScreen'))
sm.add_widget(PinBoreSettingScreen(name='PinBoreSettingScreen'))
sm.add_widget(OldHorizontalSettingScreen(name='OldHorizontalSettingScreen'))
sm.add_widget(AppSettingScreen(name='AppSettingScreen'))
sm.add_widget(AddNewUserScreen(name='AddNewUserScreen'))


class WisecoProgramsMaker(MDApp):

    def build(self):
        # TO CONTROL SIZE OF THE SCREEN (Window.size = (WIDTH, HEIGHT))
        Window.size = (900, 650)
        # TO CHOOSE BACKGROUND MODE OF APP WHETHER DARK OR LIGHT
        self.theme_cls.theme_style = "Dark"
        # TO SET DEFAULT COLOR OF APP ELEMENTS(LABELS,BUTTONS...ETC)
        self.theme_cls.primary_palette = "Red"
        # TO SET DEFAULT COLOR CONCENTRATION(DARKNESS AND BRIGHTNESS) OF APP ELEMENTS(LABELS,BUTTONS...ETC)
        self.theme_cls.primary_hue = "900"
        # LOAD (BuilderScreen) TO USE IT IN THE APP
        BuilderScreen = Builder.load_string(Screens_Builder)
        # LOAD (BuilderDialog) TO USE IT IN THE APP
        # BuilderDialog = Builder.load_string(Dialog_Builder)

        # TO DEFINE (Screen() THAT USED TO DISPLAY THE APP) AS (AppScreen) TO USE LATER
        AppScreen = Screen()

        ## BoxLayout FOR ENTIRE APP INCLUDE ALL WIDGETS AND ELEMENTS, SHOULD ADD ALL APP COMPONENTS FOR THIS BOX LAYOUT.
        # (orientation='vertical') TO ORGANIZE APP ELEMENTS VERTICALLY,
        # (spacing=20) TO MAKE SPACE BETWEEN APP ELEMENTS, (padding=15) TO MAKE SPACE BETWEEN WALL BORDERS AND APP ELEMENTS,
        # (md_bg_color= [32/255.0, 32/255.0, 32/255.0, 1]) TO CHANGE THE COLOR BY ADJUSTING RGB VALUE(CHECK: https://www.w3schools.com/colors/colors_picker.asp?colorhex=edfeff)
        AppBoxLayout = MDBoxLayout(orientation='vertical', spacing=20, padding=15,
                                   md_bg_color=[32 / 255.0, 32 / 255.0, 32 / 255.0, 1])

        # TO ADD PICTURE FOR THE APP FROM WEBSITE
        AppImage = AsyncImage(source='https://www.wiseco.com/Images/Downloads/Wiseco_Black_CMYK.gif', size_hint_y=None,
                              height=70, allow_stretch=True, pos_hint={'center_x': 0.5, 'center_y': 0.10},
                              color=[150 / 255.0, 0 / 255.0, 0 / 255.0, 1])
        # TO ADD AppImage TO AppBoxLayout TO DISPLAY IT IN THE APP SCREEN
        AppBoxLayout.add_widget(AppImage)

        # TO ADD Screens_Builder THAT'S CREATE ABOVE
        AppBoxLayout.add_widget(BuilderScreen)
        # TO ADD Dialog_Builder THAT'S CREATE ABOVE
        # AppBoxLayout.add_widget(BuilderDialog)

        # ADD AppBoxLayout THAT CONTAIN ALL ELEMENTS AND WIDGETS OF THE APP TO AppScreen TO DISPLAY IT IN THE APP SCREEN.
        AppScreen.add_widget(AppBoxLayout)
        return AppScreen


WisecoProgramsMaker().run()
