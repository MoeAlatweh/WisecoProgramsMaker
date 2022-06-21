# USE PANDAS LIBRARY TO READ HORIZONTAL DATA SHEET
import pandas as pd
# USE openpyxl LIBRARY TO LOAD THE WORK SHEET THAT HAS OUR HORIZONTAL DATA
from openpyxl import load_workbook, worksheet, workbook, writer
# IMPORT MATH LIBRARY TO USE SIN(ANGLE)
# import math

# import os
# import webbrowser
import subprocess


#**************************************#
# Date Variable , TO SET DATE OF TODAY BY DEFAULT
from datetime import date
today = date.today()
todaydate = today.strftime("%m/%d/%Y")

#**************************************#
# Colors for Code
RED = '\033[1;31;48m'           # RED COLOR
WHITE = '\033[1;30;48m'         # WHITE COLOR
LIGHT_BLUE = '\033[1;34;48m'    # LIGHT BLUE COLOR
YELLOW = '\033[1;33;48m'        # YELLOW COLOR
PURPLE = '\033[1;35;48m'        # PURPLE COLOR

# TO READ EXCEL FILE THAT INCLUDE FORGING DATA (WE USE -sheet_name=Forge Raw Data- TO READ THIS SHEET ONLY)
FORGING_DATA_SHEET = pd.read_excel(r'H:\CNC_Programming\Moe A\Forging Spec Data Live V8 07-29-2020.xlsm', sheet_name='Forge Raw Data')
# TO LOAD EXCEL FILE THAT INCLUDE HORIZONTAL DATA TO USE TO CHECK SHEETS EXIST
FORGING_Workbook = load_workbook(r'H:\CNC_Programming\Moe A\Forging Spec Data Live V8 07-29-2020.xlsm')
# TO PRINT SHEETS COlUMNS
print(FORGING_DATA_SHEET.keys())
print()
# TO PRINT SHEET DATA BY USING SHEET NAME
##print(FORGING_DATA_SHEET['Forge Raw Data'])
#**************************************#    col['tblForgeSpecAll_vcrIdentifier']   .at[0,'Depth']
# TO ACCESS FORGING NUMBERS WE USE COLUMN NAME TO PRINT ALL FORGINGS NUMBERS
##print(FORGING_DATA_SHEET['tblForgeSpecAll_vcrIdentifier'])
# TO ACCESS SPECIFIC FORGING BY USE ROW INDEX AND COLUMN NAME
##print(FORGING_DATA_SHEET.at[0,'tblForgeSpecAll_vcrIdentifier'])
## DEFINE FORGING NUMBER AS STRING
#Forging_Number = 'FXXXXX'
# ENTER FORGING NUMBER OF THE JOB
forging_number = 'F6398X'
# DEFINE LIST TO ADD ALL FORGING TO THE LIST
FORGING_LIST = []
# MAKE for LOOP TO READ ALL FORGING IN COLUMN OF 'tblForgeSpecAll_vcrIdentifier', AND ADDED TO FORGING_LIST TO USE THEM LATER
for forging in FORGING_DATA_SHEET['tblForgeSpecAll_vcrIdentifier']:
    # print(forging)
    FORGING_LIST.append(forging)
# print("Forging list:", FORGING_LIST)
print(FORGING_LIST)
print(len(FORGING_LIST))
# IF FORGING EXIST IN THE LIST, FIND THE index OF FORGING IN THE LIST TO USE IT TO ACCESS ANY COLUMN IN DATABASE(LIKE: ForgeRefLength, (U) Boss Outsd Spacing...ETC )
# (index = FORGING_LIST.index(FORGING_NUMBER)) USED TO SAVE index OF FORGING IN THE FORGING LIST
# (FORGING_DATA_SHEET.loc[index,'decForgeRefLength']) USED TO ACCESS SPECIFIC CELL BY PUT ROW index AND COLUMN NAME,
# WE CREATE index THING HERE BECAUSE ROW index SHOULD BE integer, NOT ACCEPT string
if forging_number in FORGING_LIST:
    print("Forging is : " + forging_number)
    index = FORGING_LIST.index(forging_number)
    # TO ACCESS SPECIFIC FORGING BY FORGING_NUMBER AND COLUMN NAME
    # (CHANGE COLUMN NAME TO ACCESS ANY OTHER NUMBERS LIKE:(U) Boss Outsd Spacing,(B) Forge O.D..ETC)
    print('(F) Forge Ref Length is:')
    print(FORGING_DATA_SHEET.loc[index,'decForgeRefLength'])
else:
    print("Forging " + forging_number + " Does Not Exist")

#++---- FILE OF HORIZ TEMPLATE ---++#

Part_Number = 'XXXXXX'

HorizontalProgram = []
HORIZONTAL_TEMPLATE_PATH = 'H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\HorizontalTemplate 01-05-21.MIN'
with open(HORIZONTAL_TEMPLATE_PATH,'rt') as CurrentProgram:
    for line in CurrentProgram:  # For each line in the file,
        HorizontalProgram.append(line.rstrip('\n'))  # strip newline and add to list.

    # TO REPLACE THE FIRST ITEM OF LIST WITH JOB NUMBER AND TODAYDATE
    HorizontalProgram[0] = ('(' + Part_Number + ' -- ' + todaydate + ' <SYS>' +')')
    ## WE CAN USE CODE BELOW TO ADD THE TOOL LIST AFTER CHECK EACH TOOL LOGIC
    # EACH TOOL NEED LOGIC TO CHECK IF WE NEED TO USE IT OR NOT
    LedgeCut = 1  # JUST FOR NOW
    for line in HorizontalProgram:
        # WE CAN USE ((T01 IS A 2.250 PILOT BORE)) OR ((**********TOOL LIST**********)) IF (T01 IS A 2.250 PILOT BORE) NOT ALWAYS TRUE
        if '(T01 IS A 2.250 PILOT BORE)' in HorizontalProgram :
            # TO FIND THE INDEX OF ITEM INSIDE LIST
            index = HorizontalProgram.index('(T01 IS A 2.250 PILOT BORE)')
            # USED TO ADD 1 TO INDEX TO GO TO PUT THE NEXT TOOL IN THE NEXT ELEMENT OF LIST
            index += 1
            # NEED TO WORK ON LOGIC OF EACH ONE OF THE TOOLS, MAKE SURE TO NOT REPEAT THE LOOP
            if (LedgeCut == 1):
                # USE insert() TO ADD NEW ELEMENTS TO THE LEST INSTEAD REPLACE THEM
                HorizontalProgram.insert(index , '(T06 IS A 5/8 END MILL - LEDGE TOOL)')
                index += 1
                LedgeCut = 0
            # HorizontalProgram[index]=('(T06 IS A 5/8 END MILL - LEDGE TOOL)')
            ##index += 1
            ##HorizontalProgram[index]=('(T11 IS A 24MM STI DRILL)')
            ##index += 1
            ## HorizontalProgram[index]=('\n')
    # HorizontalProgram[index] = ('')


    # WE CAN USE CODE BELOW TO ADJUST EACH VARIABLE DEPENDS ON JOB WE NEED TO MAKE PROGRAM FOR
    # MAKE for LOOP FOR EACH VARIABLE
    LedgeCut = 1  # JUST FOR NOW
    for line in HorizontalProgram:  # TO READ EACH LINE IN THE TEMPLATE
        substr = "VC118"    # VARIABLE WE LOOKING FOR
        # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
        index = line.find(substr)
        # print(index)
        # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
        # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
        if (index == 0):
            # IT IS RETURN index OF LINE(ELEMENT) OF THE LIST WHERE substr FOUND
            # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE LIST THAT'S CONTAIN substr WE FOUND)
            INDEX= HorizontalProgram.index(line)
            # print(INDEX)
            print(HorizontalProgram[INDEX])
            # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
            HorizontalProgram[INDEX] = ('VC118='+format(LedgeCut)+ '  (LedgeCut)')




    print(HorizontalProgram)
    print()
    print('\n'.join(HorizontalProgram))
    print()
    # THIS ONE WILL TAKE THE NEW NUMBER
    NEW_PROGRAM_CREATED = "CNCFILE1.MIN"
    try:
        # TO CREATE THE NEW FILE ON THE PATH YOU WANT
        # open() WITH "x" IT will create a file, returns an error if the file exist(THAT WHY WE USE try/except)
        # IF IT DOES NOT EXIST IT WILL CREATE THE NEW FILE , IF IT IS EXIST IT WILL GO TO except BLOCK AND CHECK IF NEED TO SAVE OVER THE EXISTING FILE
        NEW_HORIZ_PROGRAM = open("H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\\" + NEW_PROGRAM_CREATED , "x")
        print(NEW_PROGRAM_CREATED+" PROGRAM HAS BEEN CREATED")
        # NEW_HORIZ_PROGRAM.write("PROGRAM HAS BEEN CREATED!")
        NEW_HORIZ_PROGRAM.write('\n'.join(HorizontalProgram))
        NEW_HORIZ_PROGRAM.close()

        # webbrowser.open("H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\MYFILE1.txt")
        # os.system("H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\MYFILE1.txt")



    # print(NEW_HORIZ_PROGRAM.read())
    except:
        print(" Program is Exist ")
        print(RED + '# ANSWER 0 IF YOU DONT , AND 1 IF YOU DO')
        CREATE_OVER_CHECK = input(YELLOW + 'DO YOU WANT CREATE OVER THE EXIST FILE ?       ')  # ANSWER 0 IF NONE , AND 1 IS YES
        if ((CREATE_OVER_CHECK == '1') or (CREATE_OVER_CHECK == 'Y')):
            NEW_HORIZ_PROGRAM = open("H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\ " + NEW_PROGRAM_CREATED , "w")
        elif ((CREATE_OVER_CHECK == '0') or (CREATE_OVER_CHECK == 'N')):
            pass
        else: # JUST FOR NOW
            print("YOU ENTER UNEXPECTED VALUE, RUN THE CODE AGAIN ")


    programName = "Notepad.exe"
    fileName = "H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\CNCFILE1.MIN"
    subprocess.Popen([programName, fileName])

    # TO OPEN FILE ON WINDOWS USE CODE BELOW
    # (subprocess): BUILT IN FUNCTION USED TO OPEN FILES
    # (Popen): USED TO OPEN THE FILE
    # (ProgramName): KIND OF APP THAT USED TO OPEN THE FILE (LIKE NOTEPAD,EXCEL,WORD...ETC),
    # THE APP IS CIMCO IN OUR CASE BUT YOU WANT TO PUT THE PATH FOR IT TO LET WINDOW FIND IT
    # (FileName): PATH OF FILE YOU WANT TO OPEN
    ApplicationName = "C:\CIMCO\CIMCOEdit8\CIMCOEdit.EXE"
    FileName = "H:\CNCProgs\MazakLathe34 Originals\PWD-08438.EIA"
    subprocess.Popen([ApplicationName, FileName])





    # TO OPEN FILE ON WINDOWS USE LINE BELOW
    # (start): TO START THE APP YOU WANT TO OPEN
    # (NOTEPAD.EXE): APP KIND(LIKE NOTEOAD,EXCEL,WORD...ETC) YOU WANT TO OPEN
    # (H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\MYFILE1.txt): FILE YOU WANT TO OPEN
    # os.system("start NOTEPAD.EXE H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\MYFILE1.txt")

    # os.system("start C:\CIMCO\CIMCOEdit8\CIMCOEdit.EXE H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\HorizontalTemplate 01-05-21.MIN")
    # file= "H:\CNCProgs\MazakLathe34 Originals\PWD-08438.EIA"
    # os.system("start 'C:\CIMCO\CIMCOEdit8\CIMCOEdit.EXE' +file+'")


    # os.startfile("C:\CIMCO\CIMCOEdit8\CIMCOEdit.EXE H:\CNCProgs\MazakLathe34\WD-08438\\PWD-08438.EIA", 'open')

    # os.system("C:\CIMCO\CIMCOEdit8\CIMCOEdit.EXE H:\CNCProgs\MazakLathe34 Originals\PWD-08438.EIA")

    # os.system("start C:\CIMCO\CIMCOEdit8\CIMCOEdit.EXE H:\CNCProgs\MazakLathe34\WD-08438\\PWD-08438.EIA")


##+++----HORIZONTAL_TOOL_LIST----+++

# TO READ EXCEL FILE THAT CONTAIN HORIZONTAL TOOL LIST AND PROBE PROGRAMS (WE USE -sheet_name=None- TO READ ALL THE SHEETS)
horizontal_tool_list_file = pd.read_excel(r'H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\HORIZONTAL_SHEETS_FOR_AUTOMATION.xlsx', sheet_name=None)
# MAYBE WE DON'T NEED IT
# TO LOAD EXCEL FILE THAT CONTAIN HORIZONTAL TOOL LIST AND PROBE PROGRAMS
HORIZONTAL_TOOL_LIST_WORKBOOK = load_workbook(r'H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\HORIZONTAL_SHEETS_FOR_AUTOMATION.xlsx')
# TO PRINT SHEETS NAME
print(horizontal_tool_list_file.keys())
# TO PRINT SHEET DATA BY USING SHEET NAME('FINISH_BORE_TOOL_LIST' AS EXAMPLE)
##print(HORIZONTAL_TOOL_LIST_FILE['FINISH_BORE_TOOL_LIST'])

# LOGIC TO ACCESS HORIZONTAL_TOOL_LIST_FILE (EXAMPLE OF 'FINISH_BORE_TOOL_LIST' SHEET)(DO SAME THING FOR OTHER SHEETS)
# just for now
PinHoleDiameter = 0.551
# DEFINE LIST TO ADD ALL FINISH_BORE_TOOL_LIST TO THE LIST
finish_bore_tool_list = []
# MAKE for LOOP TO READ ALL FINISH_BORE_TOOL_LIST IN SHEET['FINISH_BORE_TOOL_LIST'] IN COLUMN OF 'PIN_BORE_DIAMETER', AND ADDED TO THE LIST TO USE THEM
# HORIZONTAL_TOOL_LIST_FILE: THE EXCEL FILE THAT CONTAIN HORIZONTAL TOOL LIST AND PROBE PROGRAMS
# ['FINISH_BORE_TOOL_LIST']: THE SHEET THAT'S CONTAIN TOOL LIST OF FINISH BORE
# ['PIN_BORE_DIAMETER']: THE COLUMN THAT'S CONTAIN PIN_BORE_DIAMETER SIZES WE HAVE
for tool in horizontal_tool_list_file['FINISH_BORE_TOOL_LIST']['PIN_BORE_DIAMETER']:
    # print(tool)
    finish_bore_tool_list.append(tool)
print("FINISH_BORE_TOOL_LIST:", finish_bore_tool_list)
# print(FINISH_BORE_TOOL_LIST)

# FIND THE index OF THE FINISH TOOL IN THE LIST TO USE IT TO ACCESS ANY COLUMN IN SHEET(LIKE: TOOL_NUMBER(FOR_MACHINES_27/28/29) (T00), DESCRIPTION...ETC )
# index = FINISH_BORE_TOOL_LIST.index(PinHoleDiameter) USED TO SAVE index OF TOOL IN THE TOOL LIST
# (FORGING_DATA_SHEET.loc[index,'decForgeRefLength']) USED TO ACCESS SPECIFIC CELL BY PUT ROW index AND COLUMN NAME,
# WE CREATE index THING HERE BECAUSE ROW index SHOULD BE integer, NOT ACCEPT string
index = finish_bore_tool_list.index(PinHoleDiameter)
# TO ACCESS SPECIFIC TOOL BY PinHoleDiameter AND COLUMN NAME
# ['FINISH_BORE_TOOL_LIST']: SHEET NAME
# (CHANGE COLUMN NAME TO ACCESS ANY OTHER NUMBERS LIKE: (TOOL_NUMBER(FOR_MACHINES_27/28/29) (T00)), (DESCRIPTION)...ETC
print('DESCRIPTION:')
print(horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[index, 'DESCRIPTION'])
print('TOOL NUMBER FOR OLD HORIZONTAL:')
print(horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[index, 'TOOL_NUMBER(FOR_MACHINES_27/28/29) (T00)'])

#**************************************#



