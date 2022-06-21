from kivymd.app import MDApp
from kivy.lang.builder import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivymd.uix.button import MDRectangleFlatButton
from kivymd.uix.dialog import MDDialog

# **************************************#
# USE PANDAS LIBRARY TO READ HORIZONTAL DATA SHEET
import pandas as pd
# IMPORT ExcelWriter TO BA ABLE TO UPDATE THE EXCEL SHEET
from pandas import ExcelWriter

# **************************************#
# IMPORT openpyxl TO LOAD THE WORK SHEET THAT HAS OUR FORGING DATA
import openpyxl
from openpyxl import load_workbook, worksheet, workbook, writer

screen_helper = """
ScreenManager:
    LoginScreen:
    MenuScreen:
    SettingScreen:
    AddNewUser:
<LoginScreen>:
    name: 'Login'
    MDLabel:
        text: 'Enter Login Information if You are One of Programming Team'
        pos_hint: {'center_x':0.70,'center_y':0.8}
        font_size: '18sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDLabel:
        text: 'If You Are New Programmer, Ask One of Programmers to Add you'
        pos_hint: {'center_x':0.74,'center_y':0.7}
        font_size: '14sp'
        bold: True
        italic: True
        theme_text_color: "Secondary"    
    MDLabel:
        text: 'If You Are NOT Programmer, Get Out From Here!'
        pos_hint: {'center_x':0.82,'center_y':0.6}
        font_size: '13sp'
        bold: True
        italic: True
        theme_text_color: "Error"       
    MDTextField:
        id: Email
        hint_text: "Enter Email Address"
        helper_text: "Use Work Email."
        helper_text_mode: "on_focus"
        pos_hint: {'center_x': 0.50, 'center_y': 0.45}
        size_hint_x:None
        width:300
        height:10      
    MDTextField:
        id: Password
        hint_text: "Enter Password"
        helper_text: "Use The Shared Password You Got by Email."
        helper_text_mode: "on_focus"
        password: True
        pos_hint: {'center_x': 0.50, 'center_y': 0.35}
        size_hint_x:None
        width:300
        height:10    
    MDRectangleFlatButton:
        text: 'LOGIN'
        pos_hint: {'center_x':0.5,'center_y':0.25}
        on_press : 
            root.Login_Check() 
            
        
             
   
<MenuScreen>:
    name: 'menu'
    MDRectangleFlatButton:
        text: 'APP SETTING'
        pos_hint: {'center_x':0.5,'center_y':0.2}
        on_press : 
            root.Admin_Check() 
    MDRectangleFlatButton:
        text: 'LOGOUT'
        pos_hint: {'center_x':0.5,'center_y':0.1}
        on_press : 
            root.Logout()
       

<SettingScreen>:
    name: 'SettingScreen'
    MDTextField:
        id: EmailAddressesList
        hint_text: "Email Addresses List Path"
        text: "H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\EMAIL_ADDRESS_LIST.xlsx"
        pos_hint: {'center_x': 0.50, 'center_y': 0.75}
        size_hint_x:None
        width:800
        height:50   
    MDRectangleFlatButton:
        text: 'Add New User'
        pos_hint: {'center_x':0.5,'center_y':0.2}
        on_press: root.manager.current = 'AddNewUser'          
    MDRectangleFlatButton:
        text: 'Back'
        pos_hint: {'center_x':0.5,'center_y':0.1}
        on_press: root.manager.current = 'menu'
      
        
        
<AddNewUser>:
    name: 'AddNewUser'
    MDLabel:
        text: 'Enter New User Information'
        pos_hint: {'center_x':0.74,'center_y':0.7}
        font_size: '14sp'
        bold: True
        italic: True
        theme_text_color: "Secondary"    
    MDTextField:
        id: UserName
        hint_text: "Enter User Name"
        helper_text: "Enter First Name,Last Name."
        helper_text_mode: "on_focus"
        pos_hint: {'center_x': 0.50, 'center_y': 0.55}
        size_hint_x:None
        width:300
        height:10          
    MDTextField:
        id: NewRWBEmail
        hint_text: "Enter RWB Email Address"
        helper_text: "Use Work Email."
        helper_text_mode: "on_focus"
        pos_hint: {'center_x': 0.50, 'center_y': 0.45}
        size_hint_x:None
        width:300
        height:10
    MDTextField:
        id: NewWisecoEmail
        hint_text: "Enter Wiseco Email Address -IF APPLICABLE-"
        helper_text: "If There Is No Wiseco Email. Leave It Blank."
        helper_text_mode: "on_focus"
        pos_hint: {'center_x': 0.50, 'center_y': 0.35}
        size_hint_x:None
        width:300
        height:10                  
    MDRectangleFlatButton:
        text: 'SING UP'
        pos_hint: {'center_x':0.5,'center_y':0.25}
        on_press : 
            root.Add_New_User()        
    MDRectangleFlatButton:
        text: 'Back'
        pos_hint: {'center_x':0.5,'center_y':0.1}
        on_press: root.manager.current = 'menu'         

"""


class LoginScreen(Screen):

    def Login_Check(self):
        print(self.ids["Email"].text)
        global EMAIL_ADDRESS_LIST_FILE_PATH
        EMAIL_ADDRESS_LIST_FILE_PATH = self.manager.get_screen('SettingScreen').ids["EmailAddressesList"].text
        print(EMAIL_ADDRESS_LIST_FILE_PATH)
        # TO READ EXCEL FILE THAT CONTAIN EMAIL ADDRESS OF USERS (WE USE -sheet_name=None- TO READ ALL THE SHEETS)
        global EMAIL_ADDRESS_LIST_FILE
        EMAIL_ADDRESS_LIST_FILE = pd.read_excel(EMAIL_ADDRESS_LIST_FILE_PATH, sheet_name=None)
        print(EMAIL_ADDRESS_LIST_FILE)
        # ADD WISECO_EMAIL_ADDRESS TO THE list
        WISECO_EMAIL_ADDRESS_LIST = []
        for user in EMAIL_ADDRESS_LIST_FILE['Email']['Wiseco Email Address']:
            WISECO_EMAIL_ADDRESS_LIST.append(user)
        print("USERS_LIST:", WISECO_EMAIL_ADDRESS_LIST)
        # ADD RWB_EMAIL_ADDRESS TO THE list
        RWB_EMAIL_ADDRESS_LIST = []
        for user in EMAIL_ADDRESS_LIST_FILE['Email']['RWB Email Address']:
            RWB_EMAIL_ADDRESS_LIST.append(user)
        print("USERS_LIST:", RWB_EMAIL_ADDRESS_LIST)
        PASSWORD_LIST = []
        for password in EMAIL_ADDRESS_LIST_FILE['Email']['Pass']:
            PASSWORD_LIST.append(password)
        print("PASSWORD_LIST:", PASSWORD_LIST)
        if ((self.ids["Email"].text in RWB_EMAIL_ADDRESS_LIST or self.ids["Email"].text in WISECO_EMAIL_ADDRESS_LIST)
                and (self.ids["Password"].text) in PASSWORD_LIST):
            print("LOGIN SUCCESS")
            # THAT'S HOW YOU CAN GO TO OTHER SCREENS INSIDE THE CLASS
            self.manager.current = 'menu'

        else:
            print("Wrong Email or Password, Try Again")
            Close_Button = MDRectangleFlatButton(text='Close', on_release=self.CloseDialog)
            self.Warning_Dialog = MDDialog(title='Warning Message:', text=("Wrong Email or Password, Try Again"),
                                           size_hint=(0.7, 1.0), buttons=[Close_Button])
            # TO OPEN THE DIALOG WINDOW
            self.Warning_Dialog.open()

    def CloseDialog(self, obj):
        self.Warning_Dialog.dismiss()


class MenuScreen(Screen):
    def Logout(self):
        # THAT'S HOW YOU CAN GO TO OTHER SCREENS INSIDE THE CLASS
        self.manager.current = 'Login'
        # TO RESET LOGIN FIELDS
        self.manager.get_screen('Login').ids["Email"].text = ""
        self.manager.get_screen('Login').ids["Password"].text = ""

    def Admin_Check(self):
        ADMIN = "Put Admin Email"
        if (self.manager.get_screen('Login').ids["Email"].text == ADMIN):
            self.manager.current = 'SettingScreen'
        else:
            print("SORRY, YOU ARE NOT AUTHORIZED TO ACCESS THIS SCREEN")
            Close_Button = MDRectangleFlatButton(text='Close', on_release=self.CloseDialog)
            self.Warning_Dialog = MDDialog(title='Warning Message:',
                                           text=("SORRY, YOU ARE NOT AUTHORIZED TO ACCESS THIS SCREEN"),
                                           size_hint=(0.7, 1.0), buttons=[Close_Button])
            # TO OPEN THE DIALOG WINDOW
            self.Warning_Dialog.open()

    def CloseDialog(self, obj):
        self.Warning_Dialog.dismiss()


# region ========Description=========
class SettingScreen(Screen):
    pass


# endregion


class AddNewUser(Screen):
    def Add_New_User(self):
        print("New User Added Successfully, Make Sure You Send Him The Shared Password By Email")
        print("USER NAME:", self.ids["UserName"].text)
        print("RWB EMAIL:", self.ids["NewRWBEmail"].text)
        print("WISECO EMAIL:", self.ids["NewWisecoEmail"].text)
        # ADD NEW USER INFORMATION FOR EXCEL FILE           HORIZONTAL_TOOL_LIST_FILE['MISCELLANEOUS_TOOL_LIST']
        USER_NAME = EMAIL_ADDRESS_LIST_FILE['Email'].loc[self.ids["UserName"].text, 'Users'] = self.ids["UserName"].text
        WISECO_EMAIL = EMAIL_ADDRESS_LIST_FILE['Email'].loc[self.ids["NewWisecoEmail"].text, 'Wiseco Email Address'] = \
        self.ids["NewWisecoEmail"].text
        RWB_EMAIL = EMAIL_ADDRESS_LIST_FILE['Email'].loc[self.ids["NewRWBEmail"].text, 'RWB Email Address'] = self.ids[
            "NewRWBEmail"].text
        # CREATE DATA FRAME WITH THE NEW USER DATA TO ADD THEM TO THE EXCELL SHEET
        NEW_USER_DATA = pd.DataFrame(
            data={'Users': [USER_NAME], 'Wiseco Email Address': [WISECO_EMAIL], 'RWB Email Address': [RWB_EMAIL]})
        # LOAD THE EXCEL SHEET TO BE ABLE TO ADD NEW DATA
        Email_Workbook = openpyxl.load_workbook(EMAIL_ADDRESS_LIST_FILE_PATH)
        # ACCESS THE FORGING SHEET AND USE (mode= 'a') TO ADD THE NEW DATA
        EMAIL_SHEET_UPDATE = pd.ExcelWriter(EMAIL_ADDRESS_LIST_FILE_PATH, engine='openpyxl', mode='a')
        # SET (EMAIL_SHEET_UPDATE) AS CURRENT EXCEL BOOK (EXCEL FILE IN ANOTHER WORD)
        EMAIL_SHEET_UPDATE.book = Email_Workbook
        # LOOP THROUGH THE EXCEL FILE TO SCAN ALL THE SHEETS (MUST PUT THIS LINE OF CODE WHEN WE USE
        # THIS MODE (mode= 'a') )
        EMAIL_SHEET_UPDATE.sheets = dict((ws.title, ws) for ws in Email_Workbook.worksheets)
        print("EMAIL_SHEET_UPDATE.sheets:", EMAIL_SHEET_UPDATE.sheets)
        # UPDATE THE EMAIL SHEET WITH ADDING THE NEW USER DATA
        NEW_USER_DATA.to_excel(EMAIL_SHEET_UPDATE, sheet_name='Email', startrow=Email_Workbook['Email'].max_row,
                               startcol=0, header=False, index=False)
        # SAVE CHANGES
        EMAIL_SHEET_UPDATE.save()
        # CLOSE SHEET
        EMAIL_SHEET_UPDATE.close()

        # SHOW MESSAGE OF New User Added Successfully
        Close_Button = MDRectangleFlatButton(text='Close', on_release=self.CloseDialog)
        self.Warning_Dialog = MDDialog(title='', text=(
            "New User Added Successfully, Make Sure You Send Him The Shared Password By Email"),
                                       size_hint=(0.7, 1.0), buttons=[Close_Button])
        # TO OPEN THE DIALOG WINDOW
        self.Warning_Dialog.open()

    # ***ADD USER WORK FINE , THE ONLY PROBLEM IT IS CORRUPT THE EXCEL FILE, WE NEED TO FIGURE THAT LATER***
    def CloseDialog(self, obj):
        self.Warning_Dialog.dismiss()
        # TO RESET ADD USER FIELDS TO START OVER
        self.ids["UserName"].text = ""
        self.ids["NewRWBEmail"].text = ""
        self.ids["NewWisecoEmail"].text = ""


# Create the screen manager
sm = ScreenManager()
sm.add_widget(LoginScreen(name='Login'))
sm.add_widget(MenuScreen(name='menu'))
sm.add_widget(SettingScreen(name='SettingScreen'))
sm.add_widget(AddNewUser(name='AddNewUser'))


class DemoApp(MDApp):

    def build(self):
        screen = Builder.load_string(screen_helper)
        return screen


DemoApp().run()
