# region <<<<=====================================[Application Requirements]=====================================>>>>

# +++++-------Code Beginning-------+++++#
# FROM kivy.config IMPORT Config TO CONTROL APP CONFIGURATION SETTINGS.
from kivy.config import Config

# MAKE THE APP HAVE FIXED CONFIGURATION(BY PUT False) THAT'S MAKE THE USER CAN'T CHANGE ANY THING AS MEXIMIZE THE SCREEN FOR FULL SCREEN OR CHANGE THE SIZE, TO KEEP THE APP ORGANIZED.
Config.set('graphics', 'resizable', False)
# IMPORT (MDApp) TO CREATE THE APP
from kivymd.app import MDApp
# FROM kivy.core IMPORT Window TO BE ABLE TO CONTROL THE WINDOW SIZE.
from kivy.core.window import Window
# FROM kivy.uix.image IMPORT (AsyncImage) IF NEED TO USE IMAGE FROM WEBSITE , USE (Image) IF PHOTO ON LOCAL COMPUTER
from kivy.uix.image import AsyncImage
# FROM kivy.lang IMPORT Builder THAT'S A METHOD TO CREATE THE TEXT INPUT
from kivy.lang.builder import Builder
# FROM kivy.uix.screenmanager IMPORT ScreenManager, Screen TO CREATE APP SCREEN AND MANEGE THEM
from kivy.uix.screenmanager import ScreenManager, Screen
# FROM kivymd.uix IMPORT ALL WIDGETS(LABELS, BUTTONS,BoxLayout,...) THAT USED IN THE APP.
from kivymd.uix.label import MDLabel
from kivymd.uix.button import MDRaisedButton, MDRaisedButton, MDFloatingActionButton, MDRoundFlatButton, \
    MDFillRoundFlatButton
from kivymd.uix.gridlayout import MDGridLayout
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.textfield import MDTextField
from kivy.uix.textinput import TextInput
# USE MDDialog TO SHOW DIALOG WINDOW OF MESSAGES
from kivymd.uix.dialog import MDDialog
from kivymd.uix.list import MDCheckbox, OneLineAvatarListItem
from kivy.uix.widget import Widget

from kivy.uix.scrollview import ScrollView
from kivy.uix.boxlayout import BoxLayout
# from kivy._event import fbind
# from kivy.core.text import
from kivymd.uix.button import MDFlatButton

from kivy.properties import StringProperty

from kivy.properties import ObjectProperty

# IMPORT MATH LIBRARY TO USE SIN(ANGLE)
import math

# USE glob() (BUILT IN FUNCTION IN PYTHON) TO SEARCH FILES INSIDE FOLDER USING FORGING NUMBER .
import glob

# **************************************#
# USE subprocess (BUILD IN FUNCTION) TO OPEN THE PROGRAM IN CIMCO
import subprocess

# **************************************#
# USE PANDAS LIBRARY TO READ HORIZONTAL DATA SHEET
import pandas as pd
# IMPORT ExcelWriter TO BA ABLE TO UPDATE THE EXCEL SHEET
from pandas import ExcelWriter

# **************************************#
# IMPORT openpyxl TO LOAD THE WORK SHEET THAT HAS OUR FORGING DATA
import openpyxl
from openpyxl import load_workbook, worksheet, workbook, writer

# IMPORT EMAIL PACKAGE
import smtplib
# IMPORT PASSWORD PACKAGE
import keyring

# **************************************#
# Date Variable , TO SET DATE OF TODAY BY DEFAULT
from datetime import date

today = date.today()
todaydate = today.strftime("%m/%d/%Y")

# endregion <<<<====================================[Application Requirements]====================================>>>>


# region <<<<========================================[Screen Builder KV]=========================================>>>>

Screens_Builder = """
ScreenManager:
    LoginScreen:
    HomeScreen:
    PinBoreScreen:
    OldHorizontalScreen:
    NewHorizontalScreen:
    SettingScreen:
    PinBoreSettingScreen:
    OldHorizontalSettingScreen:
    AppSettingScreen:
    AddNewUserScreen:


<LoginScreen>:
    name: 'LoginScreen'
    MDLabel:
        text: 'Enter Login Information if You are One of Programming Team.'
        pos_hint: {'center_x':0.69,'center_y':0.85}
        font_size: '20sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDLabel:
        text: 'If You Are New Programmer, Ask One of Programmers to Add You.'
        pos_hint: {'center_x':0.73,'center_y':0.75}
        font_size: '16sp'
        bold: True
        italic: True
        theme_text_color: "Secondary"    
    MDLabel:
        text: ''
        pos_hint: {'center_x':0.81,'center_y':0.65}
        font_size: '15sp'
        bold: True
        italic: True
        theme_text_color: "Error"       
    MDTextField:
        id: Email
        text: self.text.lower() if self.text is not None else ''
        hint_text: "Enter Email Address"
        helper_text: "Use Email With Extension @rwbteam or @wiseco."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.45}
        size_hint_x:None
        width:300
        height:10      
    MDTextField:
        id: Password
        hint_text: "Enter Password"
        helper_text: "Use The Shared Password You Got by Email."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        password: True
        pos_hint: {'center_x': 0.50, 'center_y': 0.33}
        size_hint_x:None
        width:300
        height:10    
    MDRaisedButton:                                                                         
        text: 'LOGIN'
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        pos_hint: {'center_x':0.5,'center_y':0.23}
        on_press : 
            root.login_check()     
    MDRaisedButton:                                                                         
        text: 'Version 1.0.0'
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        pos_hint: {'center_x':0.07,'center_y':0.03}  
        on_press : 
            root.application_version_features()        
    MDLabel:
        text: 'Created by: Moemen Alatweh'
        pos_hint: {'center_x':1.32,'center_y':0.04}
        font_style: 'Caption'
        theme_text_color: "Custom"
        text_color: 175/255.0, 0/255.0, 0/255.0, 1
    MDLabel:
        text: 'malatweh@rwbteam.com'
        pos_hint: {'center_x':1.33,'center_y':0.01}
        font_style: 'Caption'
        theme_text_color: "Custom"
        text_color: 175/255.0, 0/255.0, 0/255.0, 1           


<HomeScreen>:
    name: 'HomeScreen'
    on_pre_enter: root.set_user_name()
    MDLabel:
        text: 'Welcome to Wiseco Programs Maker'
        pos_hint: {'center_x':0.70,'center_y':0.8}
        font_size: '32sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDLabel:
        id: UserName
        text: ''
        halign:'center'
        valign: 'middle'
        pos_hint: {'center_y':0.7}
        font_size: '24sp'
        bold: True
        italic: True
        theme_text_color: "Error"  
    MDLabel:
        text: 'Choose Type of Program'
        pos_hint: {'center_x':0.885,'center_y':0.55}
        font_size: '18sp'
        bold: True
        italic: True
        theme_text_color: "Secondary"    
    MDRaisedButton:
        text: 'Pin Bore'
        pos_hint: {'center_x':0.5,'center_y':0.45}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'PinBoreScreen'    
    MDRaisedButton:
        text: 'Setting'
        halign:'left'
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'SettingScreen'
    MDRaisedButton:
        text: 'LOGOUT'
        pos_hint: {'center_x':0.945,'center_y':0.035}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : 
            root.logout()         




<PinBoreScreen>:
    name: 'PinBoreScreen'
    MDLabel:
        text: 'Pin Bore Programs Maker'
        pos_hint: {'center_x':0.80,'center_y':0.8}
        font_size: '30sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDLabel:
        text: 'Choose Machine That Need Program'
        pos_hint: {'center_x':0.83,'center_y':0.6}
        font_size: '18sp'
        bold: True
        italic: True
        theme_text_color: "Secondary"    
    MDRaisedButton:
        text: 'Old Horizontal Machines (28,29,32)'
        pos_hint: {'center_x':0.5,'center_y':0.5}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'OldHorizontalScreen'
    MDRaisedButton:
        text: 'New Horizontal Machine (127)'
        pos_hint: {'center_x':0.5,'center_y':0.4}                    
        md_bg_color: 120/255, 0/255, 0/255, 1   
        font_size: "15sp"
        on_press : root.still_work_on_it(object) 
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.5,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'    



<OldHorizontalScreen>:
    name: 'OldHorizontalScreen'
    MDLabel:
        text: 'Old Horizontal Machines (28,29,32)'
        pos_hint: {'center_x':0.73,'center_y':0.8}
        font_size: '30sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDTextField:
        id: JobNumber
        text: self.text.upper() if self.text is not None else ''
        hint_text: "Enter Job Number"
        helper_text: "Job number MUST match with number in Spec."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.60}
        size_hint_x:None
        width:300
        height:10      
    MDRaisedButton:
        text: 'Submit'
        pos_hint: {'center_x':0.5,'center_y':0.45}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : root.create_program_for_old_horizontal_machine(object)    
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: 
            root.manager.current = 'PinBoreScreen'   
            root.reset_old_horizontal_screen_fields()     
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: 
            root.manager.current = 'HomeScreen'        
            root.reset_old_horizontal_screen_fields() 


<NewHorizontalScreen>:
    name: 'NewHorizontalScreen'
    MDLabel:
        text: 'New Horizontal Machine (127)'
        pos_hint: {'center_x':0.84,'center_y':0.8}
        font_size: '20sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDTextField:
        id: JobNumber
        hint_text: "Enter Job Number"
        helper_text: "Job number MUST match with number in Spec."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.60}
        size_hint_x:None
        width:300
        height:10      
    MDRaisedButton:
        text: 'Submit'
        pos_hint: {'center_x':0.5,'center_y':0.45}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : root.create_program_for_new_horizontal_machine()

    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'PinBoreScreen'  
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'     


<SettingScreen>:
    name: 'SettingScreen'
    MDLabel:
        text: 'Setting'
        pos_hint: {'center_x':0.93,'center_y':0.9}
        font_size: '36sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDRaisedButton:
        text: 'Pin Bore Setting'
        pos_hint: {'center_x':0.5,'center_y':0.7}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'PinBoreSettingScreen'
    MDRaisedButton:
        text: 'Application Setting'
        pos_hint: {'center_x':0.5,'center_y':0.6}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : 
            root.admin_check() 
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.5,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'   


<PinBoreSettingScreen>:
    name: 'PinBoreSettingScreen'
    MDLabel:
        text: 'Pin Bore Setting'
        pos_hint: {'center_x':0.86,'center_y':0.9}
        font_size: '36sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDRaisedButton:
        text: 'Old Horizontal Setting Screen'
        pos_hint: {'center_x':0.5,'center_y':0.7}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'OldHorizontalSettingScreen'
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'SettingScreen'



<OldHorizontalSettingScreen>:
    name: 'OldHorizontalSettingScreen'
    MDLabel:
        text: 'Old Horizontal Setting Screen'
        pos_hint: {'center_x':0.77,'center_y':0.9}
        font_size: '32sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDTextField:
        id: HorizontalTemplate
        hint_text: "Horizontal Template Path"
        text: "H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\HorizontalTemplate 01-05-21.MIN"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.75}
        size_hint_x:None
        width:800
        height:50       
    MDTextField:
        id: HorizontalToolList
        hint_text: "Horizontal Tool List Excel File Path"
        text: "H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\HORIZONTAL_SHEETS_FOR_AUTOMATION.xlsx"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.65}
        size_hint_x:None
        width:800
        height:50          
    MDTextField:
        id: ProbePrograms
        hint_text: "Probe Programs File Path"
        text: "H:\CNCProgs\HOREBORE\Probe Programs"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.55}
        size_hint_x:None
        width:800
        height:50              
    MDTextField:
        id: RunningFolderPathOfOldHorizontalMachine
        hint_text: "Horizontal Programs (Running Folder) Path"
        text: "H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\Horizontal"
        helper_text: "Folder that Use on Machine to Load the Program."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.45}
        size_hint_x:None
        width:800
        height:50
    MDTextField:
        id: OriginalFolderPathOfOldHorizontalMachine
        hint_text: "Horizontal Programs (Original Folder) Path"
        text: "H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\Horizontal Original"
        helper_text: "Folder that Use as Backup For Programs."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.35}
        size_hint_x:None
        width:800
        height:50                
    MDTextField:
        id: CimcoEditorPath
        hint_text: "CIMCO Editor Path"
        text: "C:\CIMCO\CIMCOEdit8\CIMCOEdit.EXE"
        helper_text: "Path Should be Where CIMCO App Installed in User Computer."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.25}
        size_hint_x:None
        width:800
        height:50                          
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'PinBoreSettingScreen'      


<AppSettingScreen>:
    name: 'AppSettingScreen'
    MDLabel:
        text: 'Application Setting'
        pos_hint: {'center_x':0.87,'center_y':0.9}
        font_size: '32sp'
        bold: True
        italic: True
        theme_text_color: "Primary"
    MDTextField:
        id: EmailAddressList
        hint_text: "Email Address List File Path"
        text: "H:\CNC_Programming\Moe A\WisecoApplications\TestStuff\EMAIL_ADDRESS_LIST.xlsx"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.75}
        size_hint_x:None
        width:800
        height:50   
    MDTextField:
        id: TrelloEmailAddress
        hint_text: "Trello Board Email Address"
        text: "moemenalatweh1+sqa4wcni54jz6erwnpbj@boards.trello.com"                   
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.65}
        size_hint_x:None
        width:800
        height:50       
    MDRaisedButton:
        text: 'Add New User'
        pos_hint: {'center_x':0.5,'center_y':0.25}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'AddNewUserScreen'    
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'HomeScreen'
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: root.manager.current = 'SettingScreen' 



<AddNewUserScreen>:
    name: 'AddNewUserScreen'
    MDLabel:
        text: 'Add New User'
        pos_hint: {'center_x':0.90,'center_y':0.9}
        font_size: '32sp'
        bold: True
        italic: True
        theme_text_color: "Primary"   
    MDTextField:
        id: UserName
        hint_text: "Enter User Name"
        helper_text: "First Name, Last Name."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.65}
        size_hint_x:None
        width:400
        height:10          
    MDTextField:
        id: NewRWBEmail
        hint_text: "Enter RWB Email Address"
        helper_text: "Email With Extension @rwbteam."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.55}
        size_hint_x:None
        width:400
        height:10
    MDTextField:
        id: NewWisecoEmail
        hint_text: "Enter Wiseco Email Address"
        helper_text: "Email With Extension @wiseco, If Not Applicable. Leave It Blank."
        helper_text_mode: "on_focus"
        color_mode: 'custom'
        line_color_focus: 1, 1, 1, 1
        pos_hint: {'center_x': 0.50, 'center_y': 0.45}
        size_hint_x:None
        width:400
        height:10                  
    MDRaisedButton:
        text: 'SING UP'
        pos_hint: {'center_x':0.5,'center_y':0.30}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press : 
            root.add_new_user()        
    MDRaisedButton:
        text: 'Home'
        pos_hint: {'center_x':0.4,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: 
            root.manager.current = 'HomeScreen'
            root.reset_new_user_screen_fields() 
    MDRaisedButton:
        text: 'Back'
        pos_hint: {'center_x':0.6,'center_y':0.1}
        md_bg_color: 120/255, 0/255, 0/255, 1
        font_size: "15sp"
        on_press: 
            root.manager.current = 'AppSettingScreen'  
            root.reset_new_user_screen_fields() 

"""


# endregion <<<<=======================================[Screen Builder KV]========================================>>>>


# region <<<<==================================[Load Horizontal Sheets Function]=================================>>>>

# CREATE FUNCTION (LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION) TO LOAD HORIZONTAL TOOL LIST SHEET TO USE IT LATER FOR
# ALL HORIZONTAL PROGRAMS WHETHER 4-CYCLE, 2-CYCLE, OLD, OR NEW MACHINE
def load_horizontal_machine_tool_list_sheets(self):
    # DEFINE VARIABLE(HORIZONTAL_TOOL_LIST_FILE) TO STORE HORIZONTAL_TOOL_LIST_FILE PATH FROM SETTING SCREEN
    # TO GET VALUE FROM ANOTHER SCREEN(SettingScreen IN THIS EXAMPLE) TO USE IT
    # IN ANOTHER SCREEN CLASS (OldHorizontalScreen IN THIS EXAMPLE) FOR LOGIC PURPOSES>>
    # >> USE (self.manager.get_screen('SettingScreen').ids["HorizontalToolList"].text)
    # self.manager.get_screen('SettingScreen'): PUT SCREEN YOU WANT TO ACCESS
    # ids["HorizontalToolList"].text : PUT WIDGET OR VALUE YOU NEED
    # TO GET(IN THIS EXAMPLE WE GET THE TEXT OF MDTextField THAT NAME "HorizontalToolList")
    horizontal_tool_list_file_path = self.manager.get_screen('OldHorizontalSettingScreen').ids[
        "HorizontalToolList"].text
    print(horizontal_tool_list_file_path)
    # TO READ EXCEL FILE THAT CONTAIN HORIZONTAL TOOL LIST AND PROBE PROGRAMS
    # (WE USE -sheet_name=None- TO READ ALL THE SHEETS)
    global horizontal_tool_list_file
    horizontal_tool_list_file = pd.read_excel(horizontal_tool_list_file_path, sheet_name=None)

    # try:
    #     print("try LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
    #     global HORIZONTAL_TOOL_LIST_FILE
    #     HORIZONTAL_TOOL_LIST_FILE = pd.read_excel(HORIZONTAL_TOOL_LIST_FILE_PATH, sheet_name=None)
    # except Exception as error:
    #     print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
    #
    #     fail_messages_of_creating_old_horizontal_machine_program.append(
    #         "Failed to Access Horizontal Tool List File ." + "\n" + "An Error has occurred :" + "\n" +
    #         '[color=ff1a1a]' + str(
    #             error) + '[/color]' + "\n" + "Double Check Network and File Path Location.")
    #     # email_messages_of_creating_old_horizontal_machine_program.append
    #     ("\n".join(fail_messages_of_creating_old_horizontal_machine_program))
    #     Failed_to_Create_Horizontal_Program(self)
    #     return

    # LEAVE IT FOR TEST
    # print(HORIZONTAL_TOOL_LIST_FILE)
    # TO PRINT SHEETS NAME
    # print(HORIZONTAL_TOOL_LIST_FILE.keys())
    # TO PRINT SHEET DATA BY USING SHEET NAME('FINISH_BORE_TOOL_LIST' AS EXAMPLE)
    # print(HORIZONTAL_TOOL_LIST_FILE['FINISH_BORE_TOOL_LIST'])
    # DEFINE CIMCO PATH
    global cimco_editor_path
    cimco_editor_path = self.manager.get_screen('OldHorizontalSettingScreen').ids["CimcoEditorPath"].text
    print("CIMCO EDITOR PATH: ", cimco_editor_path)

    # DEFINE LIST TO ADD ALL FINISH_BORE_TOOL_LIST TO THE LIST
    global finish_bore_tool_list
    finish_bore_tool_list = []
    # MAKE for LOOP TO READ ALL FINISH_BORE_TOOL_LIST IN SHEET['FINISH_BORE_TOOL_LIST'] IN COLUMN OF'PIN_BORE_DIAMETER',
    # AND ADDED TO THE LIST TO USE THEM
    # HORIZONTAL_TOOL_LIST_FILE: THE EXCEL FILE THAT CONTAIN HORIZONTAL TOOL LIST AND PROBE PROGRAMS
    # ['FINISH_BORE_TOOL_LIST']: THE SHEET THAT'S CONTAIN TOOL LIST OF FINISH BORE
    # ['PIN_BORE_DIAMETER']: THE COLUMN THAT'S CONTAIN PIN_BORE_DIAMETER SIZES WE HAVE
    for tool in horizontal_tool_list_file['FINISH_BORE_TOOL_LIST']['PIN_BORE_DIAMETER']:
        # print(tool)
        finish_bore_tool_list.append(tool)
    print("FINISH_BORE_TOOL_LIST:", finish_bore_tool_list)

    # DEFINE LIST TO ADD ALL ROUGH_BORE_TOOL_LIST TO THE LIST
    global rough_bore_tool_list
    rough_bore_tool_list = []
    # MAKE for LOOP TO READ ALL ROUGH_BORE_TOOL_LIST IN SHEET['ROUGH_BORE_TOOL_LIST'] IN COLUMN OF 'DRILL_DIAMETER',
    # AND ADDED TO THE LIST TO USE THEM
    for tool in horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST']['DRILL_DIAMETER']:
        # print(tool)
        rough_bore_tool_list.append(tool)
    print("ROUGH_BORE_TOOL_LIST:", rough_bore_tool_list)

    # DEFINE LIST TO ADD ALL LOCK_RING_AND_CFREN_TOOL TO THE LIST
    global lock_ring_and_cfren_tool_list
    lock_ring_and_cfren_tool_list = []
    # MAKE for LOOP TO READ ALL LOCK_RING_AND_CFREN_TOOL_LIST IN SHEET['LOCK_RING_AND_CFREN_TOOL_LIST'] IN COLUMN OF
    # 'TOOL_WIDTH', AND ADDED TO THE LIST TO USE THEM
    for tool in horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST']['TOOL_WIDTH']:
        # print(tool)
        lock_ring_and_cfren_tool_list.append(tool)
    print("LOCK_RING_AND_CFREN_TOOL_LIST:", lock_ring_and_cfren_tool_list)

    # DEFINE LIST TO ADD ALL MISCELLANEOUS_TOOL_LIST TO THE LIST
    global miscellaneous_tool_list
    miscellaneous_tool_list = []
    # MAKE for LOOP TO READ ALL MISCELLANEOUS_TOOL_LIST IN SHEET['MISCELLANEOUS_TOOL_LIST'] IN COLUMN OF 'TOOL_USAGE',
    # AND ADDED TO THE LIST TO USE THEM
    for tool in horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST']['TOOL_USAGE']:
        # print(tool)
        miscellaneous_tool_list.append(tool)
    print("MISCELLANEOUS_TOOL_LIST:", miscellaneous_tool_list)

    # DEFINE LIST TO ADD ALL HORIZONTAL_SLOT_NUMBERS TO THE LIST
    global horizontal_slot_numbers
    horizontal_slot_numbers = []
    # MAKE for LOOP TO READ ALL HORIZONTAL_SLOT_NUMBERS IN SHEET['HORIZONTAL_SLOT_NUMBERS'] IN COLUMN OF
    # 'PIN_BORE_DIAMETER', AND ADDED TO THE LIST TO USE THEM
    for tool in horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS']['PIN_BORE_DIAMETER']:
        # print(tool)
        horizontal_slot_numbers.append(tool)
    print("HORIZONTAL_SLOT_NUMBERS:", horizontal_slot_numbers)

    # DEFINE LIST TO ADD ALL PROBE_PROGRAMS TO THE LIST
    # global PROBE_PROGRAMS_LIST
    # PROBE_PROGRAMS_LIST = []
    # # MAKE for LOOP TO READ ALL PROBE_PROGRAMS IN SHEET['PROBE_PROGRAMS'] IN COLUMN OF 'FORGING_NUMBER',
    # AND ADDED TO THE LIST TO USE THEM
    # for program in HORIZONTAL_TOOL_LIST_FILE['PROBE_PROGRAMS']['FORGING_NUMBER']:
    #     # print(tool)
    #     PROBE_PROGRAMS_LIST.append(program)
    # print("PROBE_PROGRAMS:", PROBE_PROGRAMS_LIST)

    # # define list here to store confirnation message
    # global confirmation_email_messages_of_creating_old_horizontal_machine_program
    # confirmation_email_messages_of_creating_old_horizontal_machine_program = []


# endregion <<<<=================================[Load Horizontal Sheets Function]================================>>>>


# region <<<<===================================[Four Cycle Pin Bore Function]===================================>>>>

# NEED TO ADD ALL 4 CYCLE PINBORE VARIABLES TO THIS FUNCTION TO AVOID REPETITION
def four_cycle_pin_bore_variables():
    # NEED TO ADD LOGIC FOR SOME VARIABLES TO NOT PROCEED IF DOES NOT FIND THE VALUE FROM QANTEL DATA BASE

    global pin_hole_diameter
    pin_hole_diameter = 0.927  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global pilot_bore_depth
    pilot_bore_depth = 01.8481  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global pilot_to_pin
    pilot_to_pin = -0.25  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global X_distance_from_origin_to_pin_center
    X_distance_from_origin_to_pin_center = 0  # JUST DEFINE THAT TO USE IT ON MATH LATER
    # WE NEED LOGIC HERE TO CHECK THE OFFSET DIRECTION
    global offset_amount  # WE JUST USE EMPTY STRING ("") as test WHEN CAN'T DETECT THE OFFSET FROM DATA BASE
    offset_amount = 0.03  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    # WE JUST USE EMPTY STRING ("") as test WHEN CAN'T DETECT THE OFFSET FROM DATA BASE
    global offset_direction  # OFFSET To0      OFFSET To180        OFFSET EACH WAY
    offset_direction = "OFFSET To180"
    global rough_bore_speed
    rough_bore_speed = 8000  # THIS NUMBER COMES FROM HORIZONTAL TEMPLATE
    global rough_bore_feed
    rough_bore_feed = 100  # THIS NUMBER COMES FROM HORIZONTAL TEMPLATE
    global value_used_in_Z_value_finish_bore_bottom
    value_used_in_Z_value_finish_bore_bottom = 1.156  # THIS NUMBER COMES FROM HORIZONTAL TEMPLATE
    global ledge_tool_diameter
    ledge_tool_diameter = 0.625  # THIS NUMBER COMES FROM HORIZONTAL TEMPLATE
    # WE DEFINE IT HERE NOT BELOW BECAUSE sometimes APP CAN'T DECIDE THE STATUS, IT WILL ASK FOE USER INPUT
    global ledge_cut_availability_status
    ledge_cut_availability_status = 1
    global lock_ring_cutter_width
    lock_ring_cutter_width = 0  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global lock_ring_ID_spacing
    lock_ring_ID_spacing = 0  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global lock_ring_diameter
    lock_ring_diameter = 0  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global lock_ring_tool_diameter
    lock_ring_tool_diameter = 0  # JUST DEFINE THAT TO USE IT ON MATH LATER ,THIS NUMBER COMES FROM TOOLLISTSHEET
    global cfren_cutter_width
    cfren_cutter_width = 0.077  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global cfren_ID_spacing
    cfren_ID_spacing = 1.625  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global cfren_diameter
    cfren_diameter = 0.969  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global cfren_tool_diameter
    cfren_tool_diameter = 0  # JUST DEFINE THAT TO USE IT ON MATH LATER ,THIS NUMBER COMES FROM TOOLLISTSHEET
    # NEED LOGIC RELATE ON QANTEL DATABASE TO KNOW IF JOB USE SAME TOOL FOR LOCKRING AND CFREN OR
    # DEFFERNT OR JUST HAVE CFREN

    global notch_angle_first_location
    notch_angle_first_location = 135  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global notch_angle_second_location
    notch_angle_second_location = 225  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global X_distance_from_origin_to_circlip_notch
    X_distance_from_origin_to_circlip_notch = 0  # JUST DEFINE THAT TO USE IT ON MATH OF NOTCH LOGIC
    global Y_distance_from_origin_to_circlip_notch
    Y_distance_from_origin_to_circlip_notch = 0  # JUST DEFINE THAT TO USE IT ON MATH OF NOTCH LOGIC
    global double_oil_hole_slot_ID_spacing
    double_oil_hole_slot_ID_spacing = 01.7823  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    # MAYBE NEED TO ADD LOGIC HERE TO DECIDE WHAT THE ID SPACING SHOULD BE, TAKE IT FROM QANTEL,
    # OR MAKE CALCULATION BY ADD LR ID SPACING + LR WIDTH IF HORIZ SLOTS STOP AT LOCKRING
    global horizontal_slots_ID_spacing
    horizontal_slots_ID_spacing = 2.153  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    global horizontal_slots_arc_diameter
    horizontal_slots_arc_diameter = 0.375  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER

    # MAYBE THE WAY WE PROGRAM HORIZONTAL SLOTS WILL CHANGE SOON , BUT FOR NOW WE WILL USE THE WAY WE HAVE

    global i_start_horizontal_slot
    i_start_horizontal_slot = 0  # JUST DEFINE THAT TO USE IT LATER , THIS NUMBER COMES FROM TOOL LIST SHEET
    global j_start_horizontal_slot
    j_start_horizontal_slot = 0  # JUST DEFINE THAT TO USE IT LATER , THIS NUMBER COMES FROM TOOL LIST SHEET
    global horizontal_slot_radius
    horizontal_slot_radius = 0  # JUST DEFINE THAT TO USE IT LATER , THIS NUMBER COMES FROM TOOL LIST SHEET
    global ledge_counterbore_diameter
    ledge_counterbore_diameter = 0  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    # MAYBE NEED VARIABLE OF DISTANCE OF LEDGE_COUNTERBORE TO LOCKRING HERE AND IN THE HORIZ TAMPLATE

    # JUST DEFINE THAT TO USE IT ON MATH LATER
    global X_distance_of_375_slots_from_center_of_bore_to_center_of_horizontal_slot
    X_distance_of_375_slots_from_center_of_bore_to_center_of_horizontal_slot = 0
    global Y_distance_of_375_slots_from_center_of_bore_to_center_of_horizontal_slot
    Y_distance_of_375_slots_from_center_of_bore_to_center_of_horizontal_slot = 0

    global forging_number
    forging_number = "F6473X"  # JUST FOR NOW, NEED TO ADD IT FROM QANTEL DATABASE LATER
    # NEED LOGIC TO CHECK IF WE HAVE THE FORGING IN THE FORGING DATA BASE
    # NEED LOGIC TO CHECK IF WE HAVE THE FORGE_REF_LENGTH NUMBER IN THE FORGING DATA BASE
    # FORGE REF LENGTH IS (F) VALUE IN EMSS APP
    global forge_ref_length
    forge_ref_length = 2.514  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER
    # FORGING_DIAMETER IS FIRST_B_NUMBER_IN_EMSS [(B) Forge.O.D.]
    global forging_diameter
    forging_diameter = 04.2  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER
    # FORGING_DIAMETER_OD_AT_ROUGHER IS SECOND_B_NUMBER_IN_EMSS [(B) O.D.At_Rougher]
    global forging_diameter_OD_at_rougher
    forging_diameter_OD_at_rougher = 4.12  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER
    # FORGING_OUTSIDE_BOSS_SPACING IS_U_NUMBER_IN_EMSS
    global forging_outside_boss_spacing
    forging_outside_boss_spacing = 2.38  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER
    # FORGING_INSIDE_BOSS_SPACING IS_J_NUMBER_IN_EMSS
    global forging_inside_boss_spacing
    forging_inside_boss_spacing = 0.800  # JUST FOR NOW, NEED TO ADD IT FROM FORGING DATABASE LATER

    # define list here to avoid reset the list when call (create_old_horizontal_program) function
    global old_horizontal_program_confirmation_email_message_list
    old_horizontal_program_confirmation_email_message_list = []


# endregion <<<<==================================[Four Cycle Pin Bore Function]==================================>>>>


# region <<<<===========================================[Login Screen]===========================================>>>>

class LoginScreen(Screen):
    # FUNCTION TO SHOW APP VERSION
    def application_version_features(self):
        print("application_version_features" + " is called")
        application_version_features_list = \
            ["An App to create CNC programs according to the type of operation and machine. ",
             "[color=ff1a1a]Version :[/color] 1.0.0"
             + "[color=ff1a1a]                                                                   "
               "Release Date :[/color] 06/28/2021",
             "--------------------------------------------------------------------------"
             "-----------------------------------------------------------------",
             "[color=ff1a1a]Features :[/color]", "[#] Create Pin Bore program for Horizontal machines (28,29.32). "]
        close_button = MDRaisedButton(text='Close', on_release=self.close_login_screen_window, font_size=16)
        self.login_screen_message_window = MDDialog(title='[b][color=ffffff]Wiseco Programs Maker App[/color][/b]',
                                                    text=('[color=ffffff]' + '\n'.join(
                                                        application_version_features_list) + '[/color]'),
                                                    size_hint=(0.7, 1.0), buttons=[close_button], auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.login_screen_message_window.open()

    def login_check(self):
        global user_email_address
        user_email_address = self.ids["Email"].text
        print("USER: ", user_email_address)
        # global EMAIL_ADDRESS_LIST_FILE_PATH
        # EMAIL_ADDRESS_LIST_FILE_PATH = self.manager.get_screen('AppSettingScreen').ids["EmailAddressList"].text
        # print(EMAIL_ADDRESS_LIST_FILE_PATH)
        # TO READ EXCEL FILE THAT CONTAIN EMAIL ADDRESS OF USERS (WE USE -sheet_name=None- TO READ ALL THE SHEETS)
        # global EMAIL_ADDRESS_LIST_FILE
        # EMAIL_ADDRESS_LIST_FILE = pd.read_excel(EMAIL_ADDRESS_LIST_FILE_PATH, sheet_name=None)
        try:
            global email_address_list_file_path
            email_address_list_file_path = self.manager.get_screen('AppSettingScreen').ids["EmailAddressList"].text
            # TO READ EXCEL FILE THAT CONTAIN EMAIL ADDRESS OF USERS (WE USE -sheet_name=None- TO READ ALL THE SHEETS)
            global email_address_list_file
            email_address_list_file = pd.read_excel(email_address_list_file_path, sheet_name=None)
            # ADD USERS NAME TO THE list
            global users_name_list
            users_name_list = []
            for user in email_address_list_file['Email']['Users']:
                users_name_list.append(user)
            # print("USERS_NAME_LIST:", USERS_NAME_LIST)
            # ADD WISECO_EMAIL_ADDRESS TO THE list
            global wiseco_email_address_list
            wiseco_email_address_list = []
            for user in email_address_list_file['Email']['Wiseco Email Address']:
                wiseco_email_address_list.append(user)
            # print("WISECO_USERS_LIST:", WISECO_EMAIL_ADDRESS_LIST)
            # ADD RWB_EMAIL_ADDRESS TO THE list
            global rwb_email_address_list
            rwb_email_address_list = []
            for user in email_address_list_file['Email']['RWB Email Address']:
                rwb_email_address_list.append(user)
            # print("RWB_USERS_LIST:", RWB_EMAIL_ADDRESS_LIST)
            # ADD Trello Users name TO THE list
            global trello_users_name_list
            trello_users_name_list = []
            for user in email_address_list_file['Email']['Trello Users Name']:
                trello_users_name_list.append(user)
            # print("RWB_USERS_LIST:", RWB_EMAIL_ADDRESS_LIST)
            # ADD PASSWORD TO THE list
            global password_list
            password_list = []
            for password in email_address_list_file['Email']['Pass']:
                password_list.append(password)
        except Exception as error:
            close_button = MDRaisedButton(text='Close', on_release=self.close_login_screen_window, font_size=16)
            self.login_screen_message_window = MDDialog(title='[color=990000]Warning Message[/color]',
                                                        text=("[color=ffffff]Failed to Find, Load, or Access" +
                                                              '[b][u][color=ffffff] Email Address List [/color][/u][/b]'
                                                              + "\n" + "An Error has occurred [/color]" + "\n" +
                                                              '[color=ff1a1a]' + str(error) + '[/color]' + "\n" +
                                                              "[color=ffffff]Double Check Network, "
                                                              "and File Location.[/color]"),
                                                        size_hint=(0.7, 1.0), buttons=[close_button],
                                                        auto_dismiss=False)
            # TO OPEN THE DIALOG WINDOW
            self.login_screen_message_window.open()
            return
        # TO CHECK LOGIN INFORMATION
        if (((self.ids["Email"].text in rwb_email_address_list and self.ids[
            "Email"].text != 'malatweh@rwbteam.com') or (self.ids["Email"].text in wiseco_email_address_list))
                and ((self.ids["Password"].text) in password_list)):
            print("LOGIN SUCCESS")
            # THAT'S HOW YOU CAN GO TO OTHER SCREENS INSIDE THE CLASS
            self.manager.current = 'HomeScreen'
        elif (self.ids["Email"].text == 'malatweh@rwbteam.com' and (
                self.ids["Password"].text == 'moe' + password_list[6])):
            print("ADMIN LOGIN SUCCESS")
            # THAT'S HOW YOU CAN GO TO OTHER SCREENS INSIDE THE CLASS
            self.manager.current = 'HomeScreen'
        else:
            print(
                "Wrong Email or Password, Try Again")
            close_button = MDRaisedButton(text='Close', on_release=self.close_login_screen_window, font_size=16)
            self.login_screen_message_window = MDDialog(title='[color=990000]Warning Message[/color]', text=(
                '[color=ffffff]Wrong Email or Password, Try Again[/color]'), size_hint=(0.7, 1.0),
                                                        buttons=[close_button], auto_dismiss=False)
            # TO OPEN THE DIALOG WINDOW
            self.login_screen_message_window.open()

    def close_login_screen_window(self, obj):
        self.login_screen_message_window.dismiss()


# endregion <<<<==========================================[Login Screen]==========================================>>>>


# region <<<<===========================================[Home Screen]============================================>>>>
class HomeScreen(Screen):
    # THAT'S HOW TO CALL A FUNCTION JUST BY ENTERING THR SCREEN, NEED TO ADD IT ON SCREEN ABOVE AS WELL
    def on_pre_enter(self):
        # YOU PUT THE FUNCTION YOU WANT TO CALL TO DO SOME ACTION
        self.set_user_name()

    def set_user_name(self):
        # print(USERS_NAME_LIST)
        # global UserName
        # current user name who use the app
        global connected_user_name
        global trello_user_name
        # self.ids["UserName"].text = USERS_NAME_LIST[3]
        # CHECK IF EMAIL THAT ENTERED IS EXIST OF ONE OF LISTS (RWB_EMAIL or WISECO_EMAIL),
        # THEN FIND index OF EMAIL TO USE IT IN (USERS_NAME_LIST) list TO PRINT IT ON SCREEN
        if (self.manager.get_screen('LoginScreen').ids["Email"].text in rwb_email_address_list):
            # print(self.manager.get_screen('LoginScreen').ids["Email"].text)
            # TO FIND INDEX OF USER IN EXCEL FILE
            user_index = rwb_email_address_list.index(self.manager.get_screen('LoginScreen').ids["Email"].text)
            # print(USER_INDEX)
            self.ids["UserName"].text = users_name_list[user_index]
            connected_user_name = self.ids["UserName"].text
            trello_user_name = trello_users_name_list[user_index]
            # print(trello_user_name)

        elif (self.manager.get_screen('LoginScreen').ids["Email"].text in wiseco_email_address_list):
            # print(self.manager.get_screen('LoginScreen').ids["Email"].text)
            # TO FIND INDEX OF USER IN EXCEL FILE
            user_index = wiseco_email_address_list.index(self.manager.get_screen('LoginScreen').ids["Email"].text)
            # print(USER_INDEX)
            self.ids["UserName"].text = (users_name_list[user_index])
            connected_user_name = self.ids["UserName"].text
            trello_user_name = trello_users_name_list[user_index]
            # print(trello_user_name)

    def logout(self):
        # THAT'S HOW YOU CAN GO TO OTHER SCREENS INSIDE THE CLASS
        self.manager.current = 'LoginScreen'
        # TO RESET LOGIN FIELDS
        self.manager.get_screen('LoginScreen').ids["Email"].text = ""
        self.manager.get_screen('LoginScreen').ids["Password"].text = ""


# endregion <<<<==========================================[Home Screen]===========================================>>>>


# region <<<<=====================================[Pin Bore Machines Screen]=====================================>>>>
class PinBoreScreen(Screen):
    def still_work_on_it(self, obj):
        close_button = MDRaisedButton(text='Close', on_release=self.close_pin_bore_screen_window, font_size=16)
        self.pin_bore_screen_message_window = MDDialog(title='[color=cccc00]Warning Message[/color]', text=(
            '[color=ffffff]Still Work On It, Thanks for your Patience. [/color]'), size_hint=(0.7, 1.0),
                                                       buttons=[close_button], auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.pin_bore_screen_message_window.open()

    def close_pin_bore_screen_window(self, obj):
        print("Close_PinBore_Dialog" + " is called")
        # TO CLOSE THE DIALOGE
        self.pin_bore_screen_message_window.dismiss()


# endregion <<<<====================================[Pin Bore Machines Screen]====================================>>>>


# region <<<<============================[Old Horizontal Machines(28,29,32) Items List]===========================>>>>

# CREATE THE CLASS FOE ITEM THAT'S DISPLAY IN CONFIRMATION DIALOG
class OldHorizontalMachineItem(OneLineAvatarListItem):
    divider = None


# endregion <<<<==========================[Old Horizontal Machines(28,29,32) Items List]==========================>>>>


# region <<<<===========================[Old Horizontal Machines(28,29,32) Functions]============================>>>>

def create_old_horizontal_machine_program_in_original_folder(self):
    print("create_old_horizontal_machine_program_in_original_folder" + " FUNCTION is called")

    # DEFINE VARIABLES ON THIS FUNCTION AS global TO BE ABLE USE THEM OUT SIDE THE FUNCTION
    global original_folder_path_of_old_horizontal_machine
    original_folder_path_of_old_horizontal_machine = self.manager.get_screen('OldHorizontalSettingScreen').ids[
        "OriginalFolderPathOfOldHorizontalMachine"].text

    global new_horizontal_program_for_old_machine_in_original_folder
    global new_horizontal_program_To0_direction_for_old_machine_in_original_folder
    global new_horizontal_program_To180_direction_for_old_machine_in_original_folder

    # WE USE COPY METHOD TO COPY (HorizontalProgramLines) THAT'S CONTAIN ORIGINAL LINES BEFORE
    # MAKE 2 SEPARATE PROGRAMS To0 AND To180
    # WE MAKE NEW LISTS TO BE ABLE TO MODIFY THEM WITHOUT ADJUST THE ORIGINAL LIST(HorizontalProgramLines)
    # BECAUSE WE WANT TO USE IT LATER
    global horizontal_program_lines_To0_direction_for_old_machine_in_original_folder
    horizontal_program_lines_To0_direction_for_old_machine_in_original_folder = HorizontalProgramLines.copy()

    global horizontal_program_lines_To180_direction_for_old_machine_in_original_folder
    horizontal_program_lines_To180_direction_for_old_machine_in_original_folder = HorizontalProgramLines.copy()

    # TO CREATE THE NEW FILE ON THE PATH YOU WANT
    # open() WITH "x" IT will create a file, returns an error if the file exist(THAT WHY WE USE try/except)
    # IF IT DOES NOT EXIST IT WILL CREATE THE NEW FILE , IF IT IS EXIST IT WILL GO TO except BLOCK
    # AND CHECK IF NEED TO SAVE OVER THE EXISTING FILE
    try:
        print("try original is called")
        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):

            # To0 PRROGRAM

            # print("ORIGINAL_FOLDER_TO0")
            # print(HorizontalProgramLines)
            new_horizontal_program_To0_direction_for_old_machine_in_original_folder = (
                    original_folder_path_of_old_horizontal_machine + "\\" + "P" + new_program_number_for_old_horizontal_machine + "TO0.MIN")
            horizontal_program_lines_To0_direction_for_old_machine_in_original_folder[0] = (
                    '(PART ' + new_program_number_for_old_horizontal_machine + " TO 0" + ' -- ' + todaydate + ' <SYS>' + ')')

            try:
                create_new_horizontal_program_To0_direction_for_old_machine_in_original_folder = open(
                    new_horizontal_program_To0_direction_for_old_machine_in_original_folder, "x")
                create_new_horizontal_program_To0_direction_for_old_machine_in_original_folder.write(
                    '\n'.join(horizontal_program_lines_To0_direction_for_old_machine_in_original_folder))
                create_new_horizontal_program_To0_direction_for_old_machine_in_original_folder.close()
            # TO HANDLE ERROR OF NOT FINDING ORIGINAL FOLDER
            except PermissionError or FileNotFoundError as error:
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=ffff00] ORIGINAL [/color][/b]' +
                    "Folder location to Save the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]'
                    + str(error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Original Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

            # TO 180

            # print("ORIGINAL_FOLDER_TO180")
            # print(HorizontalProgramLines)

            new_horizontal_program_To180_direction_for_old_machine_in_original_folder = (
                    original_folder_path_of_old_horizontal_machine + "\\" + "P" + new_program_number_for_old_horizontal_machine + "TO180.MIN")
            horizontal_program_lines_To180_direction_for_old_machine_in_original_folder[0] = (
                    '(PART ' + new_program_number_for_old_horizontal_machine + " TO 180" + ' -- ' + todaydate + ' <SYS>' + ')')
            horizontal_program_lines_To180_direction_for_old_machine_in_original_folder[VC155_VARIABLE_INDEX] = (
                    'VC155=' + format((-1) * offset_amount) + '  (Offset)')
            # print("OFFSET_AMOUNT for original folder", OFFSET_AMOUNT)
            horizontal_program_lines_To180_direction_for_old_machine_in_original_folder[VC173_VARIABLE_INDEX] = (
                    'VC173=' + format(float(Y_distance_from_origin_to_circlip_notch) - (2 * offset_amount),
                                      '.4f') + '  (yCirclipNotch)')

            try:
                create_new_horizontal_program_To180_direction_for_old_machine_in_original_folder = open(
                    new_horizontal_program_To180_direction_for_old_machine_in_original_folder, "x")
                create_new_horizontal_program_To180_direction_for_old_machine_in_original_folder.write(
                    '\n'.join(horizontal_program_lines_To180_direction_for_old_machine_in_original_folder))
                create_new_horizontal_program_To180_direction_for_old_machine_in_original_folder.close()
            # TO HANDLE ERROR OF NOT FINDING ORIGINAL FOLDER
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=ffff00] ORIGINAL [/color][/b]' +
                    "Folder location to Save the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]'
                    + str(error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Original Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

        # IF OFFSET DIRECTION IS NOT EACHWAY
        else:

            new_horizontal_program_for_old_machine_in_original_folder = (
                    original_folder_path_of_old_horizontal_machine + "\\" + "P" + new_program_number_for_old_horizontal_machine + ".MIN")
            try:
                create_new_horizontal_program_for_old_machine_in_original_folder = open(
                    new_horizontal_program_for_old_machine_in_original_folder, "x")
                create_new_horizontal_program_for_old_machine_in_original_folder.write(
                    '\n'.join(HorizontalProgramLines))
                create_new_horizontal_program_for_old_machine_in_original_folder.close()
            # TO HANDLE ERROR OF NOT FINDING ORIGINAL FOLDER
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=ffff00] ORIGINAL [/color][/b]' +
                    "Folder location to Save the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]'
                    + str(error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Original Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

        #  TO ADD THE CONFIRMATION DETAILS THAT WAS NEEDED TO CREATE THE PROGRAM
        if (old_horizontal_program_confirmation_email_message_list != []):
            email_messages_of_creating_old_horizontal_machine_program.append(
                "\n".join(old_horizontal_program_confirmation_email_message_list) + "\n")
        success_messages_of_creating_old_horizontal_machine_program = \
            ["\n" + "Program has been **CREATED** successfully in **ORIGINAL** Folder." + "\n"]

        email_messages_of_creating_old_horizontal_machine_program.append("\n".join(
            success_messages_of_creating_old_horizontal_machine_program))

        close_button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_window_of_original_folder,
                                      font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
                '[b][i][u][color=0099ff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' +
                '[color=ffffff] Program Has been Created Successfully in [/color]' + '[color=ffff00]ORIGINAL[/color]' +
                '[color=ffffff] Folder.[/color]'), size_hint=(0.7, 1.0), buttons=[close_button], auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        print()
        # print('\n'.join(HorizontalProgramLines))
    # IF PROGRAM IS EXIST IT WILL CHECK IF NEED TO SAVE OVER THE EXISTING FILE
    except(FileExistsError):
        print("exept original is called")
        yes_button = MDRaisedButton(text='Yes',
                                    on_release=self.replace_existing_old_horizontal_machine_program_in_original_folder,
                                    font_size=16)
        no_button = MDRaisedButton(text='No', on_release=self.close_old_horizontal_window_of_original_folder,
                                   font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=0066ff]Confirmation Message[/color]',
                                                      text=('[b][i][u][color=0099ff]' + self.ids[
                                                          "JobNumber"].text + '[/color][/u][/i][/b]' +
                                                            '[color=ffffff] Program already Exists in [/color]' +
                                                            '[color=ffff00]ORIGINAL[/color]' +
                                                            '[color=ffffff] Folder.[/color]' + '\n' +
                                                            '[color=ffffff]Do you want to replace it ?[/color]'),
                                                      size_hint=(0.7, 1.0), buttons=[yes_button, no_button],
                                                      auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()

        # MAYBE WE DO NOT NEED IT
        # TO CLEAR THE LIST TO BE EMPTY IF ALREADY THE PROGRAM EXIST
        success_messages_of_creating_old_horizontal_machine_program = []
        # maybe need logic to send this message to trello in case there is warnning make this program
        email_messages_of_creating_old_horizontal_machine_program.append("\n".join(
            success_messages_of_creating_old_horizontal_machine_program))


def create_old_horizontal_machine_program_in_running_folder(self):
    print("create_old_horizontal_machine_program_in_running_folder" + " FUNCTION is called")
    global running_folder_path_of_old_horizontal_machine
    running_folder_path_of_old_horizontal_machine = self.manager.get_screen('OldHorizontalSettingScreen').ids[
        "RunningFolderPathOfOldHorizontalMachine"].text
    global new_horizontal_program_for_old_machine_in_running_folder
    global new_horizontal_program_To0_direction_for_old_machine_in_running_folder
    global new_horizontal_program_To180_direction_for_old_machine_in_running_folder

    global horizontal_program_lines_To0_direction_for_old_machine_in_running_folder
    horizontal_program_lines_To0_direction_for_old_machine_in_running_folder = HorizontalProgramLines.copy()

    global horizontal_program_lines_To180_direction_for_old_machine_in_running_folder
    horizontal_program_lines_To180_direction_for_old_machine_in_running_folder = HorizontalProgramLines.copy()

    try:
        print("try running is called")
        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):
            # TO 0
            # print("RUNNING_FOLDER_TO0")
            # print(HorizontalProgramLines)
            new_horizontal_program_To0_direction_for_old_machine_in_running_folder = (
                    running_folder_path_of_old_horizontal_machine + "\\" + "P" + new_program_number_for_old_horizontal_machine + "TO0.MIN")
            horizontal_program_lines_To0_direction_for_old_machine_in_running_folder[0] = (
                    '(PART ' + new_program_number_for_old_horizontal_machine + " TO 0" + ' -- ' + todaydate + ' <SYS>' + ')')

            try:
                create_new_horizontal_program_To0_direction_for_old_machine_in_running_folder = open(
                    new_horizontal_program_To0_direction_for_old_machine_in_running_folder, "x")
                create_new_horizontal_program_To0_direction_for_old_machine_in_running_folder.write(
                    '\n'.join(horizontal_program_lines_To0_direction_for_old_machine_in_running_folder))
                create_new_horizontal_program_To0_direction_for_old_machine_in_running_folder.close()
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=33cc33] RUNNING [/color][/b]' + "Folder location to Save the Program."
                    + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]' + str(
                        error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Running Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

            # TO 180
            # print("RUNNING_FOLDER_TO180")
            # print(HorizontalProgramLines)
            new_horizontal_program_To180_direction_for_old_machine_in_running_folder = (
                    running_folder_path_of_old_horizontal_machine + "\\" + "P" + new_program_number_for_old_horizontal_machine + "TO180.MIN")
            horizontal_program_lines_To180_direction_for_old_machine_in_running_folder[0] = (
                    '(PART ' + new_program_number_for_old_horizontal_machine + " TO 180" + ' -- ' + todaydate + ' <SYS>' + ')')
            horizontal_program_lines_To180_direction_for_old_machine_in_running_folder[VC155_VARIABLE_INDEX] = (
                    'VC155=' + format((-1) * offset_amount) + '  (Offset)')
            horizontal_program_lines_To180_direction_for_old_machine_in_running_folder[VC173_VARIABLE_INDEX] = (
                    'VC173=' + format(float(Y_distance_from_origin_to_circlip_notch) - (2 * offset_amount),
                                      '.4f') + '  (yCirclipNotch)')

            try:
                create_new_horizontal_program_To180_direction_for_old_machine_in_running_folder = open(
                    new_horizontal_program_To180_direction_for_old_machine_in_running_folder, "x")
                create_new_horizontal_program_To180_direction_for_old_machine_in_running_folder.write(
                    '\n'.join(horizontal_program_lines_To180_direction_for_old_machine_in_running_folder))
                create_new_horizontal_program_To180_direction_for_old_machine_in_running_folder.close()
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=33cc33] RUNNING [/color][/b]' + "Folder location to Save the Program."
                    + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]' + str(
                        error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Running Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

        else:
            new_horizontal_program_for_old_machine_in_running_folder = (
                    running_folder_path_of_old_horizontal_machine + "\\" + "P" + new_program_number_for_old_horizontal_machine + ".MIN")
            try:
                create_new_horizontal_program_for_old_machine_in_running_folder = open(
                    new_horizontal_program_for_old_machine_in_running_folder, "x")
                create_new_horizontal_program_for_old_machine_in_running_folder.write('\n'.join(HorizontalProgramLines))
                create_new_horizontal_program_for_old_machine_in_running_folder.close()
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=33cc33] RUNNING [/color][/b]' + "Folder location to Save the Program."
                    + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]' + str(
                        error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Running Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

        # JUST FOR NOW
        success_messages_of_creating_old_horizontal_machine_program = \
            ["\n" + "Program has been **CREATED** successfully in **RUNNING** Folder."]
        email_messages_of_creating_old_horizontal_machine_program.append("\n".join(
            success_messages_of_creating_old_horizontal_machine_program))

        close_button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_window_of_running_folder,
                                      font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
                '[b][i][u][color=0099ff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' +
                '[color=ffffff] Program Has been Created Successfully in [/color]' +
                '[color=33cc33]RUNNING[/color]' + '[color=ffffff] Folder.[/color]' + '\n' +
                '[color=ffffff]After closing this window, the program will open on CIMCO Editor.[/color]' + '\n' +
                '[color=ffffff]Please Double Check and Report any Problem on Trello.[/color]'),
                                                      size_hint=(0.7, 1.0), buttons=[close_button], auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        print()
        # print('\n'.join(HorizontalProgramLines))
    except(FileExistsError):
        print("exept running is called")
        # print(" Program is Exist in Running Folder")
        yes_button = MDRaisedButton(text='Yes',
                                    on_release=self.replace_existing_old_horizontal_machine_program_in_running_folder,
                                    font_size=16)
        no_button = MDRaisedButton(text='No', on_release=self.close_old_horizontal_screen_window, font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=0066ff]Confirmation Message[/color]',
                                                      text=('[b][i][u][color=0099ff]' + self.ids[
                                                          "JobNumber"].text + '[/color][/u][/i][/b]' +
                                                            '[color=ffffff] Program already Exists in [/color]' +
                                                            '[color=33cc33]Running[/color]' +
                                                            '[color=ffffff] Folder.[/color]' + '\n' +
                                                            '[color=ffffff]Do you want to replace it ?[/color]'),
                                                      size_hint=(0.7, 1.0), buttons=[yes_button, no_button],
                                                      auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        # MAYBE WE DO NOT NEED IT
        # TO CLEAR THE LIST TO BE EMPTY IF ALREADY THE PROGRAM IS EXIST
        success_messages_of_creating_old_horizontal_machine_program = []
        email_messages_of_creating_old_horizontal_machine_program.append("\n".join(
            success_messages_of_creating_old_horizontal_machine_program))


def need_confirmation_to_create_old_horizontal_machine_program(self, title, sub_function, dialog_type, content):
    print("need_confirmation_to_create_old_horizontal_machine_program" + " FUNCTION is called")
    print('\n'.join(verification_messages_of_creating_old_horizontal_machine_program))
    # still need to add action to do
    enter_button = MDRaisedButton(text='Enter', on_press=sub_function,
                                  on_release=self.create_program_for_old_horizontal_machine,
                                  font_size=16)
    close_button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_screen_window, font_size=16)
    # IF CONFIRMATION NEED TO ENTER VALUE
    if (dialog_type == "custom"):
        self.Old_Horizontal_Message_Dialog = MDDialog(title=title, type=dialog_type, content_cls=content,
                                                      size_hint=(0.7, 1.0), buttons=[enter_button, close_button],
                                                      auto_dismiss=False)
    # IF CONFIRMATION NEED TO CHOOSE VALUE FROM OPTION
    elif (dialog_type == "confirmation"):
        self.Old_Horizontal_Message_Dialog = MDDialog(title=title, type=dialog_type, items=content,
                                                      size_hint=(0.7, 1.0), buttons=[enter_button, close_button],
                                                      auto_dismiss=False)
    # TO OPEN THE DIALOG WINDOW
    self.Old_Horizontal_Message_Dialog.open()


def old_horizontal_machine_program_needs_attention(self):
    print("old_horizontal_machine_program_needs_attention" + " FUNCTION is called")
    close_button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_screen_window, font_size=16)
    self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=cccc00]Warning Message[/color]', text=(
            '[color=ffffff]' + '\n'.join(warning_messages_of_creating_old_horizontal_machine_program) + '[/color]'),
                                                  size_hint=(0.7, 1.0), buttons=[close_button], auto_dismiss=False)
    # TO OPEN THE DIALOG WINDOW
    self.Old_Horizontal_Message_Dialog.open()


def failed_to_create_old_horizontal_machine_program(self):
    print("failed_to_create_old_horizontal_machine_program" + " FUNCTION is called")
    # print('\n'.join(fail_messages_of_creating_old_horizontal_machine_program))
    close_button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_screen_window, font_size=16)
    self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=990000]Warning Message[/color]', text=(
            '[color=ffffff]' + '\n'.join(fail_messages_of_creating_old_horizontal_machine_program) + '[/color]'),
                                                  size_hint=(0.7, 1.0), buttons=[close_button], auto_dismiss=False)
    # TO OPEN THE DIALOG WINDOW
    self.Old_Horizontal_Message_Dialog.open()


def send_email_about_create_old_horizontal_machine_program(self):
    try:
        service_app = "outlook.office365.com"
        # SERVER NAME
        smtp_server = "smtp.office365.com"
        port = 587
        # IT WILL USE EMAIL OF USER WHO LOGIN TO THE APP TO SENT THE EMAIL TO THE TRELLO BOARD
        sender_email = user_email_address
        # USING KEYRING PACKAGE TO GET PASSWORD FROM Windows Credential Manager WHILE THEY ARE SAVING IN USER COMPUTER
        sender_password = keyring.get_password(service_app, sender_email)
        # just for now
        # recipient_email = "moemenatweh@hotmail.com"
        trello_board_email = self.manager.get_screen('AppSettingScreen').ids[
            "TrelloEmailAddress"].text  # moemenalatweh1+sqa4wcni54jz6erwnpbj@boards.trello.com
        # creates SMTP session
        email_server = smtplib.SMTP(smtp_server, port)
        # TLS for security
        email_server.starttls()

        # authentication
        # compiler gives an error for wrong credential.
        email_server.login(sender_email, sender_password)

        #########
        if (fail_messages_of_creating_old_horizontal_machine_program != []):
            card_title = "Failed to Create Program for " + new_program_number_for_old_horizontal_machine
            card_label = "#FAILED"
            card_member = "@moemenalatweh1 " + "@" + trello_user_name
        elif (warning_messages_of_creating_old_horizontal_machine_program != []):
            card_title = new_program_number_for_old_horizontal_machine + " Program Needs Attention."
            card_label = "#WARNNING"
            card_member = "@moemenalatweh1 " + "@" + trello_user_name
        else:
            card_title = new_program_number_for_old_horizontal_machine
            card_label = "#SUCCESS"
            card_member = "@moemenalatweh1 " + "@" + trello_user_name

        #         # message to be sent to personal email
        #         email_message = f"""From: Alatweh Moemen <malatweh@rwbteam.com>
        # To: <moemenatweh@hotmail.com>
        # Subject: Testing Email by python
        #
        #
        # {"".join(email_messages_of_creating_old_horizontal_machine_program)}"""
        #
        #         email_Server.sendmail(sender_email, recipient_email, email_message)

        # message to be sent trello email
        email_message = f"""From: Alatweh Moemen <malatweh@rwbteam.com>
Subject: {card_title}  {card_label}  {card_member}


{"".join(email_messages_of_creating_old_horizontal_machine_program)}"""

        # TO SEND THE EMAIL
        email_server.sendmail(sender_email, trello_board_email, email_message)

        # terminating the session
        email_server.quit()

    except Exception as error:
        fail_messages_of_creating_old_horizontal_machine_program.append(
            "Failed to send Email to Trello board." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]' + str(
                error) + '[/color]' + "\n" + "Double Check Network and Login Authentication.")
        # email_messages_of_creating_old_horizontal_machine_program.append("\n".join(fail_messages_of_creating_old_horizontal_machine_program))
        failed_to_create_old_horizontal_machine_program(self)


# endregion <<<<=========================[Old Horizontal Machines(28,29,32) Functions]===========================>>>>


# region <<<<=============================[Old Horizontal Machines(28,29,32) Screen]=============================>>>>

class OldHorizontalScreen(Screen):
    # CALL (FOUR_CYCLE_PINBORE_VARIABLES) FUNCTION TO SET THE VARIABLES
    four_cycle_pin_bore_variables()

    def create_program_for_old_horizontal_machine(self, obj):
        print("create_program_for_old_horizontal_machine" + " FUNCTION is called")

        # print("offset value before OR AFTER ", OFFSET_AMOUNT)

        global success_messages_of_creating_old_horizontal_machine_program
        success_messages_of_creating_old_horizontal_machine_program = []

        global verification_messages_of_creating_old_horizontal_machine_program
        verification_messages_of_creating_old_horizontal_machine_program = []

        global fail_messages_of_creating_old_horizontal_machine_program
        fail_messages_of_creating_old_horizontal_machine_program = []

        global warning_messages_of_creating_old_horizontal_machine_program
        warning_messages_of_creating_old_horizontal_machine_program = []

        global email_messages_of_creating_old_horizontal_machine_program
        email_messages_of_creating_old_horizontal_machine_program = []
        email_messages_of_creating_old_horizontal_machine_program.append(self.ids["JobNumber"].text + " Program on " +
                                                                         "Old Horizontal Machine" + "\n" +
                                                                         "Created by : " + connected_user_name + "\n")

        # # **************************************#
        # # CALL (LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION) FUNCTION TO LOAD HORIZONTAL TOOL LIST SHEETS
        # LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION(self)

        # **************************************#
        # CALL (FOUR_CYCLE_PINBORE_VARIABLES) FUNCTION TO SET THE VARIABLES
        # if (self.ids["JobNumber"].text == ""):
        #      FOUR_CYCLE_PINBORE_VARIABLES()

        # self.ids["JobNumber"].text: TO ACCESS TEXT FIELD WE USE ids THAT'S DEFINED ABOVE
        global new_program_number_for_old_horizontal_machine
        new_program_number_for_old_horizontal_machine = self.ids["JobNumber"].text
        print(new_program_number_for_old_horizontal_machine)
        if (self.ids["JobNumber"].text == ""):
            fail_messages_of_creating_old_horizontal_machine_program.append("Please Enter Job Number.")
            failed_to_create_old_horizontal_machine_program(self)
            return

        # NEED LOGIC TO CHECK IF NUMBER THAT ENTERED MATCH WITH SPEC DATABASE

        try:
            # print("try LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
            # CALL (LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION) FUNCTION TO LOAD HORIZONTAL TOOL LIST SHEETS
            load_horizontal_machine_tool_list_sheets(self)
        except Exception as error:
            # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
            fail_messages_of_creating_old_horizontal_machine_program.append(
                "Failed to Find, Load, or Access" + '[b][u][color=ffffff] Horizontal Tool List [/color][/u][/b]' +
                "File to Create the Program." + "\n" + "An Error has occurred " + "\n" + '[color=ff1a1a]' + str(
                    error) + '[/color]' + "\n" + "Double Check Network, and File Location.")
            email_messages_of_creating_old_horizontal_machine_program.append(
                "Failed to Find, Load, or Access Horizontal Tool List File to Create the Program." + "\n" +
                "Double Check Network, and File Location.")
            failed_to_create_old_horizontal_machine_program(self)
            return

        # region <<<<=====================================[PROBE PROGRAMS]=====================================>>>>

        # self.manager.get_screen('SettingScreen'): SCREEN WE WANT TO ACCESS
        # ids["ProbePrograms"].text :
        # TO GET THE TEXT OF ["ProbePrograms"] THAT'S CONTAIN PROBE PROGRAMS FILE PATH TO SEARCH INSIDE THE FILE
        probe_programs_folder_path_of_old_horizontal_machine = self.manager.get_screen(
            'OldHorizontalSettingScreen').ids["ProbePrograms"].text
        print(probe_programs_folder_path_of_old_horizontal_machine)
        # DEFINE AN EMPTY LIST TO STORE RESULT OF PROBE PROGRAM SEARCH.
        try:
            result_of_probe_program_search_for_old_horizontal_machine = []
            for file in glob.glob(probe_programs_folder_path_of_old_horizontal_machine + '*\*' + forging_number + '*'):
                # RESULT_OF_PROBE_PROGRAM_SEARCH = []
                # print(file)
                result_of_probe_program_search_for_old_horizontal_machine.append(file)
            print(len(result_of_probe_program_search_for_old_horizontal_machine))
            print(result_of_probe_program_search_for_old_horizontal_machine)

            # MAYBE WE DON'T NEED IT
            # DEFINE THAT JUST TO AVOID ERROR IF RESULT OF SEARCH IS MORE THAN ONE
            # probe_program_of_old_horizontal_machine = 'OXXXX'
            # DEFINE AN EMPTY LIST TO ADD PROBE PROGRAM LINES ONE BY ONE.
            probe_programs_lines_of_old_horizontal_machine = []
            # IF ONE PROBE PROGRAM FOUND , GO A HEAD AND OPEN IT AND TAKE FIRST LINE OF
            # PROBE PROGRAM TO USE IT IN HORIZ TEMPLATE
            if ((len(result_of_probe_program_search_for_old_horizontal_machine) == 1) and (forging_number != "")):
                print("RESULT_OF_PROBE_PROGRAM_SEARCH ", result_of_probe_program_search_for_old_horizontal_machine)
                # print(RESULT_OF_PROBE_PROGRAM_SEARCH[0])
                with open(result_of_probe_program_search_for_old_horizontal_machine[0], 'rt') as CurrentProgram:
                    for line in CurrentProgram:  # For each line in the file,
                        # strip newline and add to list.
                        probe_programs_lines_of_old_horizontal_machine.append(line.rstrip('\n'))
                    # print(probe_programs_lines_of_old_horizontal_machine)
                    # print("LINE OF PROBE PROGRAM THAT NEED TO ADD TO HORIZONTAL PROGRAM: " +
                    # probe_programs_lines_of_old_horizontal_machine[0])
                    probe_program_of_old_horizontal_machine = probe_programs_lines_of_old_horizontal_machine[0]
            # maybe we don't need it
            elif (probe_programs_folder_path_of_old_horizontal_machine == ""):
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find, Load, or Access" + '[b][u][color=ffffff] Probe Programs [/color][/u][/b]' +
                    "File to Create the Program." + "\n" + "Double Check Network, and File Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find, Load, or Access Probe Programs File to Create the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return
            elif (forging_number == ""):
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Forging Number does NOT found." + '\n' + "Double Check Job Spec with Engineering and Try Again.")
                email_messages_of_creating_old_horizontal_machine_program.append("\n".join(
                    fail_messages_of_creating_old_horizontal_machine_program))
                failed_to_create_old_horizontal_machine_program(self)
                return
            # IF NO PROGRAM FOUND, WE NEED TO CHECK MANUALLY IF IT IS THERE ,
            # OTHERWISE CREATE NEW PROBE PROGRAM AND TRY AGAIN
            elif ((len(result_of_probe_program_search_for_old_horizontal_machine) == 0) and (forging_number != "")):
                # WE NEED TO CREATE DIALOG TO POP UP THE MESSAGE FOR USER
                fail_messages_of_creating_old_horizontal_machine_program = [
                    "Probe Program does NOT found, Double Check Probe Programs Folder." + '\n' +
                    "If it is NOT there, Create new Probe Program and Try Again."
                    + '\n' + '\n' + "OTHERWISE Double Check Network, and File Location."]
                print(fail_messages_of_creating_old_horizontal_machine_program)
                email_messages_of_creating_old_horizontal_machine_program.append("\n".join(
                    fail_messages_of_creating_old_horizontal_machine_program))
                failed_to_create_old_horizontal_machine_program(self)
                return
            elif ((len(result_of_probe_program_search_for_old_horizontal_machine) > 1) and (forging_number != "")):
                # WE NEED TO CREATE DIALOG TO POP UP THE MESSAGE FOR USER
                if (len(result_of_probe_program_search_for_old_horizontal_machine) <= 20):
                    fail_messages_of_creating_old_horizontal_machine_program = [
                        "Many Probe Programs found for this Forging :" + '\n' + '\n' +
                        ('\n'.join(result_of_probe_program_search_for_old_horizontal_machine)) + '\n' + '\n' +
                        "Fix the Confusion and Try Again."]
                else:
                    fail_messages_of_creating_old_horizontal_machine_program = [
                        "Many Probe Programs found for this Forging" + '\n' + "Fix the Confusion and Try Again."]
                print(fail_messages_of_creating_old_horizontal_machine_program)
                email_messages_of_creating_old_horizontal_machine_program.append("\n".join(
                    fail_messages_of_creating_old_horizontal_machine_program))
                print(len(result_of_probe_program_search_for_old_horizontal_machine))
                print(result_of_probe_program_search_for_old_horizontal_machine)
                failed_to_create_old_horizontal_machine_program(self)
                return
            # else:
            #     fail_messages_of_creating_old_horizontal_machine_program = ["UNEXPECTED issue about Probe Programs."
            #     + '\n' + "Double Check Network and probe program folder and Try Again."]
            #     print(fail_messages_of_creating_old_horizontal_machine_program)
            #     email_messages_of_creating_old_horizontal_machine_program.append("\n".join(fail_messages_of_creating_old_horizontal_machine_program))
            #     Failed_to_Create_Horizontal_Program(self)
            #     return
        except Exception as error:
            # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
            fail_messages_of_creating_old_horizontal_machine_program.append(
                "Failed to Find, Load, or Access" + '[b][u][color=ffffff] Probe Programs [/color][/u][/b]' +
                "File to Create the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]' + str(
                    error) + '[/color]' + "\n" + "Double Check Network, and File Location.")
            email_messages_of_creating_old_horizontal_machine_program.append(
                "Failed to Find, Load, or Access Probe Programs File to Create the Program." + "\n" +
                "Double Check Network, and File Location.")
            failed_to_create_old_horizontal_machine_program(self)
            return
        # endregion <<<<====================================[PROBE PROGRAMS]====================================>>>>

        # region  <<<<============================[Old Horizontal For Loop Template]============================>>>>

        # DEFINE VARIABLE(HORIZONTAL_TEMPLATE_PATH) TO STORE HorizontalTemplate PATH FROM SETTING SCREEN
        # TO GET VALUE FROM ANOTHER SCREEN(SettingScreen IN THIS EXAMPLE) TO USE IT IN ANOTHER SCREEN CLASS (OldHorizontalScreen IN THIS EXAMPLE) FOR LOGIC PURPOSES>>
        # >> USE (self.manager.get_screen('SettingScreen').ids["HorizontalTemplate"].text)
        # self.manager.get_screen('SettingScreen'): PUT SCREEN YOU WANT TO ACCESS
        # ids["HorizontalTemplate"].text : PUT WIDGET OR VALUE YOU NEED TO GET(IN THIS EXAMPLE WE GET THE TEXT OF MDTextField THAT NAME "HorizontalTemplate")
        HORIZONTAL_TEMPLATE_PATH = self.manager.get_screen('OldHorizontalSettingScreen').ids["HorizontalTemplate"].text
        ###LEAVE IT FOR TEST
        print(HORIZONTAL_TEMPLATE_PATH)

        # DEFINE list TO CONTAIN HORIZONTAL PROGRAM
        global HorizontalProgramLines
        HorizontalProgramLines = []
        # TO OPEN HORIZONTAL_TEMPLATE AND ADD EACH SINGLE LINE TO list WE CREATE ABOVE(HorizontalProgramLines)
        try:
            with open(HORIZONTAL_TEMPLATE_PATH, 'rt') as CurrentProgram:
                # try:
                for line in CurrentProgram:  # For each line in the file,
                    HorizontalProgramLines.append(line.rstrip('\n'))  # strip newline and add to list.
                # except Exception as error:
                #     # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                #     fail_messages_of_creating_old_horizontal_machine_program.append(
                #         "Failed to Find, Load, or Access Horizontal Template File to Create the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]' + str(
                #             error) + '[/color]' + "\n" + "Double Check Network, and File Location.")
                #     email_messages_of_creating_old_horizontal_machine_program.append(
                #         "Failed to Find, Load, or Access Horizontal Template File to Create the Program." + "\n" + "Double Check Network, and File Location.")
                #     Failed_to_Create_Horizontal_Program(self)
                #     return
                ###LEAVE IT FOR TEST
                ## TO PRINT ORIGINAL HORIZONTAL TEMPLATE
                # print(HorizontalProgramLines)
                # print()
                # print('\n'.join(HorizontalProgramLines))
                # print()
                ## TO REPLACE THE FIRST ITEM OF LIST WITH JOB NUMBER AND TODAYDATE
                HorizontalProgramLines[0] = \
                    ('(PART ' + new_program_number_for_old_horizontal_machine + ' -- ' + todaydate + ' <SYS>' + ')')
                ## WE CAN USE CODE BELOW TO ADD THE TOOL LIST AFTER CHECK EACH TOOL LOGIC
                # DEFINE VARIABLE TO FINISH THE while LOOP OF TOOL LIST
                HORIZONTAL_PROGRAM_TOOL_LIST_END = 0
                for line in HorizontalProgramLines:
                    while HORIZONTAL_PROGRAM_TOOL_LIST_END == 0:
                        # WE CAN USE ((**********TOOL LIST**********)) LINE IN HORIZONTAL TEMPLATE TO LOCATE THE index OF list TO START ADD OTHER TOOLS(IF APPLICABLE)(BY ADD 1 TO INDEX WE FOUND)
                        if '(**********TOOL LIST**********)' in HorizontalProgramLines:
                            # TO FIND THE INDEX OF ITEM INSIDE LIST
                            index = HorizontalProgramLines.index('(**********TOOL LIST**********)')
                            ###LEAVE IT FOR TEST
                            # print(index)
                            index += 1
                        else:
                            print("CHECK HORIZONTAL TEMPLATE AND CORRECTED TO HAVE '(**********TOOL LIST**********)'")
                        # JUST FOR NOW, NEED TO ADD LOGIC TO CHECK IF THE FORGING HAS PILOT DIAMETER WITH 2.25 OR SMALLER TO 1.70
                        PILOT_DIAMETER = 2.25
                        PILOT_AVAILABILITY_STATUS = 0
                        if (PILOT_DIAMETER == 2.25):
                            PILOT_AVAILABILITY_STATUS = 1
                            # TO ADD TOOL DESCRIPTION FROM TOOL LIST FROM EXCEL FILE,
                            # NEED TO FIND THE INDEX OF THE TOOL IN THE TOOL LIST IN EXCEL FILE BY PUT THE TOOL USAGE FROM 1ST COLUMN IN MISCELLANEOUS SHEET(WHICH IS: '2.250 PILOT BORE FOR 2.25 DIA')
                            # STORE THE INDEX OF THE TOOL IN VARIABLE(MISCELLANEOUS_TOOL_LIST_INDEX)
                            # INSERT THE NEW ELEMENT TO HorizontalProgramLines LIST BY USE insert METHOD, AND index THAT FOUND FROM PREVIOSE if STATEMENT
                            # AND USE PANDAS METHOD TO ACCESS EXCEL FILE(HORIZONTAL_TOOL_LIST_FILE), AND SHEET('MISCELLANEOUS_TOOL_LIST')>>
                            # >> WITH INDEX OF THE TOOL(MISCELLANEOUS_TOOL_LIST_INDEX) WE FOUND, AND COLUMN NAME(DESCRIPTION(FOR_MACHINES_27/28/32))
                            MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index(
                                '2.250 PILOT BORE FOR 2.25 DIA')
                            HorizontalProgramLines.insert(index,
                                                          horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                                                              MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # HorizontalProgramLines.insert(index, '(T01 IS A 2.250 PILOT BORE)')
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                            # NEED TO PUT THIS LINE IN THE LAST if STATEMENT IN THIS while LOOP TO AVOID THE for LOOP STATRT AGAIN
                            ## HorizontalProgramEnd = 1
                            # NEED TO MAKE SURE IF TEMPLATE OF 1.70 PILOT DIAMETER IS DEFFERENT FROM TEMPLATE OF 2.25 PILOT DIAMETER
                        elif (PILOT_DIAMETER == 1.70):
                            PILOT_AVAILABILITY_STATUS = 1
                            MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index(
                                '1.70 PILOT BORE FOR 1.70 DIA')
                            HorizontalProgramLines.insert(index,
                                                          horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                                                              MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # HorizontalProgramLines.insert(index, '(T57 IS A 1.7 PILOT BORE)')
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                        # NEED TO ADD LOGIC OF LEDGE TOOL THAT WILL CHECK IF (DISTANCE_FROM_ORIGIN_TO_HIGHEST_POINT_OF_BORE) SMALLER THAN>>
                        # >>((X) Outsd Ring Belt Ht FROM EMSS: IT IS DISTANCE FROM ORIGIN TO END POINT BEFORE BOSS)
                        ##OUTSIDE_RING_BELT_HIEGHT= 1  # JUST FOR NOW NEED TO HAVE THIS VALUE FROM FORGING DATABASE
                        ##DISTANCE_FROM_ORIGIN_TO_HIGHEST_POINT_OF_BORE = PILOT_TO_DOME_DEPTH- PILOT_TO_PIN_DISTANCE - PIN_HOLE_DIAMETER
                        # just for now                                              # =============NEED TO FIX HERE================
                        # global LEDGE_CUT_AVAILABILITY_STATUS
                        if (ledge_cut_availability_status == ""):
                            verification_messages_of_creating_old_horizontal_machine_program = [
                                "Ledge status can't detect," + '\n' + "Does this job need to use Ladge Tool ?"]
                            print(verification_messages_of_creating_old_horizontal_machine_program)
                            old_horizontal_program_confirmation_email_message_list.append(
                                "# It was need confirmation of :" + "\n" + "\n".join(
                                    verification_messages_of_creating_old_horizontal_machine_program))
                            self.title = '[color=0066ff]Confirmation Message[/color] ' + '\n' + '\n' + '[b][i][color=ffffff]' + \
                                         verification_messages_of_creating_old_horizontal_machine_program[
                                             0] + '[/color][/i][/b]'
                            # _Option
                            self.Need_Ledge_Tool_Option = OldHorizontalMachineItem(text="Yes, it is Need",
                                                                                   on_release=self.need_to_use_ledge_tool)
                            self.Does_Not_Need_Ledge_Tool_Option = OldHorizontalMachineItem(text="No, it Does NOT Need",
                                                                                            on_release=self.does_not_need_to_use_ledge_tool)

                            self.items = [self.Need_Ledge_Tool_Option, self.Does_Not_Need_Ledge_Tool_Option]
                            need_confirmation_to_create_old_horizontal_machine_program(self, self.title,
                                                                                       self.decide_ledge_tool_status,
                                                                                       "confirmation", self.items)

                            print("LEDGE_CUT_AVAILABILITY_STATUS in create function: ", ledge_cut_availability_status)
                            return

                        if (ledge_cut_availability_status == 1 or ledge_counterbore_diameter != 0):
                            if (pin_hole_diameter >= 0.629):
                                MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index('LEDGE TOOL 0.625 DIA')
                                HorizontalProgramLines.insert(index,
                                                              horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                                                                  MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                                index += 1
                                ##print('\n'.join(HorizontalProgramLines))
                            elif (pin_hole_diameter < 0.629):
                                MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index('LEDGE TOOL 0.375 DIA')
                                HorizontalProgramLines.insert(index,
                                                              horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                                                                  MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])

                                # WE NEED TO REPLACE THE TOOL ON ((*******CONSTANTS********)) SECTION ON TEMPLATE
                                # FIND LOCATION OF NOTE OF LEDGE TOOL ON HorizontalProgramLines List
                                LEDGE_TOOL_NOTE_INDEX = HorizontalProgramLines.index(
                                    '(T06 IS THE STD .625 CARBIDE END MILL LEDGE TOOL)')
                                # REPLACE THE NOTE TO USE TOOL #16
                                HorizontalProgramLines[LEDGE_TOOL_NOTE_INDEX] = ('(T16 IS A 3/8 END MILL - LEDGE TOOL)')

                                # FIND LOCATION OF VARIABLE ON HorizontalProgramLines List
                                VC103_VARIABLE_INDEX = HorizontalProgramLines.index('VC103=06')
                                # REPLACE THE TOOL TO USE TOOL #16
                                HorizontalProgramLines[VC103_VARIABLE_INDEX] = ('VC103=16')
                                index += 1
                                ##print('\n'.join(HorizontalProgramLines))

                        # LOGIC FOR ADD THE ROUGH BORE TOOL TO THE TOOL LIST IN THE PROGRAM AND TO THE TOOLS SECTION
                        if (0.4500 <= pin_hole_diameter < 0.4960):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('11MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.4960 <= pin_hole_diameter < 0.5355):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('12MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.5355 <= pin_hole_diameter < 0.5750):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('13MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.5750 <= pin_hole_diameter < 0.6142):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('14MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.6142 <= pin_hole_diameter < 0.6536):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('15MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.6536 <= pin_hole_diameter < 0.6930):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('16MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.6930 <= pin_hole_diameter < 0.7323):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('17MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.7323 <= pin_hole_diameter < 0.7717):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('18MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.7717 <= pin_hole_diameter < 0.8111):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('19MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.8111 <= pin_hole_diameter < 0.8504):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('20MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.8504 <= pin_hole_diameter < 0.8898):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('21MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.8898 <= pin_hole_diameter < 0.9292):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('22MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9292 <= pin_hole_diameter < 0.9686):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('23MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9686 <= pin_hole_diameter < 1.0079):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('24MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (1.0079 <= pin_hole_diameter < 1.0473):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('1_INCH')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (1.0473 <= pin_hole_diameter < 1.0942):
                            ROUGH_BORE_TOOL_LIST_INDEX = rough_bore_tool_list.index('26MM')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (ROUGH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            ROUGH_BORE_TOOL_NUMBER = horizontal_tool_list_file['ROUGH_BORE_TOOL_LIST'].loc[
                                ROUGH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (pin_hole_diameter == 0):
                            # MAYBE WE CAN PUT MESSAGE ON TOP OF PROGRAM AND ON THE DIALOG OF APP IF ANYTHING UNEXPECTED HAPPEN
                            print("PIN HOLE DIAMETER IS MISSING, SEE PROGRAMMING")
                            # JUST TO NOT GIVE AN ERROR WHEN USE IT ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE, BECAUSE THE VERIABLE CAN NOT BE NOTHING, SHOULD HAVE A VALUE
                            ROUGH_BORE_TOOL_NUMBER = 0
                        else:
                            # NEED TO BACK AND WORK ON IT
                            print("MAYBE WE CAN USE SWAP TOOL T60 OR JUST PUT MESSAGE SEE PROGRAMMING")

                        # LOGIC FOR ADD THE FINISH BORE TOOL TO THE TOOL LIST IN THE PROGRAM AND TO THE TOOLS SECTION
                        # NEED TO CHECK IF WE GONNA USE MAPEL TOOL OR BORING BAR
                        if (0.4719 <= pin_hole_diameter <= 0.4729):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.4724)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        # NEED TO CHECK IF WE GONNA USE MAPEL TOOL OR BORING BAR
                        elif (0.4895 <= pin_hole_diameter <= 0.4905):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.49)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.5115 <= pin_hole_diameter <= 0.5125):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.512)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.5505 <= pin_hole_diameter <= 0.5515):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.551)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.5905 <= pin_hole_diameter <= 0.5915):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.591)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.6295 <= pin_hole_diameter <= 0.6305):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.63)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.6685 <= pin_hole_diameter <= 0.6695):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.669)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.6715 <= pin_hole_diameter <= 0.6725):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.672)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.7085 <= pin_hole_diameter <= 0.7095):  # IT SEEMS YOU CAN USE IF YPU HAVE 0.708 AS WELL
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.709)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.7475 <= pin_hole_diameter <= 0.7485):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.748)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.7865 <= pin_hole_diameter <= 0.7875):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.787)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.7905 <= pin_hole_diameter <= 0.7915):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.791)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.7915 <= pin_hole_diameter <= 0.7925):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.792)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                            # NEED TO CHECK IF WE CAN USE ROUGH TOOL OF 20MM INSTEAD 19MM AS OLD PROGRAMS DONE
                        elif (0.8119 <= pin_hole_diameter <= 0.8129):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.8124)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.8265 <= pin_hole_diameter <= 0.8275):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.827)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.8655 <= pin_hole_diameter <= 0.8665):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.866)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9045 <= pin_hole_diameter <= 0.9055):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.905)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9115 <= pin_hole_diameter <= 0.9125):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.912)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9265 <= pin_hole_diameter <= 0.9275):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.927)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                            # ******************MAYBE NEED TO ADD 0.943 AS MAPLE USE T45**********************

                        elif (0.9445 <= pin_hole_diameter <= 0.9455):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.945)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9745 <= pin_hole_diameter <= 0.9755):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.975)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9795 <= pin_hole_diameter <= 0.9805):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.98)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9835 <= pin_hole_diameter <= 0.9845):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.984)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9895 <= pin_hole_diameter <= 0.9905):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(0.99)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (0.9995 <= pin_hole_diameter <= 1.0005):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index(1.00)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                            # MAYBE NEED TO ADD ANOTHER LOGIC TO USE 1.094 MAPEL TOOL IF NEED IT

                        elif (1.0935 <= pin_hole_diameter <= 1.0945):
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index('1.094_BORING')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                        elif (pin_hole_diameter == 0):
                            # MAYBE WE CAN PUT MESSAGE ON TOP OF PROGRAM AND ON THE DIALOG OF APP IF ANYTHING UNEXPECTED HAPPEN
                            print("PIN HOLE DIAMETER IS MISSING, SEE PROGRAMMING")
                            # JUST TO NOT GIVE AN ERROR WHEN USE IT ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE, BECAUSE THE VERIABLE CAN NOT BE NOTHING, SHOULD HAVE A VALUE
                            FINISH_BORE_TOOL_NUMBER = 0
                        else:
                            FINISH_BORE_TOOL_LIST_INDEX = finish_bore_tool_list.index('BORING_BAR_TOOL')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                pin_hole_diameter) + ' BORING BAR)')
                            # DEFINE VARIABLE (FINISH_BORE_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            FINISH_BORE_TOOL_NUMBER = horizontal_tool_list_file['FINISH_BORE_TOOL_LIST'].loc[
                                FINISH_BORE_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                            # WE NEED TO BACK TO ADD IF WE NEED SWAP TOOL TO USE T45 MAPEL

                            # NEEDS TO ADD LOGIC TO HANDLE UNEXPECTED "pin_hole_diameter" (NEEDS TO KNOW WHAT BIGGEST AND SMALLEST "pin_hole_diameter" CAN RUN ON THE MACHINE)

                            # WE NEED TO BACK TO WORK IF NEED TO MAKE LOGIC OF (T03 IS A .927 MAPAL REAMER - WILL HONE TO .928)
                            #               ************(MAYBE WE CAN MAKE THE RULES FOR ALL MAPLE TOOLS BETWEEN -0.001 AND 0.001) EXAMPLE WD-13731**************

                        # NEED LOGIC TO ADD BOTH LOCKRING AND CFREN TOOL DESCRIPTION TO THE TOOL LIST IN THE PROGRAM IF THEY ARE NOT THE SAME TOOL EX WD-13100
                        # NEED TO ADJUST HORIZONTAL TEMPLATE ALSO

                        # LOGIC FOR ADD THE LOCKRING/C_FREN TOOL TO THE TOOL LIST IN THE PROGRAM
                        if (lock_ring_cutter_width == 0.0 and cfren_cutter_width == 0.0):
                            # IT IS MEAN JOB DOESN'T HAVE LOCKRING NOR CFREN
                            pass
                            # JUST TO NOT GIVE AN ERROR WHEN USE IT ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE, BECAUSE THE VERIABLE CAN NOT BE NOTHING, SHOULD HAVE A VALUE
                            LOCK_RING_TOOL_NUMBER = 0
                        # WE SEPARATE LOGIC OF LOCKRING AND CFREN FOR THE NEXT elif STATEMENTS BECAUSE WE NEED TO USE TOOL CUTTER WIDTH IN THE DESCRIPTION
                        elif (lock_ring_cutter_width == 0.039):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.039)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                lock_ring_cutter_width) + ' SQ X .465 DIA. LOCK RING)')
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (cfren_cutter_width == 0.039):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.039)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                cfren_cutter_width) + ' SQ X .465 DIA. LOCK RING)')
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                            # NEED TO CHECK WITCH TOOL WITH 0.042 WIDTH WILL USE WHILE WE HAVE THREE TOOLS WITH DIFFERENT DIAMETER (OR ADD LOGIC TO CHECK WICTH TOOL WILL USE),
                            # NEED TO ADD ANOTHER TOOL TO EXCEL SHEET AS WELL
                        elif (lock_ring_cutter_width == 0.042):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.042)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                lock_ring_cutter_width) + ' SQ X .465 PH HORN LOCK RING)')
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (cfren_cutter_width == 0.042):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.042)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                cfren_cutter_width) + ' SQ X .465 PH HORN LOCK RING)')
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))


                        # BACK TO USE LOGIC OF LOCKRING AND CFREN IN THE ONE if STATEMENT BECAUSE WH HAVE THE TOOLS FOR THE NEXT WIDTH

                        # NEED TO CHECK WITCH TOOL WITH 0.044 WIDTH WILL USE WHILE WE HAVE THREE TOOLS WITH DIFFERENT DIAMETER (OR ADD LOGIC TO CHECK WICTH TOOL WILL USE),
                        # FOR NOW WE WILL USE (T25 IS A .044 RAD X .465 PH HORN LOCK RING)
                        elif (lock_ring_cutter_width == 0.044 or cfren_cutter_width == 0.044):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(
                                '0.044(DIA 0.465)')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (lock_ring_cutter_width == 0.047 or cfren_cutter_width == 0.047):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.047)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                        # WE SEPARATE LOGIC OF LOCKRING AND CFREN FOR THE NEXT elif STATEMENTS BECAUSE WE NEED TO USE TOOL CUTTER WIDTH IN THE DESCRIPTION

                        elif (lock_ring_cutter_width == 0.048):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.048)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                lock_ring_cutter_width) + ' RAD X .575 DIA. LOCK RING)')
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (cfren_cutter_width == 0.048):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.048)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'] + format(
                                cfren_cutter_width) + ' RAD X .575 DIA. LOCK RING)')
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                        # BACK TO USE LOGIC OF LOCKRING AND CFREN IN THE ONE if STATEMENT BECAUSE WH HAVE THE TOOLS FOR THE NEXT WIDTH

                        # NEED TO CHECK WITCH TOOL WITH 0.053 WIDTH WILL USE WHILE WE HAVE THREE TOOLS WITH DIFFERENT DIAMETER (OR ADD LOGIC TO CHECK WICTH TOOL WILL USE),
                        # FOR NOW WE WILL USE (T09 IS A .053 RAD X .618 PH HORN LOCK RING)
                        elif (lock_ring_cutter_width == 0.053 or cfren_cutter_width == 0.053):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(
                                '0.053(DIA 0.618)')
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (lock_ring_cutter_width == 0.059 or cfren_cutter_width == 0.059):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.059)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (lock_ring_cutter_width == 0.063 or cfren_cutter_width == 0.063):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.063)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (lock_ring_cutter_width == 0.065 or cfren_cutter_width == 0.065):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.065)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (lock_ring_cutter_width == 0.067 or cfren_cutter_width == 0.067):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.067)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (lock_ring_cutter_width == 0.076 or cfren_cutter_width == 0.076):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.076)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (lock_ring_cutter_width == 0.077 or cfren_cutter_width == 0.077):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.077)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (lock_ring_cutter_width == 0.088 or cfren_cutter_width == 0.088):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.088)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))
                        elif (lock_ring_cutter_width == 0.11 or cfren_cutter_width == 0.11):
                            LOCK_RING_AND_CFREN_TOOL_LIST_INDEX = lock_ring_and_cfren_tool_list.index(0.11)
                            HorizontalProgramLines.insert(index, horizontal_tool_list_file[
                                'LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (LOCK_RING_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_NUMBER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            # DEFINE VARIABLE (LOCK_RING_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "LOCKRING_CUT_RADIUS (VC163 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            LOCK_RING_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            # DEFINE VARIABLE (CFren_TOOL_DIAMETER) TO STORE TOOL DIAMETER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "CFREN_CUT_RADIUS (VC166 VARIABLE)" SECTION OF HORIZONTAL TEMPLATE
                            CFren_TOOL_DIAMETER = horizontal_tool_list_file['LOCK_RING_AND_CFREN_TOOL_LIST'].loc[
                                LOCK_RING_AND_CFREN_TOOL_LIST_INDEX, 'TOOL_DIAMETER']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                        # NEED TO BACK AND WORK HERE
                        # NEED TO CHECK IF WE CAN USE T43 WITH ANY TOOL WIDTH COMES WITH JOB(LIKE WHAT WE DO FOR BORING BAR),
                        # OR NEED TO PUT MESSAGE ON TOP OF PROGRAM OR ON DIALOG OF APP TO QUESTION THE TOOL AVAILABILITY
                        else:
                            print("LOCK RING TOOL IS NOT ON THE TOOL LIST, SEE PROGRAMMING")
                            # JUST TO NOT GIVE AN ERROR WHEN USE IT ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE, BECAUSE THE VERIABLE CAN NOT BE NOTHING, SHOULD HAVE A VALUE
                            LOCK_RING_TOOL_NUMBER = 0

                        DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS = 0  # JUST FOR NOW
                        # NEED TO DOUBLE CHECK IF PIN_HOLE_DIAMETER SHOULD BE 0.900 OR SOMETHING ELSE FROM HORIZ TEMPLATE,  MAYBE WE CAN MAKE LOGIC IN HORIZ TEMPLATE AS RANGE INSTEAD IT IS EXACT VALUE
                        # NEED TO CHECK IF WE NEED TO ADJUST THE CONDITION OF (PIN_HOLE_DIAMETER >= 0.901) TO CHECK IF PIN HOLE DIAMETER IS COVERED ON HORIZ TEMPLATE LOGIC
                        # NEED TO DOUBLE CHECK IF LOGIC IN HORIZ TEMPLATE COVER ALL FINISH BORE SIZES (LIKE: 0.9839 NOT COVERED )
                        if (double_oil_hole_slot_ID_spacing != 0 and pin_hole_diameter >= 0.901):
                            DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS = 1
                            MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index(
                                'DOUBLE OIL HOLES SLOTS(DOHS) 0.750 PH')
                            HorizontalProgramLines.insert(index,
                                                          horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc
                                                          [
                                                              MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            # DEFINE VARIABLE (DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER) TO STORE TOOL NUMBER FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "TOOLS NUMBER" SECTION OF HORIZONTAL TEMPLATE
                            DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER = horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc[
                                MISCELLANEOUS_TOOL_LIST_INDEX, 'TOOL_NUMBER(FOR_MACHINES_27/28/32) (T00)']
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                        NOTCH_AVAILABILITY_STATUS = 0
                        if (notch_angle_first_location != 0):
                            NOTCH_AVAILABILITY_STATUS = 1
                            MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index('NOTCH TOOL 5/32 DIA')
                            HorizontalProgramLines.insert(index,
                                                          horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc
                                                          [
                                                              MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            index += 1
                            ##print('\n'.join(HorizontalProgramLines))

                        HORIZONTAL_SLOTS_AVAILABILITY_STATUS = 0
                        HORIZONTAL_SLOTS_STRAIGHT_THROUGH_AVAILABILITY_STATUS = 0
                        if (horizontal_slots_arc_diameter != 0):
                            MISCELLANEOUS_TOOL_LIST_INDEX = miscellaneous_tool_list.index(
                                'HORIZONTAL SLOTS TOOL 0.375 DIA')
                            HorizontalProgramLines.insert(index,
                                                          horizontal_tool_list_file['MISCELLANEOUS_TOOL_LIST'].loc
                                                          [
                                                              MISCELLANEOUS_TOOL_LIST_INDEX, 'DESCRIPTION(FOR_MACHINES_27/28/32)'])
                            index += 1
                            # MAYBE WE NEED TO FIX THIS CONDITION LATER
                            ## WE STILL NEED TO BACK HERE AND WORK ON CONDITION, MAYBE NEED TO ADD SOMETHING FROM QANTEL DATABASE
                            if (horizontal_slots_arc_diameter != 0.375):
                                HORIZONTAL_SLOTS_AVAILABILITY_STATUS = 1
                            elif (horizontal_slots_arc_diameter == 0.375):
                                HORIZONTAL_SLOTS_STRAIGHT_THROUGH_AVAILABILITY_STATUS = 1
                            ##print('\n'.join(HorizontalProgramLines))

                            # MAYBE WE DON'T WANT THIS SECTION IF WE MAKE HORIZ SLOT STANDER WITH BORE SIZE
                        # +++++-----------------------------******************************************************************************--------------------------------------+++++#
                        # +++++-------LOGIC TO SET HORIZONTAL SLOT NUMBERS ACCORDING TO PIN_HOLE_DIAMETER TO USE THEM LATER IN VARIABLES(VC176,VC177, AND VC178) SECTIONS-------+++++#
                        # +++++-----------------------------******************************************************************************--------------------------------------+++++#
                        # MAYBE THE WAY WE PROGRAM HORIZONTAL SLOTS WILL CHANGE SOON , BUT FOR NOW WE WILL USE THE WAY WE HAVE
                        ### SOME OF THESE NUMBERS ARE MISSING, NEED TO BACK AND FIGURE OUT THESE NUMBERS OR MAKE MESSAGE LIKE: CAN'T FIND THEM OR SOMETHING SIMILER,
                        ### ALSO NEED A LOGIC TO CHECK WHEN WE HAVE TO MAKE MANUAL HORIZ SLOT PROGRAM BY MASTERCAM
                        if (0.4719 <= pin_hole_diameter <= 0.4729):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.4724)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.4895 <= pin_hole_diameter <= 0.4905):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.49)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.5115 <= pin_hole_diameter <= 0.5125):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.512)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.5505 <= pin_hole_diameter <= 0.5515):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.551)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.5905 <= pin_hole_diameter <= 0.5915):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.591)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.6295 <= pin_hole_diameter <= 0.6305):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.63)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.6685 <= pin_hole_diameter <= 0.6695):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.669)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.6715 <= pin_hole_diameter <= 0.6725):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.672)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.7085 <= pin_hole_diameter <= 0.7095):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.709)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.7278 <= pin_hole_diameter <= 0.7288):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.7283)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.7475 <= pin_hole_diameter <= 0.7485):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.748)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.7865 <= pin_hole_diameter <= 0.7875):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.787)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.7905 <= pin_hole_diameter <= 0.7915):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.791)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.7915 <= pin_hole_diameter <= 0.7925):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.792)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.7995 <= pin_hole_diameter <= 0.8005):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.800)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.8119 <= pin_hole_diameter <= 0.8129):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.8124)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.8265 <= pin_hole_diameter <= 0.8275):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.827)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.8265 <= pin_hole_diameter <= 0.8275):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.827)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.8655 <= pin_hole_diameter <= 0.8665):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.866)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.8745 <= pin_hole_diameter <= 0.8755):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.875)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9005 <= pin_hole_diameter <= 0.9015):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.901)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9045 <= pin_hole_diameter <= 0.9055):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.905)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9115 <= pin_hole_diameter <= 0.9125):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.912)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9265 <= pin_hole_diameter <= 0.9275):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.927)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9395 <= pin_hole_diameter <= 0.9405):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.94)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9445 <= pin_hole_diameter <= 0.9455):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.945)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9745 <= pin_hole_diameter <= 0.9755):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.975)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9795 <= pin_hole_diameter <= 0.9805):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.98)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9820 <= pin_hole_diameter <= 0.9830):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.9825)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9835 <= pin_hole_diameter <= 0.9845):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.984)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9895 <= pin_hole_diameter <= 0.9905):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(0.99)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (0.9995 <= pin_hole_diameter <= 1.0005):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(1.00)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (1.0305 <= pin_hole_diameter <= 1.0315):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(1.031)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        elif (1.0935 <= pin_hole_diameter <= 1.0945):
                            HORIZONTAL_SLOT_NUMBERS_INDEX = horizontal_slot_numbers.index(1.094)
                            # DEFINE VARIABLE (i_START_HORIZONTAL_SLOT) TO SET i_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC176" SECTION OF HORIZONTAL TEMPLATE
                            i_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'i_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (j_START_HORIZONTAL_SLOT) TO SET j_START_HSLOT NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC177" SECTION OF HORIZONTAL TEMPLATE
                            j_START_HORIZONTAL_SLOT = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'j_START_HORIZONTAL_SLOT']
                            # DEFINE VARIABLE (HORIZONTAL_SLOT_RADIUS) TO SET HSLOT_RADIUS NUMBER THAT'S COME FROM EXCEL FILE OF TOOL LIST TO USE IT LATER ON "VARIABLE VC178" SECTION OF HORIZONTAL TEMPLATE
                            HORIZONTAL_SLOT_RADIUS = horizontal_tool_list_file['HORIZONTAL_SLOT_NUMBERS'].loc[
                                HORIZONTAL_SLOT_NUMBERS_INDEX, 'HORIZONTAL_SLOT_RADIUS']

                        else:
                            ### STILL NEED TO BACK AND WORK HERE
                            # IN CASE THE NUMBERS NOT FOUND FOR PIN HOLE DIAMETER SIZE , MAKE A MESSAGE TO SAY THAT , MAYBE NEED TO CREATE DIALOG
                            print(
                                "LOOKS LIKE HORIZONTAL SLOT NUMBERS FOR THIS PIN HOLE DIAMETER NOT EXIST, MAYBE YOU CAN FIGURE OUT THEM OR NEED TO MAKE MANUAL HORIZ SLOTS ")

                        # END OF ADDING THE TOOL LIST FOR THE TEMPLATE
                        # NEED TO PUT THIS LINE IN END OF THE while LOOP TO AVOID THE for LOOP START AGAIN
                        HORIZONTAL_PROGRAM_TOOL_LIST_END = 1

                # WE CAN USE CODE BELOW TO ADJUST EACH VARIABLE DEPENDS ON JOB WE NEED TO MAKE PROGRAM FOR
                # MAKE for LOOP TO GO AND ITERATES THROUGH THE HORIZONTAL TEMPLATE (HorizontalProgramLines List)

                for line in HorizontalProgramLines:  # TO READ EACH LINE IN THE TEMPLATE

                    # BEGINNING OF FEATURE LIST STATUS    (*******FEATURE LIST********)

                    ### LEDGE CUT VARIABLE (VC118)
                    substr = "VC118"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC118_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC118_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC118_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        # (LEDGE_CUT_AVAILABILITY_STATUS) COMES FROM LOGIC OF LEDGE CUT STATUS OF TOOL LIST ABOVE
                        HorizontalProgramLines[VC118_VARIABLE_INDEX] = (
                                'VC118=' + format(ledge_cut_availability_status) + '  (LedgeCut)')
                        print("LEDGE_CUT_AVAILABILITY_STATUS in program: ", ledge_cut_availability_status)

                    ### PILOT VARIABLE (VC119)
                    substr = "VC119"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC119_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC119_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC119_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        # (PILOT_AVAILABILITY_STATUS) COMES FROM LOGIC OF PILOT STATUS OF TOOL LIST ABOVE
                        HorizontalProgramLines[VC119_VARIABLE_INDEX] = (
                                'VC119=' + format(PILOT_AVAILABILITY_STATUS) + '  (Pilot)')

                    ### LOCK RING VARIABLE (VC121)
                    # WE DEFINE VARIABLE TO STORE LOCK RING AVAILABILITY STATUS
                    # WE DEFINE IT HERE NOT ON TOOL LIST LOGIC ABOVE ,BECAUSE LOGIC ABOVE CHECK LOCK RING AND CFREN IN THE SAME STATMENT (MAYBE WE WILL CHANGE IT LATER, WILL SEE)
                    LOCK_RING_AVAILABILITY_STATUS = 0
                    substr = "VC121"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # MAYBE NEED TO ADJUST THE CONDITION BELOW TO NOT CHANGE THE STATUS IF WE DON'T HAVE THE TOOL WIDTH IN THE TOOL LIST
                    if (index == 0 and lock_ring_cutter_width != 0):
                        LOCK_RING_AVAILABILITY_STATUS = 1
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC121_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC121_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC121_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC121_VARIABLE_INDEX] = (
                                'VC121=' + format(LOCK_RING_AVAILABILITY_STATUS) + '  (LockRing)')

                    ### CFREN VARIABLE (VC122)
                    # WE DEFINE VARIABLE TO STORE CFREN AVAILABILITY STATUS
                    # WE DEFINE IT HERE NOT ON TOOL LIST LOGIC ABOVE ,BECAUSE LOGIC ABOVE CHECK LOCK RING AND CFREN IN THE SAME STATMENT (MAYBE WE WILL CHANGE IT LATER, WILL SEE)
                    CFren_AVAILABILITY_STATUS = 0
                    substr = "VC122"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0 and cfren_cutter_width != 0):
                        CFren_AVAILABILITY_STATUS = 1
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC122_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC122_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC122_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC122_VARIABLE_INDEX] = (
                                'VC122=' + format(CFren_AVAILABILITY_STATUS) + '  (C-Fren)')

                    ### DOUBLE_OIL_HOLE_SLOT VARIABLE (VC123)
                    substr = "VC123"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC123_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC123_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC123_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        # (DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS) COMES FROM LOGIC OF DOUBLE_OIL_HOLE_SLOT OF TOOL LIST ABOVE
                        HorizontalProgramLines[VC123_VARIABLE_INDEX] = (
                                'VC123=' + format(DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS) + '  (DOHS)')

                    ### NOTCH VARIABLE (VC124)
                    substr = "VC124"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC124_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC124_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC124_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        # (NOTCH_AVAILABILITY_STATUS) COMES FROM LOGIC OF NOTCH STATUS OF TOOL LIST ABOVE
                        HorizontalProgramLines[VC124_VARIABLE_INDEX] = (
                                'VC124=' + format(NOTCH_AVAILABILITY_STATUS) + '  (CirclipNotch)')

                    ### HORIZONTAL SLOTS VARIABLE (VC125)
                    substr = "VC125"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC125_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC125_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC125_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        # (HORIZONTAL_SLOTS_AVAILABILITY_STATUS) COMES FROM LOGIC OF HORIZONTAL_SLOTS STATUS OF TOOL LIST ABOVE
                        HorizontalProgramLines[VC125_VARIABLE_INDEX] = (
                                'VC125=' + format(HORIZONTAL_SLOTS_AVAILABILITY_STATUS) + '  (H-Slots)')

                    ### LEDGE COUNTERBORE VARIABLE (VC126)
                    # WE DEFINE VARIABLE TO STORE LEDGE COUNTERBORE AVAILABILITY STATUS
                    LEDGE_COUNTERBORE_AVAILABILITY_STATUS = 0
                    substr = "VC126"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0 and ledge_counterbore_diameter != 0):
                        LEDGE_COUNTERBORE_AVAILABILITY_STATUS = 1
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC126_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC126_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC126_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC126_VARIABLE_INDEX] = (
                                'VC126=' + format(LEDGE_COUNTERBORE_AVAILABILITY_STATUS) + '  (LedgeCounterbore)')

                    ### DOUBLE NOTCH VARIABLE (VC127)
                    # WE DEFINE VARIABLE TO STORE DOUBLE NOTCH AVAILABILITY STATUS
                    DOUBLE_NOTCH_AVAILABILITY_STATUS = 0
                    substr = "VC127"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0 and notch_angle_second_location != 0):
                        DOUBLE_NOTCH_AVAILABILITY_STATUS = 1
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC127_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC127_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC127_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC127_VARIABLE_INDEX] = (
                                'VC127=' + format(DOUBLE_NOTCH_AVAILABILITY_STATUS) + '   (Double Notch)')

                    ### HORIZONTAL_SLOTS_STRAIGHT_THROUGH(0.375 DIA) VARIABLE (VC129)
                    substr = "VC129"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR

                    ## WE STILL NEED TO BACK HERE AND WORK ON CONDITION, MAYBE NEED TO ADD SOMETHING FROM QANTEL DATABASE
                    if (index == 0 and horizontal_slots_arc_diameter == 0.375):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC129_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC129_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC129_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC129_VARIABLE_INDEX] = ('VC129=' + format(
                            HORIZONTAL_SLOTS_STRAIGHT_THROUGH_AVAILABILITY_STATUS) + '   (.375SlotsStraightThough)')

                    # END OF FEATURE LIST STATUS (*******FEATURE LIST********)

                    # BIGINNIG OF TOOLS NUMBER (*****TOOLS*****)

                    ### ROUGH BORE TOOL NUUMBER VARIABLE (VC130)
                    substr = "VC130"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC130_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC130_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC130_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        # WE JUST USE if STATEMENT BELOW TO MAKE THE TOOL NUMBER ALWAYS PRINTED WITH TWO DIGITS EVEN IT WAS ONE DIGIT(LIKE: 1,3,5,7)
                        # **IT IS NOT NECESSARY JUST TO KEEP TOOLS ORGANIZED AND HAVE 2 DIGITS**
                        if (ROUGH_BORE_TOOL_NUMBER < 10):
                            HorizontalProgramLines[VC130_VARIABLE_INDEX] = (
                                    'VC130=0' + format(ROUGH_BORE_TOOL_NUMBER) + '  (RoughBoreToolNo)')
                        elif (ROUGH_BORE_TOOL_NUMBER >= 10):
                            HorizontalProgramLines[VC130_VARIABLE_INDEX] = (
                                    'VC130=' + format(ROUGH_BORE_TOOL_NUMBER) + '  (RoughBoreToolNo)')

                    ### FINISH BORE TOOL NUUMBER VARIABLE (VC132)
                    substr = "VC132"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC132_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC132_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC132_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        # WE JUST USE if STATEMENT BELOW TO MAKE THE TOOL NUMBER ALWAYS PRINTED WITH TWO DIGITS EVEN IT WAS ONE DIGIT(LIKE: 1,3,5,7)
                        # **IT IS NOT NECESSARY JUST TO KEEP TOOLS ORGANIZED AND HAVE 2 DIGITS**
                        if (FINISH_BORE_TOOL_NUMBER < 10):
                            HorizontalProgramLines[VC132_VARIABLE_INDEX] = (
                                    'VC132=0' + format(FINISH_BORE_TOOL_NUMBER) + '  (FinishBoreToolNo)')
                        elif (FINISH_BORE_TOOL_NUMBER >= 10):
                            HorizontalProgramLines[VC132_VARIABLE_INDEX] = (
                                    'VC132=' + format(FINISH_BORE_TOOL_NUMBER) + '  (FinishBoreToolNo)')

                    ### LOCK RING TOOL NUUMBER VARIABLE (VC134)
                    substr = "VC134"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC134_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC134_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC134_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        # WE JUST USE if STATEMENT BELOW TO MAKE THE TOOL NUMBER ALWAYS PRINTED WITH TWO DIGITS EVEN IT WAS ONE DIGIT(LIKE: 1,3,5,7)
                        # **IT IS NOT NECESSARY JUST TO KEEP TOOLS ORGANIZED AND HAVE 2 DIGITS**
                        if (LOCK_RING_TOOL_NUMBER < 10):
                            HorizontalProgramLines[VC134_VARIABLE_INDEX] = (
                                    'VC134=0' + format(LOCK_RING_TOOL_NUMBER) + '  (LockRingToolNo)')
                        elif (LOCK_RING_TOOL_NUMBER >= 10):
                            HorizontalProgramLines[VC134_VARIABLE_INDEX] = (
                                    'VC134=' + format(LOCK_RING_TOOL_NUMBER) + '  (LockRingToolNo)')

                    ### DOUBLE_OIL_HOLE_SLOT TOOL NUUMBER VARIABLE (VC136)
                    substr = "VC136"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0 and DOUBLE_OIL_HOLE_SLOT_AVAILABILITY_STATUS == 1):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC136_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC136_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC136_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        # WE JUST USE if STATEMENT BELOW TO MAKE THE TOOL NUMBER ALWAYS PRINTED WITH TWO DIGITS EVEN IT WAS ONE DIGIT(LIKE: 1,3,5,7)
                        # **IT IS NOT NECESSARY JUST TO KEEP TOOLS ORGANIZED AND HAVE 2 DIGITS**
                        # ** WE KNOW THE TOOL NUMBER USED FOR DOUBLE_OIL_HOLE_SLOT IS (5) WHICH MEAN WE DON'T NEED THE elif STATEMENT, BUT PUT HERE IN CASE THE TOOL NUMBER CHANGE ON FUTURE
                        if (DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER < 10):
                            HorizontalProgramLines[VC136_VARIABLE_INDEX] = (
                                    'VC136=0' + format(DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER) + '  (DOHSToolNo)')
                        elif (DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER >= 10):
                            HorizontalProgramLines[VC136_VARIABLE_INDEX] = (
                                    'VC136=' + format(DOUBLE_OIL_HOLE_SLOT_TOOL_NUMBER) + '  (DOHSToolNo)')

                    # END OF TOOLS NUMBER (*****TOOLS*****)

                    # BEGINNING OF DIMENSIONAL VARIABLES (****DIMENSIONAL VARIABLES****)

                    # +++++-------PIN_HOLE_DIAMETER VARIABLE (VC149)-------+++++#
                    substr = "VC149"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and PIN_HOLE_DIAMETER != 0) : JUST IN CASE (PIN_HOLE_DIAMETER) IS MISSING FOR SOME REASON
                    if (index == 0 and pin_hole_diameter != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC149_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC149_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC149_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC149_VARIABLE_INDEX] = (
                                'VC149=' + format(pin_hole_diameter) + '  (PinHoleDiameter)')

                    # +++++-------FORGE_REF_LENGTH VARIABLE (VC150)-------+++++#
                    substr = "VC150"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and FORGE_REF_LENGTH__IS_F_NUMBER_IN_EMSS != 0) : JUST IN CASE (FORGE_REF_LENGTH__IS_F_NUMBER_IN_EMSS) IS MISSING FOR SOME REASON
                    if (index == 0 and forge_ref_length != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC150_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC150_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC150_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC150_VARIABLE_INDEX] = (
                                'VC150=' + format(forge_ref_length) + '  (ForgeRefLength)')

                    # +++++-------PILOT_BORE_DEPTH VARIABLE (VC151)-------+++++#
                    substr = "VC151"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and PILOT_BORE_DEPTH != 0) : JUST IN CASE (PILOT_BORE_DEPTH) IS MISSING FOR SOME REASON
                    if (index == 0 and pilot_bore_depth != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC151_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC151_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC151_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC151_VARIABLE_INDEX] = (
                                'VC151=' + format(pilot_bore_depth) + '  (PilotBoreDepth)')

                    # +++++-------(ROUGH_BORE_SPEED) VARIABLE (VC152)-------+++++#
                    substr = "VC152"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and FINISH_BORE_TOOL_NUMBER == 28) : CHECK IF WE USE BORING BAR TOOL (TOOL NUMBER IS 28), WE NEED TO SLOW THE SPEED TO 6000, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (8000)
                    if (index == 0 and FINISH_BORE_TOOL_NUMBER == 28):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC152_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC152_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC152_VARIABLE_INDEX])
                        # WE NEED TO SLOW THE SPEED TO 6000 WHEN USING THE BORING BAR TOOL
                        ROUGH_BORE_SPEED = 6000
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC152_VARIABLE_INDEX] = (
                                'VC152=' + format(ROUGH_BORE_SPEED) + '  (RoughBoreSpeed)')

                    # +++++-------(ROUGH_BORE_FEED) VARIABLE (VC153)-------+++++#
                    substr = "VC153"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and FINISH_BORE_TOOL_NUMBER == 28) : CHECK IF WE USE BORING BAR TOOL (TOOL NUMBER IS 28), WE NEED TO SLOW THE FEED TO 60, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (100)
                    if (index == 0 and FINISH_BORE_TOOL_NUMBER == 28):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC153_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC153_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC153_VARIABLE_INDEX])
                        # WE NEED TO SLOW THE FEED TO 60 WHEN USING THE BORING BAR TOOL
                        ROUGH_BORE_FEED = 60
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC153_VARIABLE_INDEX] = (
                                'VC153=' + format(ROUGH_BORE_FEED) + '  (RoughBoreFeed)')

                    # +++++-------(X_DISTANCE_FROM_ORIGIN_TO_PIN_CENTER) VARIABLE (VC154)-------+++++#
                    substr = "VC154"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and FINISH_BORE_TOOL_NUMBER == 28) : CHECK IF WE USE BORING BAR TOOL (TOOL NUMBER IS 28), WE NEED TO SLOW THE FEED TO 60, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (100)
                    if (index == 0 and pilot_bore_depth != 0 and pilot_to_pin != 0 and pin_hole_diameter != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC154_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC154_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC154_VARIABLE_INDEX])
                        # MATH CALCULATION OF X_DISTANCE_FROM_ORIGIN_TO_PIN_CENTER
                        # (abs(PILOT_TO_PIN)): WE USE abs HERE TO MAKE THE PILOT_TO_PIN NUMBER POSITIVE BECAUSE MOST LIKELY IT IS NEGETIVE IN QANTEL DATA BASE
                        # (, '.4f'): TO MAKE THE NUMBER HAVE JUST 4 DIGITS (0.0000)
                        # WD DEFINE (X_DISTANCE_FROM_ORIGIN_TO_PIN_CENTER) AS global VARIABLE AGAIN TO CAN USE LATER ON NOTCH LOGIC
                        global X_distance_from_origin_to_pin_center
                        X_distance_from_origin_to_pin_center = format(
                            pilot_bore_depth - abs(pilot_to_pin) - (pin_hole_diameter / 2), '.4f')
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC154_VARIABLE_INDEX] = (
                                'VC154=-' + format(X_distance_from_origin_to_pin_center) + '  (xPinCenter)')

                    # +++++-------(OFFSET) VARIABLE (VC155)-------+++++#
                    substr = "VC155"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):

                        # **********STILL WORK HER************
                        # just for now
                        # something should happen here
                        # verification_messages_of_creating_old_horizontal_machine_program = ["Offset can't find, Please Enter the value"]
                        # OFFSET_AMOUNT= self.Confirmation_MDTextField.text
                        if (offset_amount == ""):
                            verification_messages_of_creating_old_horizontal_machine_program = [
                                "Offset Amount does NOT found, Please Enter the value"]
                            print(verification_messages_of_creating_old_horizontal_machine_program)
                            old_horizontal_program_confirmation_email_message_list.append(
                                "# It was need confirmation of :" + "\n" + "\n".join(
                                    verification_messages_of_creating_old_horizontal_machine_program))
                            self.title = '[color=0066ff]Confirmation Message[/color]' + '\n' + '\n' + '[b][i][color=ffffff]' + \
                                         verification_messages_of_creating_old_horizontal_machine_program[
                                             0] + '[/color][/i][/b]'
                            # ******************STILL WORK HERE================
                            # MAKE IT MATCH WITH HEIGHT OF (Confirmation_MDTextField) TO NOT HAVE ANY EMPTY SPACE
                            self.Dialog_BoxLayout = BoxLayout(height=30)
                            # # NEED TO BAKE HERE AND MAKE (text) ARGUMENT TAKE VARIABLE TEXT RELATED ON THE SITUATION
                            # Confirmation_MDLabel = MDLabel(text=('\n'.join(verification_messages_of_creating_old_horizontal_machine_program)),
                            #                                theme_text_color='Custom', text_color=(1, 1, 1, 1))
                            # self.Dialog_BoxLayout.add_widget(Confirmation_MDLabel)
                            # global Confirmation_MDTextField
                            self.Confirmation_MDTextField = TextInput(hint_text="Enter Value", multiline=False,
                                                                      input_filter="float",
                                                                      background_color=[60 / 255.0, 60 / 255.0,
                                                                                        60 / 255.0, 1],
                                                                      foreground_color=[1, 1, 1, 1], size_hint_y=None,
                                                                      height=30)  # ,hint_text= "color_mode = 'accent'"   ,mode="fill",fill_color= (1, 1, 1, 1)   line_color_focus
                            self.Dialog_BoxLayout.add_widget(self.Confirmation_MDTextField)

                            need_confirmation_to_create_old_horizontal_machine_program(self, self.title,
                                                                                       self.enter_offset_value,
                                                                                       "custom", self.Dialog_BoxLayout)
                            # OFFSET_AMOUNT = self.Confirmation_MDTextField.text
                            print("offset value before ", offset_amount)
                            return

                        # **************************************************************************************#
                        if (offset_direction == "" and offset_amount != 0):
                            verification_messages_of_creating_old_horizontal_machine_program = [
                                "Offset Direction does NOT found, Please Choose the direction"]
                            print(verification_messages_of_creating_old_horizontal_machine_program)
                            old_horizontal_program_confirmation_email_message_list.append(
                                "# It was need confirmation of :" + "\n" + "\n".join(
                                    verification_messages_of_creating_old_horizontal_machine_program))

                            # (OFFSET_AMOUNT != 0)
                            self.title = '[color=0066ff]Confirmation Message[/color] ' + '\n' + '\n' + '[b][i][color=ffffff]' + \
                                         verification_messages_of_creating_old_horizontal_machine_program[
                                             0] + '[/color][/i][/b]'

                            self.OFFSET_To0_Option = OldHorizontalMachineItem(text="OFFSET To0",
                                                                              on_release=self.set_offset_direction_To0)
                            self.OFFSET_To180_Option = OldHorizontalMachineItem(text="OFFSET To180",
                                                                                on_release=self.set_offset_direction_To180)
                            self.OFFSET_EACH_WAY_Option = OldHorizontalMachineItem(text="OFFSET EACH WAY",
                                                                                   on_release=self.set_offset_direction_each_way)
                            # self.items = [self.OFFSET_To0_Option, self.OFFSET_To180_Option,self.OFFSET_EACH_WAY_Option]
                            self.items = [self.OFFSET_To0_Option, self.OFFSET_To180_Option, self.OFFSET_EACH_WAY_Option]
                            # '[b][i][u][color=ffffff]' + verification_messages_of_creating_old_horizontal_machine_program[0] + '[/color][/u][/i][/b]'

                            # self.Dialog_BoxLayout.add_widget(items)

                            need_confirmation_to_create_old_horizontal_machine_program(self, self.title,
                                                                                       self.choose_offset_direction,
                                                                                       "confirmation", self.items)
                            # OFFSET_AMOUNT = self.Confirmation_MDTextField.text
                            print("offset direction before ", offset_direction)
                            return
                        # # just for now
                        # verification_messages_of_creating_old_horizontal_machine_program = ["Offset can't find, Please Enter the value"]
                        # # OFFSET_AMOUNT= self.Confirmation_MDTextField.text
                        # if (verification_messages_of_creating_old_horizontal_machine_program != 0):
                        #     Need_Confermation_to_Create_Horizontal_Program(self)
                        #     # OFFSET_AMOUNT = self.Confirmation_MDTextField.text
                        #     print("offset section ", OFFSET_AMOUNT)
                        #     return

                        # ADJUST OFFSET ACCORDING THE DIRECTION
                        # global OFFSET_AMOUNT
                        if (offset_amount == 0):
                            DIRECTION = 1
                            # just to make thing work
                        elif (offset_direction == "OFFSET EACH WAY" or offset_direction == "OFFSET To0"):
                            DIRECTION = 1
                        elif (offset_direction == "OFFSET To180"):
                            DIRECTION = -1
                        else:
                            need_confirmation_to_create_old_horizontal_machine_program(self, self.title,
                                                                                       self.enter_offset_value,
                                                                                       "custom", self.Dialog_BoxLayout)
                            return

                        # print("offset section ", OFFSET_AMOUNT)

                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        global VC155_VARIABLE_INDEX
                        VC155_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC155_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC155_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC155_VARIABLE_INDEX] = (
                                'VC155=' + format(offset_amount * DIRECTION) + '  (Offset)')

                    # +++++-------(Z_VALUE_PIN_BORE_TOP) VARIABLE (VC156)-------+++++#
                    substr = "VC156"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC156_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC156_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC156_VARIABLE_INDEX])
                        if (forging_diameter_OD_at_rougher != 0):
                            # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                            HorizontalProgramLines[VC156_VARIABLE_INDEX] = ('VC156=[' + format(
                                forging_diameter_OD_at_rougher) + '/2]' + '  (zPinBoreTop - ? IS Forging Diameter)')
                        elif (forging_diameter != 0):
                            # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                            HorizontalProgramLines[VC156_VARIABLE_INDEX] = ('VC156=[' + format(
                                forging_diameter) + '/2]' + '  (zPinBoreTop - ? IS Forging Diameter)')

                    # +++++-------(Z_VALUE_ROUGH_BORE_BOTTOM) VARIABLE (VC157)-------+++++#
                    # NO NEED TO DO ANYTHING HERE, IT WILL JUST USE HOW IT IS IN HORIZONTAL TEMPLATE

                    # +++++-------(Z_VALUE_FINISH_BORE_BOTTOM) VARIABLE (VC158)-------+++++#
                    substr = "VC158"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and FINISH_BORE_TOOL_NUMBER == 28) : CHECK IF WE USE BORING BAR TOOL (TOOL NUMBER IS 28), WE NEED TO CHANGE THE NUMBER TO 0.1, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (1.156)
                    if (index == 0 and FINISH_BORE_TOOL_NUMBER == 28):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC158_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC158_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC158_VARIABLE_INDEX])
                        # WE NEED TO TO CHANGE THE NUMBER TO 0.1 WHEN USING THE BORING BAR TOOL
                        VALUE_USED_IN_Z_VALUE_FINISH_BORE_BOTTOM = 0.1
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC158_VARIABLE_INDEX] = ('VC158=-[VC156+' + format(
                            VALUE_USED_IN_Z_VALUE_FINISH_BORE_BOTTOM) + ']' + '  (zFinishBoreBottom)')

                    # +++++-------(FORGING_OUTSIDE_BOSS_SPACING__IS_U_NUMBER_IN_EMSS ) VARIABLE (VC159)-------+++++#
                    substr = "VC159"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    if (index == 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC159_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC159_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC159_VARIABLE_INDEX])
                        if (forging_outside_boss_spacing != 0):
                            # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                            HorizontalProgramLines[VC159_VARIABLE_INDEX] = ('VC159=' + format(
                                forging_outside_boss_spacing) + '  (OutsideBossSpacing)')
                        elif (forging_outside_boss_spacing == 0):
                            if (forging_diameter_OD_at_rougher != 0):
                                # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                                HorizontalProgramLines[VC159_VARIABLE_INDEX] = ('VC159=' + format(
                                    forging_diameter_OD_at_rougher) + '  (OutsideBossSpacing)')
                            elif (forging_diameter != 0):
                                # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                                HorizontalProgramLines[VC159_VARIABLE_INDEX] = ('VC159=' + format(
                                    forging_diameter) + '  (OutsideBossSpacing)')

                    # +++++-------(LEDGE_TOOL_DIAMETER) VARIABLE (VC160)-------+++++#
                    substr = "VC160"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and (LEDGE_CUT_AVAILABILITY_STATUS == 1 or LEDGE_COUNTERBORE_DIAMETER != 0)) : TO CHECK IF WE USE THE LEDGE TOOL OR NOT,
                    if (index == 0 and (ledge_cut_availability_status == 1 or ledge_counterbore_diameter != 0)):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC160_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC160_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC160_VARIABLE_INDEX])
                        # WE NEED TO CHANGE DIAMETER OF LEDGE TOOL TO 0.375 (TOOL DIAMETER OF TOOL WE USE WHEN HAVE SMALL BORE) IF PIN HOLE DIAMETER SMALLER THAN THE STD LEDGE TOOL DIAMETER(0.625),
                        # OTHERWISE LEAVE IT AS IT IS IN TEMPLATE (0.625)
                        if (pin_hole_diameter < 0.629):
                            LEDGE_TOOL_DIAMETER = 0.375
                            HorizontalProgramLines[VC160_VARIABLE_INDEX] = (
                                    'VC160=' + format(LEDGE_TOOL_DIAMETER) + '  (LedgeToolDiameter)')

                    # +++++-------(Z_VALUE_OF_TOP_OF_LOCKRING) VARIABLE (VC161)-------+++++#
                    substr = "VC161"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and LOCK_RING_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS LOCK RING, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                    if (index == 0 and lock_ring_cutter_width != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC161_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC161_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC161_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC161_VARIABLE_INDEX] = ('VC161=[' + format(
                            lock_ring_ID_spacing) + '/2]' + '  (zTopLockRing = Replace 0 with Lockring ID Spacing)')

                    # +++++-------(Z_VALUE_OF_BOTTOM_OF_LOCKRING) VARIABLE (VC162)-------+++++#
                    substr = "VC162"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and LOCK_RING_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS LOCK RING, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                    if (index == 0 and lock_ring_cutter_width != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC162_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC162_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC162_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC162_VARIABLE_INDEX] = ('VC162=-[VC161+' + format(
                            lock_ring_cutter_width) + ']' + '  (zBottomLockRing IS zTopLockRing + LR CUTTER WIDTH)')

                    # +++++-------(LOCKRING_CUT_RADIUS) VARIABLE (VC163)-------+++++#
                    substr = "VC163"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and LOCK_RING_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS LOCK RING, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                    if (index == 0 and lock_ring_cutter_width != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC163_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC163_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC163_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC163_VARIABLE_INDEX] = (
                                'VC163=[[' + format(lock_ring_diameter) + '-' + format(
                            LOCK_RING_TOOL_DIAMETER) + ']/2]' + '  (LRCutRadius)')

                    # +++++-------(Z_VALUE_OF_TOP_OF_CFREN) VARIABLE (VC164)-------+++++#
                    substr = "VC164"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and CFren_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS CFREN, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                    if (index == 0 and cfren_cutter_width != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC164_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC164_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC164_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC164_VARIABLE_INDEX] = ('VC164=[' + format(
                            cfren_ID_spacing) + '/2]' + '  (zTopCFREN = Replace 0 with CFren ID Spacing)')

                    # +++++-------(Z_VALUE_OF_BOTTOM_OF_CFREN) VARIABLE (VC165)-------+++++#
                    substr = "VC165"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and CFren_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS CFren, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                    if (index == 0 and cfren_cutter_width != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC165_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC165_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC165_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC165_VARIABLE_INDEX] = ('VC165=-[VC164+' + format(
                            cfren_cutter_width) + ']' + '  (zBottomCFREN IS zTopCFREN + LR CUTTER WIDTH)')

                    # +++++-------(CFREN_CUT_RADIUS) VARIABLE (VC166)-------+++++#
                    substr = "VC166"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and CFren_CUTTER_WIDTH != 0) : TO CHECK IF JOB HAS LOCK RING, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE
                    if (index == 0 and cfren_cutter_width != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC166_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC166_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC166_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC166_VARIABLE_INDEX] = (
                                'VC166=[[' + format(cfren_diameter) + '-' + format(
                            CFren_TOOL_DIAMETER) + ']/2]' + '  (CFCutRadius)')

                    # +++++-------(CFREN_CUT_OFFSET) VARIABLE (VC167)-------+++++#
                    # NO NEED TO DO ANYTHING HERE FOR NOW, IT WILL JUST USE HOW IT IS IN HORIZONTAL TEMPLATE

                    # +++++-------(DOUBLE_OIL_HOLE_SLOT_ID_SPACING) VARIABLE (VC168)-------+++++#
                    substr = "VC168"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and DOUBLE_OIL_HOLE_SLOT_ID_SPACING != 0 and PIN_HOLE_DIAMETER >= 0.901) : CHECK IF JOB HAS DOUBLE_OIL_HOLE_SLOT OR NOT AND PIN HOLE SIZE IS BIGGER THAN 0.901,
                    # OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    if (index == 0 and double_oil_hole_slot_ID_spacing != 0 and pin_hole_diameter >= 0.901):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC168_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC168_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC168_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC168_VARIABLE_INDEX] = (
                                'VC168=' + format(double_oil_hole_slot_ID_spacing) + '  (DOHS_ID_Spacing)')

                    # +++++-------NOTCH CALCULATION TO USE ON VARIABLES (VC172 AND VC173) SECTIONS LATER -------+++++#

                    # NEED TO BACK LATER AND EXPLAIN HOW YOU FIGURED OUT THING BELOW
                    # (and OFFSET_AMOUNT != ""): JUST TO MAKE THE CODE WORK IF CAN'T DETECT THE OFFSET FROM DATA BASE
                    if (notch_angle_first_location != 0 and offset_amount != ""):
                        # ADJUST OFFSET ACCORDING THE DIRECTION
                        if (offset_amount == 0):
                            DIRECTION = 1
                        elif (offset_direction == "OFFSET EACH WAY" or offset_direction == "OFFSET To0"):
                            DIRECTION = 1
                        elif (offset_direction == "OFFSET To180"):
                            DIRECTION = -1
                        else:
                            # just to make the code work
                            DIRECTION = 1
                        #     Need_Confermation_to_Create_Horizontal_Program(self , self.title ,self.EnterOffsetValue , "custom", self.Dialog_BoxLayout)
                        #     return
                        # Need_Confermation_to_Create_Horizontal_Program(self)

                        # print("in notch section ", OFFSET_AMOUNT)

                        # IT FIGURES OUT BY DRAWING TRIANGLE IN SOLID WORK AND USE TRIGONOMETRIC MATH
                        # WE USE (math.radians(NOTCH_ANGLE)) BECAUSE WE NEED PUT ANGLE IN DEGREES
                        ## print('NOTCH_ANGLE =', NOTCH_ANGLE_FIRST_LOCATION)
                        ## print("OFFSET =", OFFSET)
                        Y_VALUE_FOR_NOTCH_MATH = format(
                            (pin_hole_diameter / 2) * (math.sin(math.radians(notch_angle_first_location))), '.4f')

                        global Y_distance_from_origin_to_circlip_notch
                        Y_distance_from_origin_to_circlip_notch = format(
                            float(Y_VALUE_FOR_NOTCH_MATH) + (float(offset_amount) * DIRECTION), '.4f')
                        ## print("Y_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH =", Y_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH)

                        # IT IS USE PYTHAGORAS THEOREM FORMULA (HYPOTENUSE^2=(PERPENDICULAR^2)+(BASE^2))>>>(PERPENDICULAR=sqrt((HYPOTENUSE^2)-(BASE^2))
                        X_VALUE_FOR_NOTCH_MATH = math.sqrt(
                            (math.pow((pin_hole_diameter / 2), 2)) - (math.pow(float(Y_VALUE_FOR_NOTCH_MATH), 2)))

                        if (notch_angle_first_location <= 90):
                            X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH = format(
                                float(X_distance_from_origin_to_pin_center) - float(X_VALUE_FOR_NOTCH_MATH), '.4f')
                        elif (notch_angle_first_location > 90):
                            X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH = format(
                                float(X_distance_from_origin_to_pin_center) + float(X_VALUE_FOR_NOTCH_MATH), '.4f')
                        ## print("X VALUE IS: " + X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH)

                    # +++++-------(X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH) VARIABLE (VC172)-------+++++#
                    substr = "VC172"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and NOTCH_ANGLE_FIRST_LOCATION != 0) : CHECK IF JOB HAS NOTCH, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    if (index == 0 and notch_angle_first_location != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC172_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC172_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC172_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC172_VARIABLE_INDEX] = (
                                'VC172=-' + format(X_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH) + '  (xCirclipNotch)')

                    # +++++-------(Y_DISTANCE_FROM_ORIGIN_TO_CIRCLIP_NOTCH) VARIABLE (VC173)-------+++++#
                    substr = "VC173"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and NOTCH_ANGLE_FIRST_LOCATION != 0) : CHECK IF JOB HAS NOTCH, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    if (index == 0 and notch_angle_first_location != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        global VC173_VARIABLE_INDEX
                        VC173_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC173_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC173_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC173_VARIABLE_INDEX] = (
                                'VC173=' + format(Y_distance_from_origin_to_circlip_notch) + '  (yCirclipNotch)')

                    # +++++-------(HORIZONTAL_SLOTS_ID_SPACING) VARIABLE (VC175)-------+++++#
                    substr = "VC175"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and HORIZONTAL_SLOTS_ID_SPACING != 0) : CHECK IF JOB HAS HORIZ SLOTS, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    # (and HORIZONTAL_SLOTS_ARC_DIAMETER != 0.375) : ***JUST FOR NOW*** CHECK IF JOB HAS HORIZ SLOTS WITH RADIUS NOT EQUAL 0.375,>>>
                    # >>>BECAUSE HORIZ SLOTS WITH 0.375 RADIUS CONSIDER (HORIZONTAL SLOTS STRAIGHT_THROUGH), WHICH NEED TO ADJUST OTHER VARIABLES
                    if (index == 0 and horizontal_slots_ID_spacing != 0 and horizontal_slots_arc_diameter != 0.375):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC175_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC175_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC175_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC175_VARIABLE_INDEX] = (
                                'VC175=' + format(horizontal_slots_ID_spacing) + '  (HSlot_ID_Spacing)')

                    # +++++-------(i_START_HORIZONTAL_SLOT) VARIABLE (VC176)-------+++++#
                    # IF WE USE STANDERD HORIZ SLOT MAYBE WE WILL MAKE IT EQUAL 1
                    substr = "VC176"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and HORIZONTAL_SLOTS_ID_SPACING != 0) : CHECK IF JOB HAS HORIZ SLOTS, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    # (and HORIZONTAL_SLOTS_ARC_DIAMETER != 0.375) : ***JUST FOR NOW*** CHECK IF JOB HAS HORIZ SLOTS WITH RADIUS NOT EQUAL 0.375,>>>
                    # >>>BECAUSE HORIZ SLOTS WITH 0.375 RADIUS CONSIDER (HORIZONTAL SLOTS STRAIGHT_THROUGH), WHICH NEED TO ADJUST OTHER VARIABLES
                    if (index == 0 and horizontal_slots_ID_spacing != 0 and horizontal_slots_arc_diameter != 0.375):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC176_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC176_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC176_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC176_VARIABLE_INDEX] = (
                                'VC176=' + format(i_START_HORIZONTAL_SLOT) + '  (iStartHSlot)')

                    # +++++-------(j_START_HORIZONTAL_SLOT) VARIABLE (VC177)-------+++++#
                    # IF WE USE STANDERD HORIZ SLOT MAYBE WE WILL MAKE IT EQUAL 1
                    substr = "VC177"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and HORIZONTAL_SLOTS_ID_SPACING != 0) : CHECK IF JOB HAS HORIZ SLOTS, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    # (and HORIZONTAL_SLOTS_ARC_DIAMETER != 0.375) : ***JUST FOR NOW*** CHECK IF JOB HAS HORIZ SLOTS WITH RADIUS NOT EQUAL 0.375,>>>
                    # >>>BECAUSE HORIZ SLOTS WITH 0.375 RADIUS CONSIDER (HORIZONTAL SLOTS STRAIGHT_THROUGH), WHICH NEED TO ADJUST OTHER VARIABLES
                    if (index == 0 and horizontal_slots_ID_spacing != 0 and horizontal_slots_arc_diameter != 0.375):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC177_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC177_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC177_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC177_VARIABLE_INDEX] = (
                                'VC177=' + format(j_START_HORIZONTAL_SLOT) + '  (jStartHSlot)')

                    # +++++-------(HORIZONTAL_SLOT_RADIUS) VARIABLE (VC178)-------+++++#
                    # IF WE USE STANDERD HORIZ SLOT MAYBE WE WILL MAKE IT EQUAL 1
                    substr = "VC178"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and HORIZONTAL_SLOTS_ID_SPACING != 0) : CHECK IF JOB HAS HORIZ SLOTS, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    # (and HORIZONTAL_SLOTS_ARC_DIAMETER != 0.375) : ***JUST FOR NOW*** CHECK IF JOB HAS HORIZ SLOTS WITH RADIUS NOT EQUAL 0.375,>>>
                    # >>>BECAUSE HORIZ SLOTS WITH 0.375 RADIUS CONSIDER (HORIZONTAL SLOTS STRAIGHT_THROUGH), WHICH NEED TO ADJUST OTHER VARIABLES
                    if (index == 0 and horizontal_slots_ID_spacing != 0 and horizontal_slots_arc_diameter != 0.375):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC178_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC178_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC178_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC178_VARIABLE_INDEX] = (
                                'VC178=' + format(HORIZONTAL_SLOT_RADIUS) + '  (HSlotRadius)')

                    # +++++-------(LEDGE_COUNTERBORE_DIAMETER) VARIABLE (VC179)-------+++++#
                    substr = "VC179"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and LEDGE_COUNTERBORE_DIAMETER != 0) : CHECK IF JOB HAS LEDGE_COUNTERBORE, OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    if (index == 0 and ledge_counterbore_diameter != 0):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC179_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC179_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC179_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC179_VARIABLE_INDEX] = (
                                'VC179=' + format(ledge_counterbore_diameter) + '  (CounterboreLedgeDiameter)')

                    # +++++-------MAYBE NEED TO ADD VARIABLE OF DISTANCE OF LEDGE_COUNTERBORE TO LOCKRING TO HORIZONTAL TEMPLATE-------+++++#

                    # +++++-------(X_DISTANCE_375_SLOTS_FROM_CENTER_OF_BORE_TO_CENTER_OF_HORIZONTAL_SLOT) VARIABLE (VC183)-------+++++#

                    # NEED TO FIGURE OUT THE MATH WE WANT TO USE HERE

                    # +++++-------(Y_DISTANCE_375_SLOTS_FROM_CENTER_OF_BORE_TO_CENTER_OF_HORIZONTAL_SLOT) VARIABLE (VC184)-------+++++#

                    # NEED TO FIGURE OUT THE MATH WE WANT TO USE HERE

                    # +++++-------(FORGING_OR_FINISHED_BOSS_WIDTH_FOR_375_SLOTS) VARIABLE (VC185)-------+++++#
                    substr = "VC185"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and HORIZONTAL_SLOTS_ARC_DIAMETER == 0.375) : CHECK IF JOB HAS (HORIZONTAL SLOTS STRAIGHT_THROUGH), OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    if (index == 0 and horizontal_slots_arc_diameter == 0.375):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        VC185_VARIABLE_INDEX = HorizontalProgramLines.index(line)
                        print(VC185_VARIABLE_INDEX)
                        print(HorizontalProgramLines[VC185_VARIABLE_INDEX])
                        # WE USE THE INDEX WE FOUND TO ADJUST THE VALUE DEPENDS ON THE JOB , WE REPLACE THE LIST ELEMENT WITH THE ADJUSTED ONE
                        HorizontalProgramLines[VC185_VARIABLE_INDEX] = ('VC185=' + format(
                            forging_inside_boss_spacing) + '  (ForgingOrFinishedBossWidthFor375Slots)')

                    substr = "CALL OXXXX"  # VARIABLE WE LOOKING FOR
                    # TO FIND INDEX OF substr, IT WILL RETURN -1 WHEN NOT FOUND ANYTHING , AND RETURN substr LOCATION WHEN FIND IT
                    index = line.find(substr)
                    # print(index)
                    # BECAUSE ALL VARIABLE IN TEMPLATE LOCATE IN THE BIGINING OF THE LINE(ie: index VALUE IS 0 FOR (index = line.find(substr)))
                    # WE USE (index == 0) TO TELL THE CODE THAT' FOUND THE substr WE LOOKING FOR
                    # (and HORIZONTAL_SLOTS_ARC_DIAMETER == 0.375) : CHECK IF JOB HAS (HORIZONTAL SLOTS STRAIGHT_THROUGH), OTHERWISE LEAVE IT AS IT IS IN TEMPLATE WITHOUT CHANGE
                    if (index == 0 and forging_number is not None):
                        # IT IS RETURN index OF LINE(ELEMENT) OF THE HorizontalProgramLines LIST WHERE substr FOUND
                        # (ie: IT WILL TELL US LOCATION OF ELEMENT INSIDE THE HORIZONTAL TEMPLATE list THAT'S CONTAIN substr WE FOUND)
                        PROBE_PROGRAM_INDEX = HorizontalProgramLines.index(line)
                        print(PROBE_PROGRAM_INDEX)
                        print(HorizontalProgramLines[PROBE_PROGRAM_INDEX])
                        HorizontalProgramLines[PROBE_PROGRAM_INDEX] = (
                                'CALL ' + format(probe_program_of_old_horizontal_machine))

        # IOError is error of not find the file of template
        except IOError as error:
            # + '[b][u][color=ffffff]Horizontal Template[/color][/u][/b]' +
            fail_messages_of_creating_old_horizontal_machine_program.append(
                "Failed to Find, Load, or Access" + '[b][u][color=ffffff] Horizontal Template [/color][/u][/b]' + "File to Create the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]' + str(
                    error) + '[/color]' + "\n" + "Double Check Network, and File Location.")
            email_messages_of_creating_old_horizontal_machine_program.append(
                "Failed to Find, Load, or Access Horizontal Template File to Create the Program." + "\n" + "Double Check Network, and File Location.")
            failed_to_create_old_horizontal_machine_program(self)
            return

        # endregion <<<<============================[Old Horizontal For Loop Template]============================>>>>

        # region  <<<<========================[Old Horizontal Checking Create Program]==========================>>>>

        # Need_Confermation_to_Create_Horizontal_Program(self)
        # NEED TO BACK AND FIX THIS CONDISION

        if (fail_messages_of_creating_old_horizontal_machine_program == [] and
                verification_messages_of_creating_old_horizontal_machine_program == [] and
                self.ids["JobNumber"].text != ""):
            # just for now
            print("actual")
            print(HorizontalProgramLines)

            create_old_horizontal_machine_program_in_original_folder(self)
            # Horizontal_Program_Has_been_Created_Successfully_Running_Folder(self)
            # JUST FOR TESTING
            # print()
            # print('\n'.join(HorizontalProgramLines))
        elif (verification_messages_of_creating_old_horizontal_machine_program != []):
            print("verification_messages_of_creating_old_horizontal_machine_program STILL HAVE SOMETHING: ",
                  verification_messages_of_creating_old_horizontal_machine_program)

        elif (self.ids["JobNumber"].text == ""):
            fail_messages_of_creating_old_horizontal_machine_program.append("Please Enter Job Number.")
            failed_to_create_old_horizontal_machine_program(self)

        else:
            failed_to_create_old_horizontal_machine_program(self)

        # endregion <<<<========================[Old Horizontal Checking Create Program]=========================>>>>

    # region <<<<=============================[Old Horizontal Machines Sub Functions]=============================>>>>

    def need_to_use_ledge_tool(self, obj):
        print("need_to_use_ledge_tool function CALLED")
        self.Need_Ledge_Tool_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
        self.Does_Not_Need_Ledge_Tool_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.Need_Ledge_Tool_Status = True
        self.Does_Not_Need_Ledge_Tool_Status = False

    def does_not_need_to_use_ledge_tool(self, obj):
        print("does_not_need_to_use_ledge_tool function CALLED")
        self.Does_Not_Need_Ledge_Tool_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
        self.Need_Ledge_Tool_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.Does_Not_Need_Ledge_Tool_Status = True
        self.Need_Ledge_Tool_Status = False

    def decide_ledge_tool_status(self, obj):
        print("decide_ledge_tool_status function CALLED")
        global ledge_cut_availability_status
        if (self.Need_Ledge_Tool_Status == True):
            ledge_cut_availability_status = 1
        elif (self.Does_Not_Need_Ledge_Tool_Status == True):
            ledge_cut_availability_status = 0
        else:
            print("Ladge Tool Status get wrong")
        print("LEDGE_CUT_AVAILABILITY_STATUS in decide function ", ledge_cut_availability_status)
        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()

    def enter_offset_value(self, obj):
        global offset_amount
        offset_amount = float(self.Confirmation_MDTextField.text)
        verification_messages_of_creating_old_horizontal_machine_program = []
        print("offset value after ", offset_amount)
        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()

        # ************** NEED TO FIX COLORS ISSUE

    def set_offset_direction_To0(self, obj):
        print("set_offset_direction_To0 FUNCTION CALLED")
        self.OFFSET_To0_Option.bg_color = (
            20 / 255, 82 / 255, 20 / 255,
            1)  # '[b][i][color=33cc33]' + self.OFFSET_To0_Option.text + '[/color][/i][/b]'
        self.OFFSET_To180_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.OFFSET_EACH_WAY_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.OFFSET_To0_Option_Status = True
        print("To0 ", self.OFFSET_To0_Option_Status)
        # self.OFFSET_To180_Option.text = '[color=ffffff]' + self.OFFSET_To180_Option.text + '[/color]'
        self.OFFSET_To180_Option_Status = False
        print("To180 ", self.OFFSET_To180_Option_Status)
        # self.OFFSET_EACH_WAY_Option.text = '[color=ffffff]' + self.OFFSET_EACH_WAY_Option.text + '[/color]'
        self.OFFSET_EACH_WAY_Option_Status = False
        print("EACH_WAY ", self.OFFSET_EACH_WAY_Option_Status)
        print(self.OFFSET_To0_Option.text)
        print()

    def set_offset_direction_To180(self, obj):
        print("set_offset_direction_To180 FUNCTION CALLED")
        self.OFFSET_To180_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
        self.OFFSET_To0_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.OFFSET_EACH_WAY_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)

        self.OFFSET_To180_Option_Status = True
        print("To180 ", self.OFFSET_To180_Option_Status)
        # self.OFFSET_To0_Option.text = '[color=ffffff]' + self.OFFSET_To0_Option.text + '[/color]'
        self.OFFSET_To0_Option_Status = False
        print("To0 ", self.OFFSET_To0_Option_Status)
        # self.OFFSET_EACH_WAY_Option.text = '[color=ffffff]' + self.OFFSET_EACH_WAY_Option.text + '[/color]'
        self.OFFSET_EACH_WAY_Option_Status = False
        print("EACH_WAY ", self.OFFSET_EACH_WAY_Option_Status)
        print(self.OFFSET_To180_Option.text)
        print()

    def set_offset_direction_each_way(self, obj):
        print("set_offset_direction_each_way FUNCTION CALLED")

        self.OFFSET_EACH_WAY_Option.bg_color = (20 / 255, 82 / 255, 20 / 255, 1)
        self.OFFSET_To0_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)
        self.OFFSET_To180_Option.bg_color = (32 / 255.0, 32 / 255.0, 32 / 255.0, 1)

        self.OFFSET_EACH_WAY_Option_Status = True
        print("EACH_WAY ", self.OFFSET_EACH_WAY_Option_Status)
        # self.OFFSET_To180_Option.text = '[color=ffffff]' + self.OFFSET_To180_Option.text + '[/color]'
        self.OFFSET_To180_Option_Status = False
        print("To180 ", self.OFFSET_To180_Option_Status)
        # self.OFFSET_To0_Option.text = '[color=ffffff]' + self.OFFSET_To0_Option.text + '[/color]'
        self.OFFSET_To0_Option_Status = False
        print("To0 ", self.OFFSET_To0_Option_Status)
        print(self.OFFSET_EACH_WAY_Option.text)

        # **********still work here************

    def choose_offset_direction(self, obj):
        print("choose_offset_direction FUNCTION CALLED")
        global offset_direction
        if (self.OFFSET_To0_Option_Status == True):
            offset_direction = self.OFFSET_To0_Option.text
        elif (self.OFFSET_To180_Option_Status == True):
            offset_direction = self.OFFSET_To180_Option.text
        elif (self.OFFSET_EACH_WAY_Option_Status == True):
            offset_direction = self.OFFSET_EACH_WAY_Option.text
        else:
            print("OFFSET_DIRECTION get wrong")
        print("offset Direction after ", offset_direction)
        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()

    # def Horizontal_Program_Has_been_Created_Successfully_Original_Folder(self):
    #     pass

    # NEED TO PUT (obj) AS PARAMETER TO MAKE FUNCTION WORK
    def replace_existing_old_horizontal_machine_program_in_original_folder(self, obj):
        # NEED TO PUT THIS ONE HERE TO MAKE SURE YOU CLOSE ALL THE DIALOG WINDOW WERE OPEN
        self.Old_Horizontal_Message_Dialog.dismiss()
        print("replace_existing_old_horizontal_machine_program_in_original_folder" + " FUNCTION CALLED")
        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):

            # TO 0
            # JUST MAKE IT global HERE AGAIN TO CAN USE IT IN BELOW FUNCTIONS
            global new_horizontal_program_To0_direction_for_old_machine_in_original_folder
            new_horizontal_program_To0_direction_for_old_machine_in_original_folder = (
                    original_folder_path_of_old_horizontal_machine + "\\" + "P" +
                    new_program_number_for_old_horizontal_machine + "TO0.MIN")
            horizontal_program_lines_To0_direction_for_old_machine_in_original_folder[0] = (
                    '(PART ' + new_program_number_for_old_horizontal_machine + " TO 0" + ' -- ' + todaydate + ' <SYS>' + ')')
            try:
                create_new_horizontal_program_To0_direction_for_old_machine_in_original_folder = open(
                    new_horizontal_program_To0_direction_for_old_machine_in_original_folder, "w")
                create_new_horizontal_program_To0_direction_for_old_machine_in_original_folder.write(
                    '\n'.join(horizontal_program_lines_To0_direction_for_old_machine_in_original_folder))
                create_new_horizontal_program_To0_direction_for_old_machine_in_original_folder.close()
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=ffff00] ORIGINAL [/color][/b]' +
                    "Folder location to Save the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]'
                    + str(error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Original Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

            # TO 180
            # JUST MAKE IT global HERE AGAIN TO CAN USE IT IN BELOW FUNCTIONS
            global new_horizontal_program_To180_direction_for_old_machine_in_original_folder
            new_horizontal_program_To180_direction_for_old_machine_in_original_folder = (
                    original_folder_path_of_old_horizontal_machine + "\\" + "P" +
                    new_program_number_for_old_horizontal_machine + "TO180.MIN")
            horizontal_program_lines_To180_direction_for_old_machine_in_original_folder[0] = (
                    '(PART ' + new_program_number_for_old_horizontal_machine + " TO 180" + ' -- ' + todaydate +
                    ' <SYS>' + ')')
            horizontal_program_lines_To180_direction_for_old_machine_in_original_folder[VC155_VARIABLE_INDEX] = (
                    'VC155=' + format((-1) * offset_amount) + '  (Offset)')
            horizontal_program_lines_To180_direction_for_old_machine_in_original_folder[VC173_VARIABLE_INDEX] = (
                    'VC173=' + format(float(Y_distance_from_origin_to_circlip_notch) - (2 * offset_amount),
                                      '.4f') + '  (yCirclipNotch)')
            try:
                create_new_horizontal_program_To180_direction_for_old_machine_in_original_folder = open(
                    new_horizontal_program_To180_direction_for_old_machine_in_original_folder, "w")
                create_new_horizontal_program_To180_direction_for_old_machine_in_original_folder.write(
                    '\n'.join(horizontal_program_lines_To180_direction_for_old_machine_in_original_folder))
                create_new_horizontal_program_To180_direction_for_old_machine_in_original_folder.close()
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=ffff00] ORIGINAL [/color][/b]' +
                    "Folder location to Save the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]'
                    + str(error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Original Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return
        else:
            new_horizontal_program_for_old_machine_in_original_folder = (
                    original_folder_path_of_old_horizontal_machine + "\\" + "P" +
                    new_program_number_for_old_horizontal_machine + ".MIN")
            try:
                create_new_horizontal_program_for_old_machine_in_original_folder = open(
                    new_horizontal_program_for_old_machine_in_original_folder, "w")
                create_new_horizontal_program_for_old_machine_in_original_folder.write(
                    '\n'.join(HorizontalProgramLines))
                create_new_horizontal_program_for_old_machine_in_original_folder.close()
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=ffff00] ORIGINAL [/color][/b]' +
                    "Folder location to Save the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]'
                    + str(error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Original Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

        close_button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_window_of_original_folder,
                                      font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
                '[b][i][u][color=0099ff]' + self.ids["JobNumber"].text +
                '[/color][/u][/i][/b]' + '[color=ffffff] Program Has been [/color]' +
                '[b][color=ffffff]REPLACED[/color][/b]' + '[color=ffffff] Successfully in [/color]' +
                '[color=ffff00]ORIGINAL[/color]' + '[color=ffffff] Folder.[/color]'), size_hint=(0.7, 1.0),
                                                      buttons=[close_button], auto_dismiss=False)

        if (old_horizontal_program_confirmation_email_message_list != []):
            email_messages_of_creating_old_horizontal_machine_program.append(
                "\n".join(old_horizontal_program_confirmation_email_message_list) + "\n")
        success_messages_of_creating_old_horizontal_machine_program = [
            "\n" + "Program has been **REPLACED** successfully in **ORIGINAL** Folder." + "\n"]
        # maybe need logic to send this message to trello in case there is warnning make this program
        email_messages_of_creating_old_horizontal_machine_program.append("\n".join(
            success_messages_of_creating_old_horizontal_machine_program))

        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        print()
        # print('\n'.join(HorizontalProgramLines))

    def replace_existing_old_horizontal_machine_program_in_running_folder(self, obj):
        # NEED TO PUT THIS ONE HERE TO MAKE SURE YOU CLOSE ALL THE DIALOG WINDOW WERE OPEN
        self.Old_Horizontal_Message_Dialog.dismiss()
        print("replace_existing_old_horizontal_machine_program_in_running_folder" + " Function is called")
        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):
            # TO 0
            # JUST MAKE IT global HERE AGAIN TO CAN USE IT IN BELOW FUNCTIONS
            global new_horizontal_program_To0_direction_for_old_machine_in_running_folder
            new_horizontal_program_To0_direction_for_old_machine_in_running_folder = (
                    running_folder_path_of_old_horizontal_machine + "\\" + "P" +
                    new_program_number_for_old_horizontal_machine + "TO0.MIN")
            horizontal_program_lines_To0_direction_for_old_machine_in_running_folder[0] = (
                    '(PART ' + new_program_number_for_old_horizontal_machine + " TO 0" + ' -- '
                    + todaydate + ' <SYS>' + ')')
            try:
                create_new_horizontal_program_To0_direction_for_old_machine_in_running_folder = open(
                    new_horizontal_program_To0_direction_for_old_machine_in_running_folder, "w")
                create_new_horizontal_program_To0_direction_for_old_machine_in_running_folder.write(
                    '\n'.join(horizontal_program_lines_To0_direction_for_old_machine_in_running_folder))
                create_new_horizontal_program_To0_direction_for_old_machine_in_running_folder.close()
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=33cc33] RUNNING [/color][/b]' + "Folder location to Save the Program."
                    + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]' + str(
                        error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Running Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

            # TO 180
            # JUST MAKE IT global HERE AGAIN TO CAN USE IT IN BELOW FUNCTIONS
            global new_horizontal_program_To180_direction_for_old_machine_in_running_folder
            new_horizontal_program_To180_direction_for_old_machine_in_running_folder = (
                    running_folder_path_of_old_horizontal_machine + "\\" + "P" +
                    new_program_number_for_old_horizontal_machine + "TO180.MIN")
            horizontal_program_lines_To180_direction_for_old_machine_in_running_folder[0] = (
                    '(PART ' + new_program_number_for_old_horizontal_machine + " TO 180" + ' -- '
                    + todaydate + ' <SYS>' + ')')
            horizontal_program_lines_To180_direction_for_old_machine_in_running_folder[VC155_VARIABLE_INDEX] = (
                    'VC155=' + format((-1) * offset_amount) + '  (Offset)')
            horizontal_program_lines_To180_direction_for_old_machine_in_running_folder[VC173_VARIABLE_INDEX] = (
                    'VC173=' + format(float(Y_distance_from_origin_to_circlip_notch) - (2 * offset_amount),
                                      '.4f') + '  (yCirclipNotch)')
            try:
                create_new_horizontal_program_To180_direction_for_old_machine_in_running_folder = open(
                    new_horizontal_program_To180_direction_for_old_machine_in_running_folder, "w")
                create_new_horizontal_program_To180_direction_for_old_machine_in_running_folder.write(
                    '\n'.join(horizontal_program_lines_To180_direction_for_old_machine_in_running_folder))
                create_new_horizontal_program_To180_direction_for_old_machine_in_running_folder.close()
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=33cc33] RUNNING [/color][/b]' +
                    "Folder location to Save the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]'
                    + str(error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Running Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

        else:
            new_horizontal_program_for_old_machine_in_running_folder = (
                    running_folder_path_of_old_horizontal_machine + "\\" + "P" +
                    new_program_number_for_old_horizontal_machine + ".MIN")
            try:
                create_new_horizontal_program_for_old_machine_in_running_folder = open(
                    new_horizontal_program_for_old_machine_in_running_folder, "w")
                create_new_horizontal_program_for_old_machine_in_running_folder.write('\n'.join(HorizontalProgramLines))
                create_new_horizontal_program_for_old_machine_in_running_folder.close()
            except PermissionError or FileNotFoundError as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                fail_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find" + '[b][color=33cc33] RUNNING [/color][/b]' +
                    "Folder location to Save the Program." + "\n" + "An Error has occurred :" + "\n" + '[color=ff1a1a]'
                    + str(error) + '[/color]' + "\n" + "Double Check Network, and Folder Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "Failed to Find Running Folder location to Save the Program." + "\n" +
                    "Double Check Network, and File Location.")
                failed_to_create_old_horizontal_machine_program(self)
                return

        close_button = MDRaisedButton(text='Close', on_release=self.close_old_horizontal_window_of_running_folder,
                                      font_size=16)
        self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=248f24]Success Message[/color]', text=(
                '[b][i][u][color=0099ff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' +
                '[color=ffffff] Program Has been [/color]' + '[b][color=ffffff]REPLACED[/color][/b]' +
                '[color=ffffff] Successfully in [/color]' + '[color=33cc33]RUNNING[/color]' +
                '[color=ffffff] Folder.[/color]' + '\n' +
                '[color=ffffff]After closing this window, the program will open on CIMCO Editor.[/color]' +
                '\n' + '[color=ffffff]Please Double Check and Report any Problem on Trello.[/color]'),
                                                      size_hint=(0.7, 1.0), buttons=[close_button], auto_dismiss=False)

        success_messages_of_creating_old_horizontal_machine_program = [
            "\n" + "Program has been **REPLACED** successfully in **RUNNING** Folder." + "\n"]
        # maybe need logic to send this message to trello in case there is warnning make this program
        email_messages_of_creating_old_horizontal_machine_program.append(
            "\n".join(success_messages_of_creating_old_horizontal_machine_program))
        # TO OPEN THE DIALOG WINDOW
        self.Old_Horizontal_Message_Dialog.open()
        print()
        # print('\n'.join(HorizontalProgramLines))

    def close_old_horizontal_window_of_original_folder(self, obj):
        print("close_old_horizontal_window_of_original_folder" + " Function is called")

        global result_of_existing_programs
        # LIST TO STORE THE FOUND JOBS WITH WHOLE PATH INCLUDING FILE NAME
        result_of_existing_programs = []
        for file in glob.glob(original_folder_path_of_old_horizontal_machine + '\*' +
                              new_program_number_for_old_horizontal_machine + '*'):
            # TO APPEND(ADD) EACH PROGRAM THAT FOUND TO THE RESULT LIST
            result_of_existing_programs.append(file)
        print("result_of_existing_programs: ", result_of_existing_programs)

        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()

        if ((((original_folder_path_of_old_horizontal_machine + "\\" + "P" +
               new_program_number_for_old_horizontal_machine + "TO0.MIN") in result_of_existing_programs) or
             ((original_folder_path_of_old_horizontal_machine + "\\" + "P" +
               new_program_number_for_old_horizontal_machine + "TO180.MIN") in result_of_existing_programs)) and
                ((original_folder_path_of_old_horizontal_machine + "\\" + "P" +
                  new_program_number_for_old_horizontal_machine + ".MIN") in result_of_existing_programs)):
            print("job saved as To0 & To180")
            warning_messages_of_creating_old_horizontal_machine_program.append('[b][i][u][color=0099ff]' + self.ids[
                "JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] job saved in several ways :[/color]' +
                                                                               '\n' + '\n' + '[color=ffffff]' + (
                                                                                   '\n'.join(
                                                                                       result_of_existing_programs)) + '[/color]' +
                                                                               '\n' + '\n' + '[color=ffffff]Double check in [/color]' + '[color=ffff00]ORIGINAL[/color]' +
                                                                               '[color=ffffff] and [/color]' + '[color=33cc33]RUNNING[/color]' +
                                                                               '[color=ffffff] Folders , and[/color]' + '\n' +
                                                                               '[color=ffffff]DELETE the Wrong Program.[/color]')

            email_messages_of_creating_old_horizontal_machine_program.append("Job saved in several ways" + '\n' + (
                '\n'.join(result_of_existing_programs)) + '\n' + "Double check in ORIGINAL and RUNNING Folders" +
                                                                             '\n' + "and DELETE the Wrong Program.")
            old_horizontal_machine_program_needs_attention(self)
            return

        # # TO CLOSE THE DIALOGE
        # self.Old_Horizontal_Message_Dialog.dismiss()
        # just for now
        # if (confirmation_email_messages_of_creating_old_horizontal_machine_program != []):
        #     email_messages_of_creating_old_horizontal_machine_program.append(
        #     "\n".join(confirmation_email_messages_of_creating_old_horizontal_machine_program)+"\n")
        # success_messages_of_creating_old_horizontal_machine_program = [
        # "Program has been created successfully in ORIGINAL Folder." + "\n"]
        # # maybe need logic to send this message to trello in case there is warnning make this program
        # email_messages_of_creating_old_horizontal_machine_program.append("\n".join(
        # success_messages_of_creating_old_horizontal_machine_program))
        # TO CALL FUNCTION OF HORIZONTAL PROGRAM IN RUNNING FOLDER
        # if self.ids["JobNumber"].text != "":

        create_old_horizontal_machine_program_in_running_folder(self)
        # TO RESET JOB NUMBER FIELD TO START OVER
        # self.ids["JobNumber"].text = ""

        # ---------------------still work here--------------------

    def close_old_horizontal_window_of_running_folder(self, obj):
        print("close_old_horizontal_window_of_running_folder" + " Function is called")

        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()
        # TO OPEN FILE ON WINDOWS USE CODE BELOW
        # subprocess.Popen([ApplicationName, FileName])
        # (subprocess): BUILT IN FUNCTION USED TO OPEN FILES
        # (Popen): USED TO OPEN THE FILE
        # (ApplicationName): KIND OF APP THAT USED TO OPEN THE FILE (LIKE NOTEPAD,EXCEL,WORD...ETC),
        # CIMCO Editor IN OUR CASE
        # (FileName): PATH OF FILE YOU WANT TO OPEN
        if (offset_direction == "OFFSET EACH WAY" and offset_amount != 0):
            try:
                subprocess.Popen([cimco_editor_path,
                                  new_horizontal_program_To0_direction_for_old_machine_in_running_folder])
                subprocess.Popen([cimco_editor_path,
                                  (new_horizontal_program_To180_direction_for_old_machine_in_running_folder)])
            except Exception as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                warning_messages_of_creating_old_horizontal_machine_program.append(
                    "Program has been created successfully, but Failed to Open it by" +
                    '[b][u][color=ffffff] CIMCO Editor. [/color][/u][/b]' + "\n" + "An Error has occurred :" + "\n" +
                    '[color=ff1a1a]' + str(error) + '[/color]' + "\n" +
                    "Double Check Network, and CIMCO Editor Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "\n" + "But Failed to Open it by CIMCO Editor." + "\n" + "An Error has occurred :" + "\n" + str(
                        error) + "\n" + "Double Check Network, and CIMCO Editor Location.")
                old_horizontal_machine_program_needs_attention(self)
        else:
            try:
                subprocess.Popen([cimco_editor_path, new_horizontal_program_for_old_machine_in_running_folder])
            except Exception as error:
                # print("except LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION")
                warning_messages_of_creating_old_horizontal_machine_program.append(
                    "Program has been created successfully, but Failed to Open it by" +
                    '[b][u][color=ffffff] CIMCO Editor. [/color][/u][/b]' + "\n" + "An Error has occurred :" + "\n" +
                    '[color=ff1a1a]' + str(error) + '[/color]' + "\n" +
                    "Double Check Network, and CIMCO Editor Location.")
                email_messages_of_creating_old_horizontal_machine_program.append(
                    "\n" + "But Failed to Open it by CIMCO Editor." + "\n" + "An Error has occurred :" + "\n" + str(
                        error) + "\n" + "Double Check Network, and CIMCO Editor Location.")
                old_horizontal_machine_program_needs_attention(self)

        # if ((((NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO0.MIN")
        # in result_of_existing_programs) or
        #      ((NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + "TO180.MIN")
        #      in result_of_existing_programs)) and
        #         ((NEW_HORIZONTAL_PROGRAM_RUNNING_FOLDER_PATH + "\\" + "P" + NEW_PROGRAM_NUMBER + ".MIN")
        #         in result_of_existing_programs)):
        #     print("job saved as To0 & To180")
        #     warning_messages_of_creating_old_horizontal_machine_program.append('[b][i][u][color=0099ff]' +
        #     self.ids["JobNumber"].text + '[/color][/u][/i][/b]' + '[color=ffffff] job saved in several ways :[/color]'
        #     + '\n' + '\n' + '[color=ffffff]' + ('\n'.join(result_of_existing_programs)) + '[/color]' + '\n' + '\n' +
        #     '[color=ffffff]Double check in [/color]' + '[color=ffff00]ORIGINAL[/color]' +
        #     '[color=ffffff] and [/color]' +  '[color=33cc33]RUNNING[/color]' +
        #     '[color=ffffff] Folders , and[/color]' + '\n' + '[color=ffffff]DELETE the Wrong Program.[/color]')
        #
        #     horizontal_Program_needs_Attention(self)

        # Close_Button = MDRaisedButton(text='Close',on_release=self.Close_Old_Horizontal_Dialog,font_size = 16)
        # Close_Old_Horizontal_Dialog_      '[color=ffffff] Folder.[/color]'
        # self.Old_Horizontal_Message_Dialog = MDDialog(title='[color=cccc00]Warning Message[/color]', text=(
        #                         '[b][i][u][color=0099ff]' + self.ids["JobNumber"].text + '[/color][/u][/i][/b]' +
        #                         '[color=ffffff] job saved in several ways :[/color]'  + '\n' + '\n' +
        #                         '[color=ffffff]' + ('\n'.join(result_of_existing_programs)) + '[/color]' + '\n'
        #                         + '\n' +  '[color=ffffff]Double check in [/color]' + '[color=ffff00]ORIGINAL[/color]'
        #                         + '[color=ffffff] and [/color]' +  '[color=33cc33]RUNNING[/color]'
        #                         + '[color=ffffff] Folders , and[/color]' + '\n'
        #                         + '[color=ffffff]DELETE the Wrong Program.[/color]')
        #                          , size_hint=(0.7, 1.0), buttons=[Close_Button],auto_dismiss=False)
        # TO OPEN THE DIALOG WINDOW
        # self.Old_Horizontal_Message_Dialog.open()

        # maybe need to add logic of warnning and need manual stuff
        # success_messages_of_creating_old_horizontal_machine_program =
        # ["Program has been created successfully in RUNNING Folder."]
        # email_messages_of_creating_old_horizontal_machine_program.append(
        # "\n".join(success_messages_of_creating_old_horizontal_machine_program))

        send_email_about_create_old_horizontal_machine_program(self)

        # TO ADD MESSAGE ON TRELLO
        print("\n".join(email_messages_of_creating_old_horizontal_machine_program))

        # TO RESET JOB NUMBER FIELD TO START OVER
        self.ids["JobNumber"].text = ""
        # Reset_Variables
        four_cycle_pin_bore_variables()

    # def EnterOffsetValue(self,obj):
    #     global OFFSET_AMOUNT
    #     OFFSET_AMOUNT = self.Confirmation_MDTextField.text
    #     global verification_messages_of_creating_old_horizontal_machine_program
    #     verification_messages_of_creating_old_horizontal_machine_program = []
    #     print("offset value after ", OFFSET_AMOUNT)
    #     return OFFSET_AMOUNT

    def close_old_horizontal_screen_window(self, obj):
        print("close_old_horizontal_screen_window" + " Function is called")
        # TO CLOSE THE DIALOGE
        self.Old_Horizontal_Message_Dialog.dismiss()
        # TO ADD MESSAGE ON TRELLO
        if (self.ids["JobNumber"].text == ""):
            pass
        elif (fail_messages_of_creating_old_horizontal_machine_program != []):
            print("\n".join(email_messages_of_creating_old_horizontal_machine_program))
            send_email_about_create_old_horizontal_machine_program(self)
        elif (warning_messages_of_creating_old_horizontal_machine_program != []):
            print("\n".join(email_messages_of_creating_old_horizontal_machine_program))
            # to avoid crash the app if cant's open it by cimco
            try:
                subprocess.Popen([cimco_editor_path, new_horizontal_program_for_old_machine_in_original_folder])
            except BaseException:
                print("Can't open program by CIMCO")
            send_email_about_create_old_horizontal_machine_program(self)

        # TO RESET JOB NUMBER FIELD TO START OVER
        self.ids["JobNumber"].text = ""
        # Reset_Variables
        four_cycle_pin_bore_variables()
        # JUST FOR NOW
        # fail_messages_of_creating_old_horizontal_machine_program = []

    def reset_old_horizontal_screen_fields(self):
        # TO RESET JOB NUMBER FIELD TO START OVER
        self.ids["JobNumber"].text = ""

        # TO CALL FUNCTION TO START OVER
        # LOAD_HORIZONTAL_SHEETS_FOR_AUTOMATION(self)
        # self.manager.current = 'OldHorizontalScreen'
        # self.Old_Horizontal_Message_Dialog.dismiss()

    # endregion <<<<===========================[Old Horizontal Machines Sub Functions]===========================>>>>


# endregion <<<<===========================[Old Horizontal Machines(28,29,32) Screen]============================>>>>


# region <<<<================================[New Horizontal Machine(127) Screen]===============================>>>>
class NewHorizontalScreen(Screen):
    # self.ids["JobNumber"].text: TO ACCESS TEXT FIELD WE USE ids THAT'S DEFINE ABOVE
    def create_program_for_new_horizontal_machine(self):
        new_program_number_for_the_new_horizontal_machine = self.ids["JobNumber"].text
        print(new_program_number_for_the_new_horizontal_machine)


# endregion <<<<==============================[New Horizontal Machine(127) Screen]=============================>>>>


# region <<<<==========================================[Setting Screen]==========================================>>>>
class SettingScreen(Screen):
    # CHECK FOR ADMIN TO NOT ALLOWED ANY ONE ELSE CHANE APP SETTING
    def admin_check(self):
        admin = "malatweh@rwbteam.com"
        if (self.manager.get_screen('LoginScreen').ids["Email"].text == admin):
            self.manager.current = 'AppSettingScreen'
        else:
            # print("SORRY, YOU ARE NOT AUTHORIZED TO ACCESS THIS SCREEN")
            close_button = MDRaisedButton(text='Close', on_release=self.close_setting_screen_window, font_size=16)
            self.setting_screen_message_window = MDDialog(title='[color=990000]Warning Message[/color]', text=(
                '[color=ffffff]Sorry, You are NOT Authorized to access this screen.[/color]'),
                                                          size_hint=(0.7, 1.0), buttons=[close_button],
                                                          auto_dismiss=False)
            # TO OPEN THE DIALOG WINDOW
            self.setting_screen_message_window.open()

    def close_setting_screen_window(self, obj):
        self.setting_screen_message_window.dismiss()


# endregion <<<<========================================[Setting Screen]=========================================>>>>


# region <<<<=====================================[Pin Bore Setting Screen]======================================>>>>
class PinBoreSettingScreen(Screen):
    pass


# endregion <<<<===================================[Pin Bore Setting Screen]=====================================>>>>


# region <<<<==========================[Old Horizontal Machines(28,29,32) Setting Screen]========================>>>>
class OldHorizontalSettingScreen(Screen):
    pass


# endregion <<<<=======================[Old Horizontal Machines(28,29,32) Setting Screen]========================>>>>


# region <<<<====================================[Application Setting Screen]====================================>>>>
class AppSettingScreen(Screen):
    pass


# endregion <<<<===================================[Application Setting Screen]==================================>>>>


# region <<<<======================================[Add New User Screen]=======================================>>>>

class AddNewUserScreen(Screen):
    # needs work
    def add_new_user(self):
        try:
            print("New User Added Successfully, Make Sure You Send Him The Shared Password By Email")
            print("USER NAME:", self.ids["UserName"].text)
            print("RWB EMAIL:", self.ids["NewRWBEmail"].text)
            print("WISECO EMAIL:", self.ids["NewWisecoEmail"].text)
            # ADD NEW USER INFORMATION FOR EXCEL FILE(USER_NAME,WISECO_EMAIL, AND RWB_EMAIL)
            new_user_name = email_address_list_file['Email'].loc[self.ids["UserName"].text, 'Users'] = self.ids[
                "UserName"].text
            new_wiseco_email = email_address_list_file['Email'].loc[
                self.ids["NewWisecoEmail"].text, 'Wiseco Email Address'] = self.ids["NewWisecoEmail"].text
            new_rwb_email = email_address_list_file['Email'].loc[self.ids["NewRWBEmail"].text, 'RWB Email Address'] = \
                self.ids["NewRWBEmail"].text
            # CREATE DATA FRAME WITH THE NEW USER DATA TO ADD THEM TO THE EXCEL SHEET
            new_user_data = pd.DataFrame(
                data={'Users': [new_user_name], 'Wiseco Email Address': [new_wiseco_email],
                      'RWB Email Address': [new_rwb_email]})
            # LOAD THE EXCEL SHEET TO BE ABLE TO ADD NEW DATA
            user_information_workbook = openpyxl.load_workbook(email_address_list_file_path)
            # ACCESS THE EXCEL SHEET AND USE (mode= 'a') TO ADD THE NEW DATA
            user_information_sheet_update = pd.ExcelWriter(email_address_list_file_path, engine='openpyxl', mode='a')
            # SET (EMAIL_SHEET_UPDATE) AS CURRENT EXCEL BOOK (EXCEL FILE IN ANOTHER WORD)
            user_information_sheet_update.book = user_information_workbook
            # LOOP THROUGH THE EXCEL FILE TO SCAN ALL THE SHEETS
            # (MUST PUT THIS LINE OF CODE WHEN WE USE THIS MODE (mode= 'a') )
            user_information_sheet_update.sheets = dict((ws.title, ws) for ws in user_information_workbook.worksheets)
            # print("EMAIL_SHEET_UPDATE.sheets:", EMAIL_SHEET_UPDATE.sheets)
            # UPDATE THE EMAIL SHEET WITH ADDING THE NEW USER DATA
            new_user_data.to_excel(user_information_sheet_update, sheet_name='Email',
                                   startrow=user_information_workbook['Email'].max_row,
                                   startcol=0, header=False, index=False)
            # SAVE CHANGES
            user_information_sheet_update.save()
            # CLOSE SHEET
            user_information_sheet_update.close()

            # SHOW MESSAGE OF New User Added Successfully
            close_button = MDRaisedButton(text='Close', on_release=self.close_new_user_screen_window, font_size=16)
            self.new_user_screen_message_window = MDDialog(title='', text=(
                "New User Added Successfully, Make Sure You Send to Him The Shared Password By Email"),
                                                           size_hint=(0.7, 1.0), buttons=[close_button],
                                                           auto_dismiss=False)
            # TO OPEN THE DIALOG WINDOW
            self.new_user_screen_message_window.open()

        # ADD USER WORK FINE ONLY FOR ADDING ONE USER ,
        # THE PROBLEM IT WILL CORRUPT THE EXCEL WHICH MAKE IT INACCESSIBLE FILE, WE NEED TO FIGURE OUT THAT LATER***
        except (PermissionError):
            # SHOW MESSAGE OF New User Added Successfully
            close_button = MDRaisedButton(text='Close', on_release=self.close_new_user_screen_window, font_size=16)
            self.new_user_screen_message_window = MDDialog(title='', text=(
                "PermissionError: Something Get Wrong to Access the File, Email malatweh@rwbteam.com "),
                                                           size_hint=(0.7, 1.0), buttons=[close_button],
                                                           auto_dismiss=False)
            # TO OPEN THE DIALOG WINDOW
            self.new_user_screen_message_window.open()

    def close_new_user_screen_window(self, obj):
        self.new_user_screen_message_window.dismiss()
        # TO RESET ADD USER FIELDS TO START OVER
        self.ids["UserName"].text = ""
        self.ids["NewRWBEmail"].text = ""
        self.ids["NewWisecoEmail"].text = ""

    def reset_new_user_screen_fields(self):
        # TO RESET ADD USER FIELDS TO START OVER
        self.ids["UserName"].text = ""
        self.ids["NewRWBEmail"].text = ""
        self.ids["NewWisecoEmail"].text = ""


# endregion <<<<====================================[Add New User Screen]=====================================>>>>


# region <<<<==========================================[Screen Manager]==========================================>>>>

# Create the screen manager
sm = ScreenManager()
sm.add_widget(LoginScreen(name='LoginScreen'))
sm.add_widget(HomeScreen(name='HomeScreen'))
sm.add_widget(PinBoreScreen(name='PinBoreScreen'))
sm.add_widget(OldHorizontalScreen(name='OldHorizontalScreen'))
sm.add_widget(NewHorizontalScreen(name='NewHorizontalScreen'))
sm.add_widget(SettingScreen(name='SettingScreen'))
sm.add_widget(PinBoreSettingScreen(name='PinBoreSettingScreen'))
sm.add_widget(OldHorizontalSettingScreen(name='OldHorizontalSettingScreen'))
sm.add_widget(AppSettingScreen(name='AppSettingScreen'))
sm.add_widget(AddNewUserScreen(name='AddNewUserScreen'))


# endregion <<<<=========================================[Screen Manager]========================================>>>>


# region <<<<=======================================[Application Builder]========================================>>>>

class WisecoProgramsMaker(MDApp):

    def build(self):
        # TO CONTROL SIZE OF THE SCREEN (Window.size = (WIDTH, HEIGHT))
        Window.size = (900, 650)
        # TO CHOOSE BACKGROUND MODE OF APP WHETHER DARK OR LIGHT
        self.theme_cls.theme_style = "Dark"
        # TO SET DEFAULT COLOR OF APP ELEMENTS(LABELS,BUTTONS...ETC)
        self.theme_cls.primary_palette = "Red"
        # TO SET DEFAULT COLOR CONCENTRATION(DARKNESS AND BRIGHTNESS) OF APP ELEMENTS(LABELS,BUTTONS...ETC)
        self.theme_cls.primary_hue = "900"
        # LOAD (builder_screen) TO USE IT IN THE APP
        builder_screen = Builder.load_string(Screens_Builder)
        # LOAD (BuilderDialog) TO USE IT IN THE APP
        # BuilderDialog = Builder.load_string(Dialog_Builder)

        # TO DEFINE (Screen() THAT USED TO DISPLAY THE APP) AS (app_screen) TO USE LATER
        app_screen = Screen()

        # BoxLayout FOR ENTIRE APP INCLUDE ALL WIDGETS AND ELEMENTS, SHOULD ADD ALL APP COMPONENTS FOR THIS BOX LAYOUT.
        # (orientation='vertical') TO ORGANIZE APP ELEMENTS VERTICALLY,
        # (spacing=20) TO MAKE SPACE BETWEEN APP ELEMENTS,
        # (padding=15) TO MAKE SPACE BETWEEN WALL BORDERS AND APP ELEMENTS,
        # (md_bg_color= [32/255.0, 32/255.0, 32/255.0, 1]) TO CHANGE THE COLOR BY ADJUSTING RGB VALUE
        # (CHECK: https://www.w3schools.com/colors/colors_picker.asp?colorhex=edfeff)
        app_box_layout = MDBoxLayout(orientation='vertical', spacing=20, padding=15,
                                     md_bg_color=[32 / 255.0, 32 / 255.0, 32 / 255.0, 1])

        # TO ADD PICTURE FOR THE APP FROM WEBSITE
        app_image = AsyncImage(source='https://www.wiseco.com/Images/Downloads/Wiseco_Black_CMYK.gif', size_hint_y=None,
                               height=70, allow_stretch=True, pos_hint={'center_x': 0.5, 'center_y': 0.10},
                               color=[150 / 255.0, 0 / 255.0, 0 / 255.0, 1])
        # TO ADD app_image TO app_box_layout TO DISPLAY IT IN THE APP SCREEN
        app_box_layout.add_widget(app_image)

        # TO ADD Screens_Builder THAT'S CREATE ABOVE
        app_box_layout.add_widget(builder_screen)
        # TO ADD Dialog_Builder THAT'S CREATE ABOVE
        # app_box_layout.add_widget(BuilderDialog)

        # ADD app_box_layout THAT CONTAIN ALL ELEMENTS AND WIDGETS OF THE APP TO app_screen
        # TO DISPLAY IT IN THE APP SCREEN.
        app_screen.add_widget(app_box_layout)
        return app_screen


WisecoProgramsMaker().run()

# endregion <<<<=====================================[Application Builder]=======================================>>>>
